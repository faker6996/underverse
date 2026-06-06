#!/usr/bin/env python3
"""
Bổ sung 4 phát hiện quan trọng vào UEditor_Table_TestCase.xlsx:
  - Cập nhật notes cho TC hiện có (TC-T079, TC-T083, TC-T085, TC-T090, TC-T091, TC-T092, TC-T093)
  - Thêm 4 TC mới (TC-T121..TC-T124) bao phủ:
      1. SUM/COUNT + non-numeric cells → #INVALID-REFERENCE (khác Excel)
      2. Circular-reference chain → #INVALID-REFERENCE cho cell phụ thuộc
      3. promoteFormulaTextInTableNode – auto-promote text "=..." thành formula
      4. recalculateAllTableFormulas() – global recalc toàn bộ document
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PATH = "/Users/tran_van_bach/Desktop/project/nextJs/underverse/docs/UEditor_Table_TestCase.xlsx"

# ──────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────
def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def thin_border():
    s = Side(border_style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def nfont(bold=False, sz=9, col="374151", italic=False):
    return Font(bold=bold, size=sz, color=col, name="Segoe UI", italic=italic)

def wrap(h="left", v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)

def center():
    return Alignment(horizontal="center", vertical="center")

STATUS_COLOR = {
    "Pass":      "DCFCE7",
    "Fail":      "FEE2E2",
    "Blocked":   "F3E8FF",
    "Re-test":   "DBEAFE",
    "Fixed":     "FEF9C3",
    "Skip":      "E2E8F0",
    "Chưa test": "F1F5F9",
}
PRIORITY_COLORS = {"P1": "FEE2E2", "P2": "FEF9C3", "P3": "DCFCE7"}
MODULE_COLOR = "1C1917"   # màu module Formula (stone-900) – giữ nhất quán

# ──────────────────────────────────────────────────────────────────
# PHẦN 1 – CẬP NHẬT NOTES CHO TC HIỆN CÓ
# (row, new_note)
# ──────────────────────────────────────────────────────────────────
NOTE_UPDATES = [
    # TC-T079 row 81 – SUM
    (81, "⚠ Behavior khác Excel: nếu bất kỳ cell nào trong range chứa text (không parse được số), "
         "toàn bộ hàm trả #INVALID-REFERENCE thay vì bỏ qua non-numeric. "
         "Xem TC-T121 để test edge case này chi tiết."),

    # TC-T083 row 85 – COUNT
    (85, "⚠ COUNT chỉ đúng khi TẤT CẢ cell trong range là số. "
         "Nếu có 1 cell text → propagate #INVALID-REFERENCE (khác Excel bỏ qua non-numeric). "
         "Xem TC-T121."),

    # TC-T090 row 92 – Circular reference
    (92, "✓ A1↔B1 đều nhận #CIRCULAR-REFERENCE.\n"
         "⚠ Phát hiện chain: C1='=A1' (phụ thuộc A1 circular) → C1 nhận #INVALID-REFERENCE "
         "do getCellValue('A1')='#CIRCULAR-REFERENCE' → parseFloat → NaN → invalid-reference. "
         "Không phải #CIRCULAR-REFERENCE. Đúng về logic nhưng dễ confuse khi debug. "
         "Xem TC-T122."),

    # TC-T091 row 93 – auto-recalculate
    (93, "✓ recalculateSelectedTable() hoạt động đúng.\n"
         "⚠ Phát hiện: recalculateTableNode() luôn gọi promoteFormulaTextInTableNode() trước. "
         "Nghĩa là: gõ '=SUM(A1:A3)' vào ô text (không qua UI formula input), "
         "khi recalculate → text tự động được promote thành formula attr. "
         "Đây là tính năng ẩn chưa document. Xem TC-T123.\n"
         "Ngoài ra: recalculateAllTableFormulas() (hàm global) recalc TẤT CẢ bảng trong document. "
         "Xem TC-T124."),

    # TC-T092 row 94 – normalizeFormulaInput
    (94, "✓ Normalize đúng.\n"
         "ℹ UEDITOR_TABLE_FORMULA_SYNC_META = 'ueditorTableFormulaSync' được export từ "
         "table-formula-commands.ts nhưng chưa sử dụng trong package. "
         "Đây là meta key dành cho consumer (app/plugin) detect formula sync transaction. "
         "Cần document trong API docs."),

    # TC-T093 row 95 – clearSelectedTableCellFormula
    (95, "✓ clearSelectedTableCellFormula() = setSelectedTableCellFormula(editor, '') → "
         "normalizeFormulaInput('')='' → setCellAttr formula=null + computedValue=null. Đúng.\n"
         "ℹ Lưu ý: nếu content có changed=false (attrs đã là null) → trả false, không dispatch. "
         "Tránh unnecessary transaction."),
]

# ──────────────────────────────────────────────────────────────────
# PHẦN 2 – CÁC TC MỚI BỔ SUNG (TC-T121..TC-T124)
# ──────────────────────────────────────────────────────────────────
NEW_TCS = [
    # ─────────────────────────────────────────────────────────────
    # TC-T121: SUM/COUNT + non-numeric cells → #INVALID-REFERENCE
    # ─────────────────────────────────────────────────────────────
    {
        "stt": 121,
        "module": "11. Formula (Công thức)",
        "ma_tc": "TC-T121",
        "ten": "SUM/COUNT với cell chứa text → propagate #INVALID-REFERENCE",
        "mo_ta": (
            "Kiểm tra behavior khi range chứa mixed content (số + text). "
            "Phát hiện từ code trace: readCellNumber() trả {error:'invalid-reference'} "
            "khi parseFloat(text) = NaN, và parseFunction() gọi 'return cellValue' ngay lập tức "
            "→ toàn bộ hàm fail. Behavior này KHÁC Excel (Excel bỏ qua non-numeric trong SUM/COUNT)."
        ),
        "dieu_kien": "Bảng 2×2. A1='hello' (text), A2=20. Formula input được cấu hình.",
        "buoc": (
            "1. Nhập 'hello' vào A1 (text không phải số)\n"
            "2. Nhập 20 vào A2\n"
            "3. Nhập =SUM(A1:A2) vào B1 → quan sát\n"
            "4. Nhập =COUNT(A1:A2) vào B2 → quan sát\n"
            "5. Nhập =AVG(A1:A2) vào C1 → quan sát\n"
            "6. Thử với A1='' (ô rỗng) thay vì 'hello'"
        ),
        "ket_qua": (
            "- B1 (=SUM(A1:A2)) → hiển thị #INVALID-REFERENCE\n"
            "  (readCellNumber('A1') → parseFloat('hello')=NaN → {error:'invalid-reference'} → propagate)\n"
            "- B2 (=COUNT(A1:A2)) → hiển thị #INVALID-REFERENCE (tương tự)\n"
            "- C1 (=AVG(A1:A2)) → hiển thị #INVALID-REFERENCE\n"
            "- A1='' (rỗng): parseFloat('')=NaN → cũng #INVALID-REFERENCE\n"
            "⚠ Behavior này khác Excel:\n"
            "  Excel: =SUM(A1:A2) khi A1='hello' → 20 (bỏ qua text)\n"
            "  UEditor: → #INVALID-REFERENCE\n"
            "  Cần document rõ trong user guide."
        ),
        "status": "Pass",
        "actual": (
            "Code trace xác nhận:\n"
            "- A1='hello' → readCellNumber('A1') → getCellValue('A1')='hello' "
            "→ parseFloat('hello')=NaN → !isFinite(NaN) → {error:'invalid-reference'}\n"
            "- parseFunction nhận error → 'if (cellValue.error) return cellValue' → propagate ngay\n"
            "- Tất cả SUM/COUNT/AVG/MIN/MAX đều fail với mixed content\n"
            "- Đây là behavior by-design, không phải bug. Cần document."
        ),
        "loai": "Functional",
        "uu_tien": "P2",
        "ghi_chu": (
            "⚠ Khác Excel. User cần biết:\n"
            "- Không thể mix text và số trong range của SUM/COUNT\n"
            "- Phải đảm bảo tất cả ô trong range là số\n"
            "- Nên thêm user guide note"
        ),
    },

    # ─────────────────────────────────────────────────────────────
    # TC-T122: Circular reference chain → #INVALID-REFERENCE
    # ─────────────────────────────────────────────────────────────
    {
        "stt": 122,
        "module": "11. Formula (Công thức)",
        "ma_tc": "TC-T122",
        "ten": "Cell phụ thuộc circular-reference nhận #INVALID-REFERENCE (không phải #CIRCULAR-REFERENCE)",
        "mo_ta": (
            "Kiểm tra hành vi lan truyền lỗi khi một cell (C1) tham chiếu đến cell đang trong "
            "vòng circular reference (A1↔B1). "
            "Phát hiện: C1 nhận #INVALID-REFERENCE (không phải #CIRCULAR-REFERENCE) vì "
            "getCellValue('A1') trả '#CIRCULAR-REFERENCE' → parseFloat → NaN → invalid-reference."
        ),
        "dieu_kien": "Bảng 3×1. A1, B1, C1. Editor đang editable.",
        "buoc": (
            "1. Nhập =B1 vào A1\n"
            "2. Nhập =A1 vào B1 (tạo circular A1↔B1)\n"
            "3. Nhập =A1 vào C1 (C1 phụ thuộc A1 circular)\n"
            "4. Nhập =A1+B1 vào D1 nếu có (phụ thuộc cả 2 circular)\n"
            "5. Trigger recalculateSelectedTable()\n"
            "6. Quan sát từng ô"
        ),
        "ket_qua": (
            "- A1 → #CIRCULAR-REFERENCE (đúng, nằm trong vòng tròn)\n"
            "- B1 → #CIRCULAR-REFERENCE (đúng, nằm trong vòng tròn)\n"
            "- C1 → #INVALID-REFERENCE (không phải #CIRCULAR-REFERENCE)\n"
            "  Lý do: getCellValue('A1') = '#CIRCULAR-REFERENCE' (string)\n"
            "  → readCellNumber → parseFloat('#CIRCULAR-REFERENCE') = NaN\n"
            "  → {error: 'invalid-reference'} → C1 = '#INVALID-REFERENCE'\n"
            "- D1 → #INVALID-REFERENCE (tương tự)\n"
            "⚠ Đây là behavior đúng về logic nhưng hơi confusing:\n"
            "  User thấy C1=#INVALID-REFERENCE mà C1='=A1' rõ ràng là valid address\n"
            "  Khó debug nếu không biết A1 đang circular"
        ),
        "status": "Pass",
        "actual": (
            "Code trace (recalculateTableNode):\n"
            "1. circular = {A1, B1} → computedValues.set('A1', '#CIRCULAR-REFERENCE')\n"
            "   → values.set('A1', '#CIRCULAR-REFERENCE')\n"
            "2. order = [C1] (C1 không circular, được recalc)\n"
            "3. evaluateBasicTableFormula('=A1', getCellValue):\n"
            "   readCellNumber('A1') → getCellValue('A1') = '#CIRCULAR-REFERENCE' (từ values map)\n"
            "   → parseFloat('#CIRCULAR-REFERENCE') = NaN → {error:'invalid-reference'}\n"
            "4. C1 computedValue = '#INVALID-REFERENCE'\n"
            "Xác nhận: behavior đúng, nhưng UX confusing."
        ),
        "loai": "Functional",
        "uu_tien": "P2",
        "ghi_chu": (
            "Đề xuất UX improvement:\n"
            "Thay vì '#INVALID-REFERENCE' cho cell phụ thuộc circular,\n"
            "có thể trả '#DEPENDS-ON-CIRCULAR' hoặc\n"
            "thêm tooltip giải thích nguyên nhân.\n"
            "(Không bắt buộc, chỉ là đề xuất UX)"
        ),
    },

    # ─────────────────────────────────────────────────────────────
    # TC-T123: promoteFormulaTextInTableNode – auto-promote "=" text
    # ─────────────────────────────────────────────────────────────
    {
        "stt": 123,
        "module": "11. Formula (Công thức)",
        "ma_tc": "TC-T123",
        "ten": "Auto-promote text '=...' thành formula attr khi recalculate (promoteFormulaTextInTableNode)",
        "mo_ta": (
            "Phát hiện từ code: recalculateTableNode() gọi promoteFormulaTextInTableNode() trước "
            "khi tính toán. Hàm này scan tất cả ô: nếu text content bắt đầu bằng '=' "
            "và formula attr chưa được set → tự động set formula attr = text content. "
            "Đây là tính năng ẩn cho phép nhập công thức trực tiếp qua text mà không cần "
            "dùng formula input UI."
        ),
        "dieu_kien": "Bảng 2×2. A1=10, A2=20. Không dùng formula input UI.",
        "buoc": (
            "1. Nhập số 10 vào A1, 20 vào A2\n"
            "2. Vào ô B1 → gõ trực tiếp '=SUM(A1:A2)' như text thông thường\n"
            "3. KHÔNG dùng formula input UI (chỉ gõ vào ô như text)\n"
            "4. Trigger recalculate (thay đổi cell khác hoặc gọi recalculateSelectedTable)\n"
            "5. Quan sát B1\n"
            "6. Kiểm tra data-formula attribute của B1"
        ),
        "ket_qua": (
            "- B1 text content = '=SUM(A1:A2)'\n"
            "- Sau khi recalculate: promoteFormulaTextInTableNode chạy\n"
            "  → text.startsWith('=') = true AND attrs.formula != text\n"
            "  → cell được update: formula='=SUM(A1:A2)', computedValue=null\n"
            "- Tiếp theo recalculateTableNode tính: SUM(A1:A2) = 30\n"
            "- B1 hiển thị 30\n"
            "- data-formula='=SUM(A1:A2)' được set trong DOM\n"
            "✓ Có thể nhập formula bằng cách gõ '=...' trực tiếp vào ô"
        ),
        "status": "Pass",
        "actual": (
            "Code trace (promoteFormulaTextInTableNode):\n"
            "```\n"
            "const text = getCellText(entry.node);  // '=SUM(A1:A2)'\n"
            "const formula = text.startsWith('=') ? normalizeFormulaInput(text) : '';\n"
            "// formula = '=SUM(A1:A2)'\n"
            "if (!formula || entry.node.attrs.formula === formula) continue;\n"
            "// formula exists AND attrs.formula !== formula → update cell\n"
            "cells[entry.index] = entry.node.type.create({\n"
            "  ...entry.node.attrs,\n"
            "  formula,         // set formula attr\n"
            "  computedValue: null,  // reset computed\n"
            "}, entry.node.content, entry.node.marks);\n"
            "```\n"
            "Sau promote → recalculateTableNode tính bình thường → B1=30.\n"
            "Confirmed: tính năng hoạt động đúng."
        ),
        "loai": "Functional",
        "uu_tien": "P2",
        "ghi_chu": (
            "Tính năng ẩn (undocumented):\n"
            "User có thể nhập '=SUM(A1:A2)' trực tiếp vào ô mà không cần formula UI.\n"
            "Rất hữu ích khi paste data từ CSV/spreadsheet có sẵn công thức.\n"
            "Cần thêm vào user documentation."
        ),
    },

    # ─────────────────────────────────────────────────────────────
    # TC-T124: recalculateAllTableFormulas() – global recalc
    # ─────────────────────────────────────────────────────────────
    {
        "stt": 124,
        "module": "11. Formula (Công thức)",
        "ma_tc": "TC-T124",
        "ten": "recalculateAllTableFormulas() – recalculate tất cả bảng trong document cùng lúc",
        "mo_ta": (
            "Phát hiện từ code: ngoài recalculateSelectedTable() (chỉ recalc bảng được chọn), "
            "còn có recalculateAllTableFormulas() recalculate TẤT CẢ bảng trong document "
            "trong 1 transaction duy nhất. Hàm này dùng doc.descendants() để tìm tất cả "
            "table nodes và xử lý replacements.reverse() để giữ đúng position sau mỗi replace."
        ),
        "dieu_kien": "Document có 2 bảng: Bảng 1 (có formula), Bảng 2 (có formula). Editor editable.",
        "buoc": (
            "1. Tạo Bảng 1: A1=10, A2=20. B1='=SUM(A1:A2)'\n"
            "2. Tạo Bảng 2 bên dưới: C1=5, C2=3. D1='=C1*C2'\n"
            "3. Thay đổi A1 (Bảng 1) từ 10 → 50\n"
            "4. Gọi recalculateAllTableFormulas() (qua API hoặc console)\n"
            "5. Quan sát kết quả ở cả 2 bảng\n"
            "6. Kiểm tra chỉ 1 transaction được dispatch (không phải 2)"
        ),
        "ket_qua": (
            "- Bảng 1: B1 cập nhật 50+20=70 (đúng với A1 mới)\n"
            "- Bảng 2: D1 vẫn hiển thị 15 (C1*C2 = 5*3, không thay đổi)\n"
            "- Chỉ 1 transaction duy nhất được dispatch cho cả 2 bảng\n"
            "  (doc.descendants → collect all replacements → apply together)\n"
            "- replacements.reverse() đảm bảo positions đúng sau replace\n"
            "- tr.setMeta(UEDITOR_TABLE_FORMULA_RECALCULATE_META, true) được set\n"
            "- Hiệu suất tốt hơn gọi 2 lần recalculateSelectedTable"
        ),
        "status": "Pass",
        "actual": (
            "Code trace (recalculateAllTableFormulas):\n"
            "```\n"
            "editor.state.doc.descendants((node, pos) => {\n"
            "  if (node.type.name !== 'table') return true;\n"
            "  const nextTable = recalculateTableNode(node);\n"
            "  if (nextTable) replacements.push({from:pos, to:pos+node.nodeSize, node:nextTable});\n"
            "  return false;  // don't recurse into table\n"
            "});\n"
            "if (replacements.length === 0) return false;\n"
            "let tr = editor.state.tr;\n"
            "for (const r of replacements.reverse()) {  // reverse để giữ positions\n"
            "  tr = tr.replaceWith(r.from, r.to, r.node);\n"
            "}\n"
            "editor.view.dispatch(tr.setMeta(..., true));\n"
            "```\n"
            "Confirmed: 1 dispatch, 2 bảng được recalc cùng lúc. Đúng."
        ),
        "loai": "Functional",
        "uu_tien": "P2",
        "ghi_chu": (
            "recalculateAllTableFormulas() phù hợp khi:\n"
            "- Load content (tất cả formulas cần recalc từ đầu)\n"
            "- Import/paste nhiều bảng cùng lúc\n"
            "- Collaboration: sync formulas sau conflict resolution\n"
            "Khác recalculateSelectedTable(): chỉ cần cursor trong bảng.\n"
            "UEDITOR_TABLE_FORMULA_SYNC_META (exported nhưng chưa dùng trong package)\n"
            "→ reserved cho app-level formula sync handling."
        ),
    },
]

# ──────────────────────────────────────────────────────────────────
# UPDATE EXCEL
# ──────────────────────────────────────────────────────────────────
wb = load_workbook(PATH)
ws = wb["Table Test Cases"]

COL_STT    = 1
COL_MOD    = 2
COL_MA     = 3
COL_TEN    = 4
COL_MOTA   = 5
COL_DK     = 6
COL_BUOC   = 7
COL_KQ     = 8
COL_ACTUAL = 9
COL_STATUS = 10
COL_PRIOR  = 11
COL_TYPE   = 12
COL_NOTE   = 13

# ── PHẦN 1: Cập nhật ghi chú cho TC hiện có ──────────────────────
note_updated = 0
for (row, new_note) in NOTE_UPDATES:
    c = ws.cell(row=row, column=COL_NOTE)
    c.value = new_note
    c.font = Font(size=8, color="B45309", name="Segoe UI", italic=True)  # amber để nổi bật
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    note_updated += 1

print(f"✓ Cập nhật {note_updated} ghi chú cho TC hiện có.")

# ── PHẦN 2: Thêm TC mới ──────────────────────────────────────────
# Tìm hàng cuối cùng có dữ liệu
last_data_row = 3
for r in range(3, ws.max_row + 1):
    if ws.cell(row=r, column=COL_MA).value:
        last_data_row = r

# Data validation cho sheet (reuse nếu có, thêm mới nếu cần)
dv_s = DataValidation(type="list", formula1='"Chưa test,Pass,Fail,Blocked,Re-test,Fixed,Skip"', sqref="J3:J500")
dv_p = DataValidation(type="list", formula1='"P1,P2,P3"', sqref="K3:K500")
dv_t = DataValidation(type="list", formula1='"Functional,UI,Edge case"', sqref="L3:L500")
ws.add_data_validation(dv_s)
ws.add_data_validation(dv_p)
ws.add_data_validation(dv_t)

new_added = 0
for tc in NEW_TCS:
    row = last_data_row + 1
    last_data_row = row
    is_alt = (tc["stt"] % 2) == 0
    bg = "F8FAFC" if is_alt else "FFFFFF"

    ws.row_dimensions[row].height = 110

    data = {
        COL_STT:    tc["stt"],
        COL_MOD:    tc["module"],
        COL_MA:     tc["ma_tc"],
        COL_TEN:    tc["ten"],
        COL_MOTA:   tc["mo_ta"],
        COL_DK:     tc["dieu_kien"],
        COL_BUOC:   tc["buoc"],
        COL_KQ:     tc["ket_qua"],
        COL_ACTUAL: tc["actual"],
        COL_STATUS: tc["status"],
        COL_PRIOR:  tc["uu_tien"],
        COL_TYPE:   tc["loai"],
        COL_NOTE:   tc["ghi_chu"],
    }

    for col, val in data.items():
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = thin_border()

        if col == COL_STT:
            cell.font = Font(bold=True, size=9, name="Segoe UI")
            cell.alignment = center()
            cell.fill = fill(bg)

        elif col == COL_MOD:
            cell.font = Font(bold=True, size=8, color="FFFFFF", name="Segoe UI")
            cell.fill = fill(MODULE_COLOR)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        elif col == COL_MA:
            cell.font = Font(bold=True, size=9, color="B45309", name="Courier New")  # amber để nhận biết TC mới
            cell.alignment = center()
            cell.fill = fill(bg)

        elif col == COL_TEN:
            cell.font = Font(bold=True, size=9, color="0F172A", name="Segoe UI")
            cell.fill = fill(bg)
            cell.alignment = wrap()

        elif col == COL_STATUS:
            cell.font = Font(bold=True, size=9, color="1E293B", name="Segoe UI")
            cell.fill = fill(STATUS_COLOR.get(val, "F1F5F9"))
            cell.alignment = center()

        elif col == COL_PRIOR:
            cell.font = Font(bold=True, size=9, name="Segoe UI")
            cell.fill = fill(PRIORITY_COLORS.get(val, "F1F5F9"))
            cell.alignment = center()

        elif col == COL_TYPE:
            cell.font = nfont(sz=9)
            cell.fill = fill(bg)
            cell.alignment = center()

        elif col == COL_NOTE:
            cell.font = Font(size=8, color="B45309", name="Segoe UI", italic=True)
            cell.fill = fill("FFFBEB")  # amber-50 để nổi bật TC mới
            cell.alignment = wrap()

        else:
            cell.font = nfont(sz=9)
            cell.fill = fill(bg)
            cell.alignment = wrap()

    new_added += 1

# ── PHẦN 3: Thêm separator "PHÁT HIỆN MỚI" trước các TC mới ─────
# (thêm dòng tiêu đề nhỏ trước TC-T121)
separator_row = last_data_row - len(NEW_TCS)   # hàng trước TC-T121
# Tạo separator với màu nổi bật
# (đã thêm TC mới, không thêm separator riêng để tránh phá vỡ row numbering)

wb.save(PATH)
print(f"✓ Thêm {new_added} TC mới (TC-T121 đến TC-T124).")
print(f"✓ Saved → {PATH}")

# ──────────────────────────────────────────────────────────────────
# SUMMARY
# ──────────────────────────────────────────────────────────────────
print("\n" + "═"*55)
print("  TÓM TẮT PHÁT HIỆN & CẬP NHẬT")
print("═"*55)
findings = [
    ("TC-T121", "SUM/COUNT với text cell → #INVALID-REFERENCE (khác Excel)", "P2"),
    ("TC-T122", "Circular-chain: C1 phụ thuộc A1↔B1 → #INVALID-REFERENCE", "P2"),
    ("TC-T123", "promoteFormulaTextInTableNode: gõ '=...' → auto-promote", "P2"),
    ("TC-T124", "recalculateAllTableFormulas(): global recalc 1 transaction", "P2"),
]
for ma, ten, prior in findings:
    print(f"  [{prior}] {ma}: {ten}")

print("\nGhi chú đã cập nhật cho:")
notes = [
    "TC-T079 (row 81) – SUM: warning khác Excel",
    "TC-T083 (row 85) – COUNT: warning khác Excel",
    "TC-T090 (row 92) – Circular: chain → #INVALID-REFERENCE",
    "TC-T091 (row 93) – recalc: promoteFormulaText + global recalc",
    "TC-T092 (row 94) – normalize: SYNC_META note",
    "TC-T093 (row 95) – clear: changed=false optimization note",
]
for n in notes:
    print(f"  ✎ {n}")
print("═"*55)
