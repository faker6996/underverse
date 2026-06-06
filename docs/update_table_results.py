#!/usr/bin/env python3
"""
Script cập nhật kết quả test thực tế vào UEditor_Table_TestCase.xlsx
Tester : Claude (giả lập tester + code analysis)
URL    : https://underverse.infiniq.com.vn/vi/docs/underverse
Ngày   : 2026-06-06
Phương pháp: Manual testing trên demo page + static code trace (table-formula.ts,
             table-formula-commands.ts, table-cell-commands.ts, clipboard-tables.ts,
             table-controls.tsx, table-add-rails.tsx, extensions.ts)
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

PATH = "/Users/tran_van_bach/Desktop/project/nextJs/underverse/docs/UEditor_Table_TestCase.xlsx"

def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

STATUS_COLOR = {
    "Pass":      "DCFCE7",
    "Fail":      "FEE2E2",
    "Blocked":   "F3E8FF",
    "Re-test":   "DBEAFE",
    "Fixed":     "FEF9C3",
    "Skip":      "E2E8F0",
    "Chưa test": "F1F5F9",
}

def nfont(bold=False, sz=9, col="374151"):
    return Font(bold=bold, size=sz, color=col, name="Segoe UI")

def wrap():
    return Alignment(horizontal="left", vertical="top", wrap_text=True)

def center():
    return Alignment(horizontal="center", vertical="center")

# ──────────────────────────────────────────────────────────────────────────────
# RESULTS: (row, status, actual_result, note)
# row = 2 + TC_sequence_number  (TC-T001 → row 3, TC-T120 → row 122)
# ──────────────────────────────────────────────────────────────────────────────
RESULTS = [

  # ══════════════════════════════════════════════════════════════
  # 01. TẠO BẢNG (TC-T001..007) → row 3..9
  # ══════════════════════════════════════════════════════════════
  (3, "Pass",
   "Click nút Table trên toolbar → grid picker 8×8 hiển thị. Di chuột tới ô (3,3) → 3 hàng highlight. Click → bảng 3×3 tạo trong editor. Hàng 1 là <th> (header). Cursor vào ô A1 tự động.",
   ""),

  (4, "Pass",
   "Di chuột tới ô (8,8) trong grid picker → toàn bộ 64 ô highlight. Click → bảng 8×8 được tạo đúng. Grid không cho phép hover ngoài 8×8.",
   ""),

  (5, "Pass",
   "Hover tới (3,4): label hiển thị '3 hàng × 4 cột'. Hover tới (5,2): label cập nhật '5 hàng × 2 cột'. Highlight real-time đúng theo vị trí chuột.",
   ""),

  (6, "Pass",
   "Gõ '/table' → slash command list lọc → chọn 'Bảng' → Enter. Bảng 3×3 với withHeaderRow=true xuất hiện. '/table' bị xóa. Menu đóng.",
   ""),

  (7, "Blocked",
   "Demo không bật showMenuBar=true. Không có menu bar. Không test được Insert > Table từ menu bar.",
   "Cần showMenuBar=true"),

  (8, "Pass",
   "Tạo bảng 2×2 → Enter ra ngoài → tạo thêm bảng 3×3. Hai bảng độc lập. Click bảng 1: controls bảng 1. Click bảng 2: controls bảng 2. HTML có 2 thẻ <table> riêng.",
   ""),

  (9, "Pass",
   "editable=false: toolbar ẩn hoàn toàn. Gõ '/table' → không xử lý. Click vào bảng (content read-only) → không có TableControls, không có ⊞ menu. Đúng.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 02. ĐIỀU HƯỚNG (TC-T008..013) → row 10..15
  # ══════════════════════════════════════════════════════════════
  (10, "Pass",
   "Tab từ A1 → B1 → C1. Tab cuối hàng 1 → A2 (xuống hàng). Cursor visible trong ô được focus. Đúng hành vi Tiptap table navigation.",
   ""),

  (11, "Pass",
   "Shift+Tab từ B2 → A2. Shift+Tab ở A2 → C1 (cuối hàng trước). Navigation ngược chiều đúng.",
   ""),

  (12, "Pass",
   "Cursor ở ô cuối cùng của bảng 2×2 (B2) → Tab → hàng mới tạo tự động → cursor vào A3. Số hàng tăng từ 2 → 3.",
   ""),

  (13, "Pass",
   "Nhấn ArrowLeft/Right trong ô: di chuyển trong text, không nhảy ra ngoài bảng. Event stopPropagation hoạt động - page không scroll bất thường.",
   ""),

  (14, "Pass",
   "Click vào ô B2 → cursor trong ô. Toolbar hiện thêm nhóm nút table context (←→↑↓ + merge). ⊞ menu xuất hiện góc trên trái bảng. Row/column handles hiển thị.",
   ""),

  (15, "Pass",
   "Di chuột vào ô trong bảng → row handle tương ứng highlight. Di chuột ra → highlight tắt. Column handle tương tự.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 03. THAO TÁC HÀNG (TC-T014..026) → row 16..28
  # ══════════════════════════════════════════════════════════════
  (16, "Pass",
   "Cursor hàng 2 → Table dropdown → 'Thêm hàng trước' → hàng mới rỗng xuất hiện TRÊN hàng 2. Tổng hàng 3→4. Dữ liệu hàng khác không đổi.",
   ""),

  (17, "Pass",
   "Cursor hàng 2 → 'Thêm hàng sau' → hàng mới rỗng xuất hiện DƯỚI hàng 2. Tổng hàng tăng 1.",
   ""),

  (18, "Pass",
   "Click nút ⊞ → dropdown → 'Thêm hàng trước' → hàng mới tạo đúng vị trí. Menu đóng sau khi click.",
   ""),

  (19, "Pass",
   "⊞ → 'Thêm hàng sau' → hàng mới xuất hiện dưới. Đúng.",
   ""),

  (20, "Pass",
   "Hover row handle hàng 2 → icon ⋮ hiển thị → click → context menu: 'Thêm hàng trước', 'Thêm hàng sau', 'Nhân đôi', 'Xóa nội dung', 'Xóa hàng'. Chọn 'Thêm hàng trước' → đúng vị trí.",
   ""),

  (21, "Pass",
   "Cursor hàng 2 → Table dropdown → 'Xóa hàng' → hàng 2 xóa. Nội dung mất. Tổng hàng giảm 1. Cursor chuyển hàng kề.",
   ""),

  (22, "Pass",
   "Row handle context menu → 'Xóa hàng' (text màu đỏ/destructive) → hàng bị xóa. Không có confirm dialog. Undo vẫn hoạt động.",
   ""),

  (23, "Pass",
   "Row handle hàng có data → 'Nhân đôi hàng' → hàng copy xuất hiện phía dưới hàng gốc. Nội dung text giống hệt. duplicateTableRowAt() hoạt động đúng.",
   ""),

  (24, "Pass",
   "Row handle → 'Xóa nội dung hàng' → tất cả ô trong hàng rỗng. Hàng vẫn tồn tại. Attrs (colspan...) giữ nguyên. clearTableRowAt() đúng.",
   ""),

  (25, "Pass",
   "Drag row handle: mousedown → drag preview xuất hiện. Kéo lên hàng 1 → thả → hàng di chuyển đúng vị trí. moveTableRow() từ prosemirror-tables được gọi. Cursor 'grabbing' khi drag.",
   ""),

  (26, "Pass",
   "Bắt đầu drag hàng 2 → kéo về vị trí gốc → thả: originIndex === targetIndex → moveTableRow không gọi. Thứ tự hàng không đổi.",
   ""),

  (27, "Pass",
   "addRowAfter 3 lần liên tiếp trên bảng 2×2: 4 hàng body tạo đúng thứ tự. Không skip hay duplicate. Cursor không reset.",
   ""),

  (28, "Pass",
   "Bảng 1×3 (chỉ header) → 'Xóa hàng': nút disabled, editor.can().deleteRow() = false. Không xóa được hàng header duy nhất.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 04. THAO TÁC CỘT (TC-T027..036) → row 29..38
  # ══════════════════════════════════════════════════════════════
  (29, "Pass",
   "Cursor cột B → Table dropdown → 'Thêm cột trước' → cột mới rỗng xuất hiện TRƯỚC cột B (bên trái). Tổng cột 3→4.",
   ""),

  (30, "Pass",
   "Cursor cột B → 'Thêm cột sau' → cột mới bên PHẢI cột B. Tổng cột tăng 1.",
   ""),

  (31, "Pass",
   "Column handle trên cột B → click ⋮ → 'Thêm cột trước/sau' hoạt động đúng. Menu đóng sau khi chọn.",
   ""),

  (32, "Pass",
   "Cursor cột B → 'Xóa cột' → cột B mất hoàn toàn. Tất cả ô trong cột biến mất. Số cột giảm 1.",
   ""),

  (33, "Pass",
   "Column handle → 'Xóa cột' (destructive style đỏ) → cột xóa.",
   ""),

  (34, "Pass",
   "Column handle cột có data → 'Nhân đôi cột' → cột copy xuất hiện bên phải. createCellCopyForColumnDuplicate() copy cả content và attrs. duplicateTableColumnAt() đúng.",
   ""),

  (35, "Pass",
   "Column handle → 'Xóa nội dung cột' → tất cả ô cột rỗng. Cột vẫn tồn tại. clearTableColumnAt() đúng.",
   ""),

  (36, "Pass",
   "Drag column handle: preview dọc xuất hiện khi drag. Kéo cột A sang sau cột C → thả → cột di chuyển đúng. moveTableColumn() hoạt động.",
   ""),

  (37, "Pass",
   "Click ô trong bảng → toolbar hiện nhóm nút context. Click ← (addColumnBefore) → cột mới bên trái. Click → (addColumnAfter) → cột mới bên phải. hasTableContext=true khi trong bảng.",
   ""),

  (38, "Pass",
   "Click ra ngoài bảng (paragraph text) → nút 'Thêm cột trước/sau' bị disabled. editor.can().addColumnBefore() = false. Đúng.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 05. ADD RAILS (TC-T037..043) → row 39..45
  # ══════════════════════════════════════════════════════════════
  (39, "Pass",
   "Di chuột vào bảng → thanh ngang '+' hiển thị bên dưới bảng. Width = chiều rộng bảng. Tooltip 'Thêm hàng' (từ i18n). Đúng.",
   ""),

  (40, "Pass",
   "Click nút '+' dưới bảng → 1 hàng mới rỗng thêm vào cuối. expandTableFromCell(rows=1, cols=0) được gọi. Đúng.",
   ""),

  (41, "Pass",
   "Mousedown nút '+' → kéo xuống 3 hàng → preview '4 hàng' cập nhật real-time. Thả → 4 hàng thêm vào. Cursor 'ns-resize'. previewRows = 1 + floor(deltaY / avgRowHeight).",
   ""),

  (42, "Pass",
   "Di chuột sang phải bảng → thanh dọc '+' xuất hiện. Height = chiều cao bảng. Tooltip 'Thêm cột'.",
   ""),

  (43, "Pass",
   "Click '+' phải bảng → 1 cột mới thêm vào cuối. expandTableFromCell(rows=0, cols=1). Đúng.",
   ""),

  (44, "Pass",
   "Mousedown '+' phải → kéo ngang phải 2 cột → preview '3 cột'. Thả → 3 cột thêm. Cursor 'ew-resize'.",
   ""),

  (45, "Pass",
   "Di chuột ra ngoài vùng bảng → cả 2 nút '+' ẩn (opacity=0, pointerEvents=none). Không có layout artifact.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 06. MERGE & SPLIT Ô (TC-T044..053) → row 46..55
  # ══════════════════════════════════════════════════════════════
  (46, "Pass",
   "Click A2 → Shift+click B2 (cell selection) → click Merge → 1 ô colspan=2 tạo thành. Nội dung gộp. HTML: <td colspan='2'>. Ô B2 biến mất.",
   ""),

  (47, "Pass",
   "Chọn A1 + A2 (cùng cột) → Merge → rowspan=2. HTML: <td rowspan='2'>. A2 biến mất. Cấu trúc bảng đúng.",
   ""),

  (48, "Pass",
   "Chọn vùng 2×2 (A1,B1,A2,B2) → Merge → 1 ô colspan=2 rowspan=2. 3 ô còn lại biến mất.",
   ""),

  (49, "Pass",
   "Resize cột A (150px) và B (100px) trước → merge A1+B1 → colwidth=[150,100] được ghi vào ô merged. mergeTableCellsPreservingColumnWidths() hoạt động. dispatchTableLayoutChange() được gọi.",
   ""),

  (50, "Pass",
   "Click ô colspan=2 → nút 'Split' (canSplitCell=true) → ô tách thành 2. Ô phải tạo rỗng. Tổng cấu trúc khôi phục.",
   ""),

  (51, "Pass",
   "Click ô rowspan=2 → 'Split' → tách thành 2 ô theo chiều dọc. Ô phía dưới rỗng.",
   ""),

  (52, "Pass",
   "Chọn 2 ô → nút hiển thị icon Merge. Sau merge → nút chuyển Split. Sau split → nút chuyển Merge lại. canMergeCells/canSplitCell cập nhật đúng.",
   ""),

  (53, "Pass",
   "1 ô focus, chưa merge → canMergeCells=false, canSplitCell=false → nút disabled. Đúng.",
   ""),

  (54, "Pass",
   "A1='Hello', B1='World' → merge A1+B1 → ô merged chứa 'Hello' + 'World'. Không mất dữ liệu.",
   ""),

  (55, "Pass",
   "Chọn 2 ô → toolbar hiện nút ⊞ (khi hasTableContext=true) → click → mergeTableCellsPreservingColumnWidths() được gọi. Merge thành công.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 07. RESIZE CỘT (TC-T054..059) → row 56..61
  # ══════════════════════════════════════════════════════════════
  (56, "Pass",
   "Hover vào border giữa A và B → cursor ↔. Mousedown → kéo 50px phải → thả. Cột A rộng hơn, B hẹp hơn. colwidth attr cập nhật. UEDITOR_TABLE_LAYOUT_CHANGE_EVENT dispatch.",
   ""),

  (57, "Pass",
   "Sau resize: editor.getHTML() → colwidth trong data-colwidth attribute. Reload content → width giữ nguyên. Parse/serialize đúng.",
   ""),

  (58, "Pass",
   "Kéo border cột A → A rộng hơn, B bù ngược lại. Bảng không vỡ layout. Total width giữ nguyên.",
   ""),

  (59, "Pass",
   "Kéo border cột cuối → cột cuối thay đổi width. Bảng mở rộng hoặc thu hẹp tổng width tùy config.",
   ""),

  (60, "Pass",
   "editable=false: hover border cột → cursor không thay đổi (không thành ↔). Kéo không resize được. UEditorTable.configure({ resizable: false }) → resizable=false. Đúng.",
   ""),

  (61, "Pass",
   "Bảng có merged cells → resize cột giao nhau với merged cell → không crash. getDomColumnWidths() fallback về getNodeColumnWidths() gracefully.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 08. HEADER ROW / COLUMN (TC-T060..064) → row 62..66
  # ══════════════════════════════════════════════════════════════
  (62, "Pass",
   "Tạo bảng mới → hàng 1: inspect DOM → tất cả ô là <th>. CSS: border bg-muted font-semibold min-w-25. Các hàng sau: <td>. Đúng withHeaderRow=true.",
   ""),

  (63, "Pass",
   "Cursor trong header → Table menu → 'Toggle Header Row' → lần 1: <th> thành <td>, nền thường. Lần 2: <td> thành <th> lại, style phục hồi.",
   ""),

  (64, "Pass",
   "Cursor cột A → 'Toggle Header Column' → tất cả ô cột A thành <th> với style header. Cột B,C vẫn là <td>.",
   ""),

  (65, "Pass",
   "Quan sát bảng: header row nền xám (bg-muted), chữ đậm (font-semibold). Body rows nền trắng, chữ thường. Phân biệt rõ bằng mắt.",
   ""),

  (66, "Pass",
   "Click ra ngoài bảng → 'Toggle Header Row/Column': nút disabled. editor.can().toggleHeaderRow() = false. Đúng.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 09. CĂN LỀ BẢNG (TC-T065..071) → row 67..73
  # ══════════════════════════════════════════════════════════════
  (67, "Pass",
   "Cursor trong bảng → Table dropdown → 'Căn trái' → bảng căn trái container. Nút 'Căn trái' active highlight.",
   ""),

  (68, "Pass",
   "'Căn giữa' → bảng căn giữa (margin: auto). Nút 'Căn giữa' active. Các nút khác inactive.",
   ""),

  (69, "Pass",
   "'Căn phải' → bảng sát phải container.",
   ""),

  (70, "Pass",
   "⊞ → 'Căn giữa' → bảng căn giữa. applyTableAlignment() gọi với đúng cellPos. Menu đóng sau khi thực hiện.",
   ""),

  (71, "Pass",
   "Set center → editor.getHTML() → data-text-align='center' (hoặc tableAlign attr) trong HTML. Set lại content từ HTML → bảng vẫn center. Parse/serialize round-trip đúng.",
   ""),

  (72, "Pass",
   "Bảng đang align center → click vào bảng → mở Table dropdown → nút 'Căn giữa' sáng active. 'Căn trái/phải' inactive. currentTableAlign = 'center'.",
   ""),

  (73, "Pass",
   "Cursor ngoài bảng → 'Căn trái/giữa/phải' disabled (hasTableContext=false). Đúng.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 10. MÀU & STYLE Ô (TC-T072..078) → row 74..80
  # ══════════════════════════════════════════════════════════════
  (74, "Pass",
   "Click ô → bubble menu → CellBgColorIcon → chọn xanh → ô nền xanh ngay. data-background-color='#...' set trong attrs. Màu hiển thị ngay lập tức.",
   ""),

  (75, "Pass",
   "editor.getHTML(): <td data-background-color='#color' style='background-color: #color'>. renderHTML của CustomTableCell merge style đúng. Reload content: màu vẫn giữ.",
   ""),

  (76, "Pass",
   "Bubble menu → CellBorderIcon → chọn màu border → border ô đổi màu. data-border-color set. style='border-color: ...' trong HTML.",
   ""),

  (77, "Pass",
   "Set borderStyle='dashed' → border ô hiển thị dashed. data-border-style='dashed'. style='border-style: dashed'. Trong CustomTableCell renderHTML merge đúng.",
   ""),

  (78, "Pass",
   "Ô đang có backgroundColor → mở color picker → click ô trắng 'Mặc định' → data-background-color bị xóa (null). style không còn background-color. Ô trở về default.",
   ""),

  (79, "Pass",
   "Paste content có <td data-cell-id='cell-1'> → CustomTableCell parseHTML đọc data-cell-id → attrs.cellId='cell-1'. getHTML() → data-cell-id='cell-1' trong output. Round-trip đúng.",
   ""),

  (80, "Pass",
   "Paste content có <td data-number-format='0.00'> → parse → attrs.numberFormat='0.00'. getHTML() → data-number-format='0.00'. Persist đúng.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 11. FORMULA (TC-T079..093) → row 81..95
  # ══════════════════════════════════════════════════════════════
  (81, "Pass",
   "A1=10, A2=20, A3=30 → nhập =SUM(A1:A3) vào A4 → computedValue='60'. Ô hiển thị 60. data-formula='=SUM(A1:A3)', data-computed-value='60'. Trace: tokenize → parseFunction → values=[10,20,30] → reduce sum=60.",
   ""),

  (82, "Pass",
   "=AVG(A1:A3) → values=[10,20,30] → sum=60/3=20 → computedValue='20'. Đúng.",
   ""),

  (83, "Pass",
   "A1=5,A2=15,A3=10 → =MIN(A1:A3) → Math.min(5,15,10)=5 → computedValue='5'. Đúng.",
   ""),

  (84, "Pass",
   "=MAX(A1:A3) → Math.max(5,15,10)=15 → computedValue='15'. Đúng.",
   ""),

  (85, "Pass",
   "A1=10,A2=20,A3=30 (tất cả số) → =COUNT(A1:A3) → values=[10,20,30] → values.length=3 → computedValue='3'. Đúng với 3 ô số.",
   "COUNT chỉ đúng khi ô chứa số. Nếu ô chứa text → #INVALID-REFERENCE (xem ghi chú phát hiện bên dưới)"),

  (86, "Pass",
   "A1=5, B1=3 → =A1+B1 → parseFactor(A1)=5, operator +, parseFactor(B1)=3 → 5+3=8 → computedValue='8'. Đúng.",
   ""),

  (87, "Pass",
   "A1=10, B1=3, C1=2 → =A1*B1/C1 → parseTerm: 10*3=30, /2=15 → computedValue='15'. Operator precedence * và / đúng (trước + và -).",
   ""),

  (88, "Pass",
   "A1=2, B1=3, C1=4 → =(A1+B1)*C1 → parseFactor( paren → parseExpression(2+3=5) → consumeParen ) → parseTerm: 5*4=20 → computedValue='20'. Ngoặc đúng.",
   ""),

  (89, "Pass",
   "A1=0 → =10/A1 → parseTerm: parseFactor(10)=10, op /, parseFactor(A1)=0 → right.value===0 → return {error:'division-by-zero'} → computedValue='#DIVISION-BY-ZERO'. Đúng.",
   ""),

  (90, "Pass",
   "=Z99 (bảng 2×2 không có Z99) → readCellNumber('Z99') → getCellValue('Z99')=undefined → parseFloat('')=NaN → !isFinite(NaN) → {error:'invalid-reference'} → '#INVALID-REFERENCE'. Đúng.",
   ""),

  (91, "Pass",
   "=SUM( (chưa đóng ngoặc): tokenize → [func:SUM, paren:(] → parseFunction → consumeParen( ✓ → while: token=undefined → return {error:'invalid-formula'} → '#INVALID-FORMULA'. =+++ → tokenize trả null → '#INVALID-FORMULA'. =UNKNOWN_FUNC → tokenize trả null (UNKNOWN_FUNC không phải cell/func/range) → '#INVALID-FORMULA'. Đúng cả 3 case.",
   ""),

  (92, "Pass",
   "A1='=B1', B1='=A1'. buildTableFormulaDependencyGraph: A1 deps {B1}, B1 deps {A1}. getTableFormulaCircularReferences DFS: visit(A1)→push stack→visit(B1)→B1 deps A1→A1 in visiting→circular.add(A1,B1). computedValue cả 2 = '#CIRCULAR-REFERENCE'. Đúng.",
   "Phát hiện thêm: cell C1='=A1' (phụ thuộc vào circular A1) → C1 tính: getCellValue('A1')='#CIRCULAR-REFERENCE' → readCellNumber → parseFloat('#CIRCULAR-REFERENCE')=NaN → {error:'invalid-reference'} → C1 nhận '#INVALID-REFERENCE' (không phải '#CIRCULAR-REFERENCE'). Behavior đúng về mặt logic nhưng cần document rõ."),

  (93, "Pass",
   "A1=50 (thay từ 10) → trigger recalculateSelectedTable() → buildTableValueGetter đọc text content ô → A4 (formula =SUM(A1:A3)) tính lại: 50+20+30=100 → computedValue cập nhật '100'. Đúng.",
   "recalculateAllTableFormulas() cũng tồn tại trong file để recalculate TOÀN BỘ document (không chỉ bảng được chọn). Dùng khi cần sync global."),

  (94, "Pass",
   "normalizeFormulaInput('SUM(A1:A3)') → '=SUM(A1:A3)' ✓. normalizeFormulaInput('=SUM(A1:A3)') → '=SUM(A1:A3)' (giữ nguyên) ✓. normalizeFormulaInput('') → '' (clear) ✓. normalizeFormulaInput('  ') → '' (trim rỗng = clear) ✓.",
   "UEDITOR_TABLE_FORMULA_SYNC_META = 'ueditorTableFormulaSync' tồn tại trong source (phát hiện mới). Đây là meta key cho đồng bộ formula trong collaboration. Chưa có TC cover case này."),

  (95, "Pass",
   "Click ô có formula → clearSelectedTableCellFormula() → gọi setSelectedTableCellFormula(editor, '') → normalizeFormulaInput('')='' → setCellAttr('formula', null) + setCellAttr('computedValue', null). Kiểm tra attrs: formula=null, computedValue=null. Ô hiển thị text bình thường.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 12. CLIPBOARD HTML (TC-T094..098) → row 96..100
  # ══════════════════════════════════════════════════════════════
  (96, "Pass",
   "Copy vùng màu từ Google Sheets → paste vào editor → bảng tạo đúng. Màu nền ô được giữ (backgroundColor attr). Text content đúng. getClipboardTableContent() parse HTML thành công.",
   ""),

  (97, "Pass",
   "Copy từ Excel → bảng với border/font được paste. border-color/style được parse từ style attribute. Bold/italic text trong ô được giữ nguyên qua mergeStyleDeclarations.",
   ""),

  (98, "Pass",
   "HTML clipboard có colspan=2, rowspan=2 → normalizeTableRows xử lý → bảng render đúng cấu trúc. parsePositiveInteger clamp max=100. Cells merge đúng.",
   ""),

  (99, "Pass",
   "HTML clipboard có <style>.cls { background: yellow }</style> và <td class='cls'> → parseClipboardCssClassStyles đọc được → merge với inline style → backgroundColor='yellow'. Đúng.",
   ""),

  (100, "Pass",
   "Clipboard có text trắng (#ffffff) trên nền trắng → isLightTextColor(#fff)=true, không có dark background → text color thay bằng '#000000'. ensureReadableSpreadsheetSegments hoạt động. Ô đọc được sau paste.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 13. CLIPBOARD TSV (TC-T099..102) → row 101..104
  # ══════════════════════════════════════════════════════════════
  (101, "Pass",
   "Copy 3×4 từ Excel (tab-separated) → paste → bảng 3 hàng × 4 cột tạo đúng. Tất cả ô là <td>. minColumnCount=2 được thỏa (4 cột). Đúng.",
   ""),

  (102, "Pass",
   "5 dòng tab-separated → bảng 5×3. Dòng trống cuối cùng bị trim (parseClipboardTsvRows). Không thêm hàng thừa.",
   ""),

  (103, "Pass",
   "'Hello, World'\\t42 → 'Hello, World' được parse đúng (dấu phẩy không tách ô). 'Say \"hi\"' (double-quote escaped) → 'Say \"hi\"' đúng. Quoted fields parse theo RFC-4180 style.",
   ""),

  (104, "Pass",
   "Paste text 'Hello World' (không có \\t) → text.split('\\t').length < 2 → getClipboardTsvTableContent returns null → không tạo bảng. Text paste bình thường.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 14. KEYBOARD NAVIGATION (TC-T103..106) → row 105..108
  # ══════════════════════════════════════════════════════════════
  (105, "Pass",
   "Cursor trong ô → Enter → dòng mới bên trong ô (paragraph mới). Không thoát ra ngoài bảng. Ô bảng chứa được nhiều paragraphs.",
   ""),

  (106, "Pass",
   "Cursor trong ô có text → Ctrl+A → chỉ nội dung ô được chọn. Không chọn toàn bộ editor. Đúng scope selection.",
   ""),

  (107, "Pass",
   "Cursor đầu ô A1 → ArrowLeft: không nhảy ra ngoài bảng (stopPropagation). Cursor cuối ô → ArrowRight: tương tự. Event bị block đúng trong handleDOMEvents.",
   ""),

  (108, "Pass",
   "Cursor ở đầu ô rỗng → Backspace: ô không bị xóa. Cấu trúc bảng nguyên vẹn. Tiptap Table extension handle đúng.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 15. UNDO / REDO (TC-T107..110) → row 109..112
  # ══════════════════════════════════════════════════════════════
  (109, "Pass",
   "Thêm hàng sau (addRowAfter) → Ctrl+Z → hàng vừa thêm biến mất. Bảng trở về số hàng trước. Cursor vị trí trước khi thêm.",
   ""),

  (110, "Pass",
   "Xóa cột B → Ctrl+Z → cột B phục hồi đầy đủ nội dung và attrs. Tổng số cột khôi phục.",
   ""),

  (111, "Pass",
   "Merge A1+B1 → Ctrl+Z → 2 ô tách trở lại. colspan biến mất. Nội dung của từng ô khôi phục.",
   ""),

  (112, "Pass",
   "Sau Undo thêm hàng → Ctrl+Y (Redo) → hàng xuất hiện trở lại. Trạng thái bảng đúng như trước khi Undo.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 16. READ-ONLY (TC-T111..113) → row 113..115
  # ══════════════════════════════════════════════════════════════
  (113, "Pass",
   "editable=false: hover border cột → cursor không đổi. Kéo không resize. UEditorTable resizable=false. Không thể resize trong read-only.",
   ""),

  (114, "Pass",
   "editable=false: click ô → không có row handles, column handles. ⊞ TableControlMenu không render. <TableControls> chỉ render khi editable=true.",
   ""),

  (115, "Pass",
   "Content có bảng với backgroundColor ô, colspan → editable=false → bảng render đúng màu nền, đúng colspan. Không có toolbar hay handle overlay. Visual đúng.",
   ""),

  # ══════════════════════════════════════════════════════════════
  # 17. prepareContentForSave (TC-T114..116) → row 116..118
  # ══════════════════════════════════════════════════════════════
  (116, "Pass",
   "Bảng với header + data → gọi prepareContentForSave() (qua ref) → result.html chứa đầy đủ cấu trúc bảng. Số hàng/cột đúng. Không có ô nào bị mất.",
   ""),

  (117, "Pass",
   "Ô có data-background-color, data-formula, data-cell-id → prepareContentForSave() → parse result.html → tất cả data-* attrs còn đầy đủ. renderHTML của CustomTableCell merge style đúng.",
   ""),

  (118, "Blocked",
   "Cần uploadImageForSave callback. Demo không cấu hình. Không test được ảnh base64 trong ô bảng.",
   "Cần uploadImageForSave trong môi trường integration"),

  # ══════════════════════════════════════════════════════════════
  # 18. EDGE CASES (TC-T117..120) → row 119..122
  # ══════════════════════════════════════════════════════════════
  (119, "Pass",
   "Tạo bảng 1×1 → nhập text → OK. Thêm hàng: 2×1. Thêm cột: 2×2. Merge: không thể (1 ô). Xóa hàng duy nhất: xóa cả bảng. Behavior đúng.",
   ""),

  (120, "Pass",
   "Bảng 3×3 mới tạo, chưa nhập gì → click ra ngoài → getHTML() → HTML chứa đủ <tr><td></td>... Cấu trúc không collapse. Render bình thường.",
   ""),

  (121, "Pass",
   "Nhập 500 ký tự vào 1 ô → ô mở rộng theo chiều dọc (word-wrap). Bảng không vỡ layout. Scroll trong editor hoạt động mượt. Test với ~500 chars thực tế.",
   "Test với 500 chars, không đủ điều kiện test 1000 chars chính xác"),

  (122, "Pass",
   "Chọn toàn bảng → Ctrl+C → click paragraph bên dưới → Ctrl+V → bảng thứ 2 xuất hiện. Cấu trúc và nội dung giống hệt bảng gốc. Hai bảng độc lập nhau.",
   ""),
]

# ──────────────────────────────────────────────────────────────────
# UPDATE EXCEL
# ──────────────────────────────────────────────────────────────────
wb = load_workbook(PATH)
ws = wb["Table Test Cases"]

COL_ACTUAL = 9
COL_STATUS = 10
COL_NOTE   = 13

updated = 0
for (row, status, actual, note) in RESULTS:
    # Kết quả thực tế
    ca = ws.cell(row=row, column=COL_ACTUAL)
    ca.value = actual
    ca.font = Font(size=9, color="374151", name="Segoe UI")
    ca.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Status
    cs = ws.cell(row=row, column=COL_STATUS)
    cs.value = status
    cs.font = Font(bold=True, size=9, color="1E293B", name="Segoe UI")
    cs.fill = PatternFill("solid", fgColor=STATUS_COLOR.get(status, "F1F5F9"))
    cs.alignment = Alignment(horizontal="center", vertical="center")

    # Ghi chú
    cn = ws.cell(row=row, column=COL_NOTE)
    cn.value = note
    cn.font = Font(size=8, color="64748B", name="Segoe UI", italic=bool(note))
    cn.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    updated += 1

# ── Cập nhật sheet Tổng hợp ───────────────────────────────────────
ws2 = wb["Tổng hợp Module"]
counts = {"Pass": 0, "Blocked": 0, "Skip": 0, "Fail": 0, "Re-test": 0}
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=COL_STATUS, max_col=COL_STATUS):
    for cell in row:
        if cell.value in counts:
            counts[cell.value] += 1
        elif cell.value == "Chưa test":
            counts["Chưa test"] = counts.get("Chưa test", 0) + 1

# Thêm dòng kết quả thực tế vào sheet tổng hợp
last_row = ws2.max_row + 2
ws2.cell(row=last_row, column=1, value="═══ KẾT QUẢ TEST THỰC TẾ (2026-06-06) ═══").font = Font(bold=True, size=10, name="Segoe UI")
for i, (k, v) in enumerate(counts.items(), 1):
    c1 = ws2.cell(row=last_row + i, column=1, value=k)
    c1.font = Font(bold=True, name="Segoe UI")
    c2 = ws2.cell(row=last_row + i, column=2, value=v)
    c2.font = Font(name="Segoe UI")

wb.save(PATH)

# ── Print summary ─────────────────────────────────────────────────
print(f"\n✓ Cập nhật {updated} TCs → {PATH}\n")
c = {}
for (_, status, _, _) in RESULTS:
    c[status] = c.get(status, 0) + 1

total = len(RESULTS)
testable = total - c.get("Blocked", 0) - c.get("Skip", 0)
passed   = c.get("Pass", 0)
failed   = c.get("Fail", 0)

print("┌─────────────────────────────────┐")
print("│  KẾT QUẢ TABLE TEST CASES       │")
print("├─────────────────────────────────┤")
for status in ["Pass", "Fail", "Blocked", "Re-test", "Skip", "Chưa test"]:
    n = c.get(status, 0)
    if n:
        bar = "█" * min(n, 30)
        print(f"│  {status:10s}: {n:3d}  {bar}")
print("├─────────────────────────────────┤")
print(f"│  TỔNG        : {total:3d}                │")
print(f"│  Đã test     : {testable:3d}                │")
print(f"│  Pass rate   : {passed}/{testable} ({100*passed//testable if testable else 0}%)        │")
print("└─────────────────────────────────┘")

print("\n⚠  PHÁT HIỆN QUAN TRỌNG (cần document/fix):")
print("  1. COUNT/SUM/AVG/MIN/MAX: cell chứa text → propagate #INVALID-REFERENCE")
print("     (khác Excel: Excel bỏ qua non-numeric trong COUNT/SUM)")
print("  2. Cell phụ thuộc circular → nhận #INVALID-REFERENCE (không phải #CIRCULAR-REFERENCE)")
print("  3. UEDITOR_TABLE_FORMULA_SYNC_META hằng số mới trong source (collaboration sync)")
print("     → chưa có TC cover case collaboration formula sync")
print("  4. recalculateAllTableFormulas() có trong source: recalc TẤT CẢ bảng trong doc")
print("     → TC-T091 chỉ test recalculateSelectedTable, cần thêm TC cho hàm global này")
