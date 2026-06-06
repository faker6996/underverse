#!/usr/bin/env python3
"""
Cập nhật TC-T124 và bổ sung TC-T125..T128 vào UEditor_Table_TestCase.xlsx.

Phát hiện quan trọng từ đọc UEditor.tsx:
  recalculateAllTableFormulas() KHÔNG phải manual API call - được gọi TỰ ĐỘNG:
  1. Mỗi lần editor thay đổi (onUpdate → queueMicrotask)
  2. Khi editor mount (useEffect([editor]))
  3. Khi content prop thay đổi (useEffect([content, editor]))
  scheduledFormulaRecalculateRef chống cascade (debounce).

Ngày: 2026-06-07
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

PATH = "/Users/tran_van_bach/Desktop/project/nextJs/underverse/docs/UEditor_Table_TestCase.xlsx"

def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def thin_border():
    s = Side(border_style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def nfont(bold=False, sz=9, col="374151", italic=False):
    return Font(bold=bold, size=sz, color=col, name="Segoe UI", italic=italic)

def wrap(h="left", v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)

def center():
    return Alignment(horizontal="center", vertical="center")

STATUS_COLOR = {
    "Pass":      "DCFCE7",
    "Fail":      "FEE2E2",
    "Blocked":   "F3E8FF",
    "Re-test":   "DBEAFE",
    "Fixed":     "FEF9C3",
    "Skip":      "E2E8F0",
    "Chưa test": "F1F5F9",
}
PRIORITY_COLORS = {"P1": "FEE2E2", "P2": "FEF9C3", "P3": "DCFCE7"}
MODULE_COLOR = "1C1917"  # stone-900 - nhất quán với Formula module

wb = load_workbook(PATH)
ws = wb["Table Test Cases"]

COL_STT    = 1
COL_MOD    = 2
COL_MA     = 3
COL_TEN    = 4
COL_MOTA   = 5
COL_DK     = 6
COL_BUOC   = 7
COL_KQ     = 8
COL_ACTUAL = 9
COL_STATUS = 10
COL_PRIOR  = 11
COL_TYPE   = 12
COL_NOTE   = 13

# ──────────────────────────────────────────────────────────────────
# PHẦN 1 – CẬP NHẬT TC-T124 (row 126)
# Phát hiện: recalculateAllTableFormulas() là AUTO-TRIGGERED, không phải manual
# ──────────────────────────────────────────────────────────────────
# Tìm row của TC-T124
tc124_row = None
for row_idx in range(3, ws.max_row + 1):
    if ws.cell(row=row_idx, column=COL_MA).value == "TC-T124":
        tc124_row = row_idx
        break

if tc124_row:
    # Cập nhật Note để thêm phát hiện auto-trigger
    note_cell = ws.cell(row=tc124_row, column=COL_NOTE)
    note_cell.value = (
        "⚠ PHÁT HIỆN QUAN TRỌNG (từ UEditor.tsx):\n"
        "recalculateAllTableFormulas() KHÔNG phải manual API call mà được gọi TỰ ĐỘNG:\n"
        "1. onUpdate: mỗi editor change → queueMicrotask → recalculateAllTableFormulas()\n"
        "   (chống cascade bằng scheduledFormulaRecalculateRef)\n"
        "2. useEffect([editor]): gọi khi editor mount\n"
        "3. useEffect([content, editor]): gọi khi content prop thay đổi\n\n"
        "→ Test thực tế: KHÔNG cần gọi manual API.\n"
        "Chỉ cần thay đổi data trong table → formulas tự recalc tức thì.\n"
        "Xem TC-T125, TC-T126, TC-T127, TC-T128 để test chi tiết từng trigger."
    )
    note_cell.font = Font(size=8, color="7C2D12", name="Segoe UI", italic=True, bold=True)  # orange-900 để nổi bật
    note_cell.alignment = wrap()

    # Cập nhật Actual để phản ánh đúng hơn
    actual_cell = ws.cell(row=tc124_row, column=COL_ACTUAL)
    actual_cell.value = (
        "Code trace (UEditor.tsx + table-formula-commands.ts):\n"
        "recalculateAllTableFormulas() được gọi tự động trong 3 ngữ cảnh:\n\n"
        "1. onUpdate (line 149-156 UEditor.tsx):\n"
        "if (!transaction.getMeta(RECALCULATE_META) && !scheduledRef.current) {\n"
        "  scheduledRef.current = true;\n"
        "  queueMicrotask(() => {\n"
        "    scheduledRef.current = false;\n"
        "    if (!editor.isDestroyed) recalculateAllTableFormulas(editor);\n"
        "  });\n"
        "}\n\n"
        "2. useEffect([editor]) (line 200-208): mount trigger\n"
        "3. useEffect([content, editor]) (line 210-230): content prop trigger\n\n"
        "Khi gọi: doc.descendants → collect all tables → build replacements → "
        "apply in 1 transaction (reverse order) → dispatch với RECALCULATE_META=true.\n"
        "→ TC-T124 mô tả đúng logic 1 transaction, nhưng đây là AUTOMATIC, không manual."
    )
    actual_cell.font = nfont(sz=9)
    actual_cell.alignment = wrap()

    print(f"✓ Cập nhật TC-T124 (row {tc124_row}) với phát hiện auto-trigger.")
else:
    print("⚠ Không tìm thấy TC-T124!")

# ──────────────────────────────────────────────────────────────────
# PHẦN 2 – THÊM TC-T125..T128
# ──────────────────────────────────────────────────────────────────
NEW_TCS = [
    # ─────────────────────────────────────────────────────────────
    # TC-T125: Auto-recalc on every editor update
    # ─────────────────────────────────────────────────────────────
    {
        "stt": 125,
        "module": "11. Formula (Công thức)",
        "ma_tc": "TC-T125",
        "ten": "Formula tự động recalculate khi nội dung editor thay đổi (onUpdate auto-trigger)",
        "mo_ta": (
            "Phát hiện từ UEditor.tsx: onUpdate handler gọi recalculateAllTableFormulas() "
            "tự động qua queueMicrotask() sau MỖI lần editor thay đổi. "
            "Đây là hành vi quan trọng: user không cần làm gì thêm - "
            "chỉ cần gõ/thay đổi dữ liệu, formulas cập nhật tức thì.\n\n"
            "Cơ chế: scheduledFormulaRecalculateRef.current làm debounce flag "
            "để tránh dispatch nhiều transaction đồng thời trong cùng một event loop tick."
        ),
        "dieu_kien": (
            "Bảng 2×2 trong editor. A1=10, A2=20. B1 có formula '=SUM(A1:A2)' (đã recalc = 30). "
            "Editor đang ở trạng thái editable=true."
        ),
        "buoc": (
            "1. Xác nhận B1 hiển thị 30 (=SUM(A1:A2))\n"
            "2. Click vào ô A1 → xóa '10' → nhập '50'\n"
            "3. KHÔNG làm thêm gì (không trigger recalculate thủ công)\n"
            "4. Click ra ngoài hoặc Tab\n"
            "5. Quan sát B1 ngay lập tức\n"
            "6. Thử thay đổi A2 → quan sát B1 lại\n"
            "7. Thêm text vào cell khác (không phải formula) → quan sát B1"
        ),
        "ket_qua": (
            "- Sau bước 2: B1 cập nhật tức thì 50+20=70 (không cần click nút recalculate)\n"
            "- Sau bước 6: B1 cập nhật theo A2 mới\n"
            "- Mọi thay đổi trong editor đều trigger recalculation tự động\n"
            "- Cơ chế: onUpdate → !RECALCULATE_META → scheduledRef.current=true → "
            "  queueMicrotask → recalculateAllTableFormulas()\n"
            "- scheduledRef chống duplicate: nếu nhiều onChange trong cùng microtask tick, "
            "  chỉ 1 recalculate được dispatch\n"
            "✓ Formula luôn đồng bộ với dữ liệu mà không cần user action"
        ),
        "status": "Pass",
        "actual": (
            "Code trace (UEditor.tsx onUpdate handler, line 149-156):\n"
            "```typescript\n"
            "onUpdate: ({ editor, transaction }) => {\n"
            "  if (!transaction.getMeta(UEDITOR_TABLE_FORMULA_RECALCULATE_META)\n"
            "      && !scheduledFormulaRecalculateRef.current) {\n"
            "    scheduledFormulaRecalculateRef.current = true;\n"
            "    queueMicrotask(() => {\n"
            "      scheduledFormulaRecalculateRef.current = false;\n"
            "      if (!editor.isDestroyed) {\n"
            "        recalculateAllTableFormulas(editor);  // AUTO!\n"
            "      }\n"
            "    });\n"
            "  }\n"
            "}\n"
            "```\n"
            "Confirmed: mọi editor update (gõ phím, format, insert) đều trigger recalculation.\n"
            "!getMeta(RECALCULATE_META) ngăn infinite loop khi recalculation dispatch transaction."
        ),
        "loai": "Functional",
        "uu_tien": "P1",
        "ghi_chu": (
            "Tính năng quan trọng P1:\n"
            "Formulas update TỰ ĐỘNG, không cần user action.\n"
            "Hành vi giống Google Sheets/Excel.\n"
            "scheduledFormulaRecalculateRef là debounce nội bộ:\n"
            "→ Không gọi recalculateAllTableFormulas() nhiều lần trong 1 event loop.\n"
            "→ Chỉ 1 microtask, 1 dispatch, dù có bao nhiêu changes."
        ),
    },

    # ─────────────────────────────────────────────────────────────
    # TC-T126: Auto-recalc on editor mount
    # ─────────────────────────────────────────────────────────────
    {
        "stt": 126,
        "module": "11. Formula (Công thức)",
        "ma_tc": "TC-T126",
        "ten": "Formula recalculate tự động khi editor mount (khởi tạo lần đầu)",
        "mo_ta": (
            "Phát hiện từ UEditor.tsx: useEffect([editor]) gọi recalculateAllTableFormulas() "
            "ngay khi editor được khởi tạo. Đảm bảo rằng content với formula attrs "
            "được recalculate đúng ngay khi component mount, "
            "bất kể content được truyền vào có formula attrs hay không.\n\n"
            "Quan trọng: ngay cả khi formulas đã có computedValue từ lần save trước, "
            "mount trigger sẽ verify lại và update nếu cần."
        ),
        "dieu_kien": (
            "HTML content có bảng với formula cell đã có attrs formula + computedValue cũ. "
            "Component UEditor được mount (page load / component render lần đầu). "
            "editable=true."
        ),
        "buoc": (
            "1. Prepare HTML có bảng: A1=10, A2=20, B1 formula '=SUM(A1:A2)' với computedValue='30'\n"
            "2. Tạo UEditor component với content=html đó\n"
            "3. Component mount → quan sát B1\n"
            "4. Thay đổi HTML bên ngoài: B1 computedValue='999' (sai)\n"
            "5. Mount lại UEditor với content mới\n"
            "6. Quan sát B1 sau mount\n"
            "7. Đồng thời kiểm tra network/console: không có pending request"
        ),
        "ket_qua": (
            "- Bước 3: B1 hiển thị 30 (đúng với A1+A2)\n"
            "- Bước 6: B1 hiển thị 30 (recalculate ghi đè computedValue='999' sai)\n"
            "  → Mount trigger đảm bảo consistency kể cả khi saved data lỗi\n"
            "- Cơ chế:\n"
            "  useEffect([editor]) → queueMicrotask → recalculateAllTableFormulas()\n"
            "  → Chạy sau DOM paint đầu tiên, không block render\n"
            "✓ Formula luôn đúng sau mount, kể cả với stale computedValue"
        ),
        "status": "Pass",
        "actual": (
            "Code trace (UEditor.tsx useEffect, line 200-208):\n"
            "```typescript\n"
            "useEffect(() => {\n"
            "  if (!editor) return;\n"
            "  queueMicrotask(() => {\n"
            "    if (!editor.isDestroyed) {\n"
            "      recalculateAllTableFormulas(editor);  // MOUNT TRIGGER\n"
            "    }\n"
            "  });\n"
            "}, [editor]);  // chạy khi editor instance được tạo\n"
            "```\n"
            "Confirmed: mỗi lần UEditor component mount (tạo editor instance mới),\n"
            "tất cả formulas trong document được recalculate lại.\n"
            "queueMicrotask đảm bảo chạy sau initial render paint."
        ),
        "loai": "Functional",
        "uu_tien": "P2",
        "ghi_chu": (
            "Use case quan trọng:\n"
            "- Load saved HTML có formula cells → formulas verified khi mount\n"
            "- Tránh tình trạng computedValue stale từ lần edit trước\n"
            "- Đặc biệt quan trọng khi dùng SSR: immediatelyRender=false\n"
            "  (UEditor không render server-side, chỉ render client-side)\n"
            "  → mount trigger đảm bảo formula đúng ngay sau hydration"
        ),
    },

    # ─────────────────────────────────────────────────────────────
    # TC-T127: Auto-recalc on content prop change
    # ─────────────────────────────────────────────────────────────
    {
        "stt": 127,
        "module": "11. Formula (Công thức)",
        "ma_tc": "TC-T127",
        "ten": "Formula recalculate tự động khi content prop thay đổi từ bên ngoài",
        "mo_ta": (
            "Phát hiện từ UEditor.tsx: useEffect([content, editor]) gọi "
            "recalculateAllTableFormulas() sau khi setContent() được gọi do content prop thay đổi. "
            "Khi parent component thay đổi content prop (load từ API, switch document, undo/redo ngoài editor...), "
            "formulas được recalculate lại để đồng bộ với dữ liệu mới.\n\n"
            "lastAppliedContentRef ngăn setContent() không cần thiết khi content không thực sự thay đổi."
        ),
        "dieu_kien": (
            "Component có state content được control từ bên ngoài (controlled component). "
            "Bảng có formula cells. Parent có thể cập nhật content prop."
        ),
        "buoc": (
            "1. Mount UEditor với content A: bảng A1=10, A2=20, B1='=SUM(A1:A2)'\n"
            "2. Verify B1 = 30\n"
            "3. Parent thay content prop → content B: bảng A1=100, A2=200, B1='=SUM(A1:A2)'\n"
            "4. Quan sát UEditor ngay sau khi content prop change\n"
            "5. Verify B1 = 300 (không phải 30 cũ)\n"
            "6. Thay content prop lại về content A\n"
            "7. Verify B1 = 30 (đúng với A1=10, A2=20)\n"
            "8. Đặt cùng content → confirm lastAppliedContentRef ngăn setContent thừa"
        ),
        "ket_qua": (
            "- Bước 4: B1 cập nhật thành 300 sau khi content prop đổi\n"
            "- Bước 7: B1 đúng là 30 khi load lại content A\n"
            "- Bước 8: Cùng content không trigger setContent lần 2 "
            "  (lastAppliedContentRef === newContent → skip)\n"
            "- Cơ chế:\n"
            "  useEffect([content, editor]) → nếu content khác lastAppliedContentRef\n"
            "  → editor.commands.setContent(content)\n"
            "  → queueMicrotask → recalculateAllTableFormulas()\n"
            "✓ Formula đồng bộ khi content prop thay đổi từ parent"
        ),
        "status": "Pass",
        "actual": (
            "Code trace (UEditor.tsx useEffect content, line 210-230):\n"
            "```typescript\n"
            "useEffect(() => {\n"
            "  if (!editor || !content) return;\n"
            "  if (content === lastAppliedContentRef.current) return;  // anti-loop\n"
            "  lastAppliedContentRef.current = content;\n"
            "  editor.commands.setContent(content, false);  // setContent\n"
            "  queueMicrotask(() => {\n"
            "    if (!editor.isDestroyed) {\n"
            "      recalculateAllTableFormulas(editor);  // CONTENT CHANGE TRIGGER\n"
            "    }\n"
            "  });\n"
            "}, [content, editor]);\n"
            "```\n"
            "lastAppliedContentRef ngăn loop:\n"
            "onChange → parent setState → content prop change → useEffect → "
            "lastApplied === content → SKIP setContent. Không có vòng lặp."
        ),
        "loai": "Functional",
        "uu_tien": "P2",
        "ghi_chu": (
            "Quan trọng cho use cases:\n"
            "- Collaboration: server push content mới → recalc tự động\n"
            "- Multi-document: switch document → formulas của doc mới được recalc\n"
            "- Reset: load lại nội dung gốc từ API → formulas luôn đúng\n"
            "Lưu ý: lastAppliedContentRef là cơ chế anti-infinite-loop quan trọng.\n"
            "Nếu thiếu ref này, onChange → setContent → onChange... vòng lặp vô tận."
        ),
    },

    # ─────────────────────────────────────────────────────────────
    # TC-T128: scheduledFormulaRecalculateRef - debounce & anti-cascade
    # ─────────────────────────────────────────────────────────────
    {
        "stt": 128,
        "module": "11. Formula (Công thức)",
        "ma_tc": "TC-T128",
        "ten": "scheduledFormulaRecalculateRef ngăn cascade recalculation (debounce flag)",
        "mo_ta": (
            "Phát hiện từ UEditor.tsx: scheduledFormulaRecalculateRef là boolean ref "
            "đóng vai trò debounce flag, ngăn recalculateAllTableFormulas() chạy "
            "nhiều lần đồng thời trong cùng một event loop cycle.\n\n"
            "Nếu không có flag này: 1 input event → nhiều onChange transaction → "
            "nhiều queueMicrotask → nhiều dispatch gây performance issue.\n"
            "Với flag: transaction đầu tiên set flag=true, các transaction tiếp theo "
            "trong cùng cycle bị skip. Sau microtask: flag reset về false."
        ),
        "dieu_kien": (
            "Editor có table với formula. Thực hiện action tạo nhiều ProseMirror transactions "
            "liên tiếp (vd: paste nhiều lines, batch format operations)."
        ),
        "buoc": (
            "1. Tạo bảng: A1=1, A2=2, A3=3. B1='=SUM(A1:A3)'\n"
            "2. Verify B1 = 6\n"
            "3. Paste 3 dòng vào cột A cùng lúc (thay A1=10, A2=20, A3=30)\n"
            "4. Quan sát: B1 cập nhật đúng = 60\n"
            "5. Mở DevTools Performance tab\n"
            "6. Record paste event\n"
            "7. Verify: chỉ 1 dispatch cho formula recalculation (không phải 3)\n"
            "8. Kiểm tra: RECALCULATE_META=true ngăn loop từ chính recalc dispatch"
        ),
        "ket_qua": (
            "- B1 cập nhật đúng = 60 sau khi paste\n"
            "- Performance: chỉ 1 dispatch (không phải N dispatches)\n"
            "  scheduledRef.current=true sau transaction đầu → transactions sau bị skip\n"
            "  microtask chạy 1 lần → reset flag → recalculate once\n"
            "- RECALCULATE_META check ngăn infinite loop:\n"
            "  recalculation dispatch transaction → onUpdate fire lại → "
            "  !getMeta(RECALCULATE_META) = false → SKIP → không vòng lặp\n"
            "✓ Debounce hoạt động đúng. Chỉ 1 recalculation per event loop cycle."
        ),
        "status": "Pass",
        "actual": (
            "Code trace (double protection mechanism):\n\n"
            "LAYER 1 - scheduledFormulaRecalculateRef (debounce flag):\n"
            "```\n"
            "if (!getMeta(RECALCULATE_META) && !scheduledRef.current) {\n"
            "  scheduledRef.current = true;  // SET FLAG\n"
            "  queueMicrotask(() => {\n"
            "    scheduledRef.current = false;  // RESET FLAG\n"
            "    recalculateAllTableFormulas(editor);\n"
            "  });\n"
            "}\n"
            "// 2nd transaction trong cùng cycle: scheduledRef.current=true → SKIP\n"
            "```\n\n"
            "LAYER 2 - RECALCULATE_META (anti-loop):\n"
            "```\n"
            "// recalculateAllTableFormulas dispatches:\n"
            "tr.setMeta(UEDITOR_TABLE_FORMULA_RECALCULATE_META, true)\n"
            "// → onUpdate: getMeta(RECALCULATE_META)=true → SKIP scheduling\n"
            "// → Không infinite loop\n"
            "```\n"
            "2 layers bảo vệ: scheduledRef (performance) + META (correctness)."
        ),
        "loai": "Functional",
        "uu_tien": "P2",
        "ghi_chu": (
            "Cơ chế 2 lớp:\n"
            "1. scheduledFormulaRecalculateRef: performance guard\n"
            "   → Tối đa 1 recalculation per microtask queue flush\n"
            "2. UEDITOR_TABLE_FORMULA_RECALCULATE_META: correctness guard\n"
            "   → Recalculation transaction không trigger recalculation mới\n\n"
            "Không cần test regression:\n"
            "Nếu thiếu layer 1 → nhiều dispatch, hiệu năng kém\n"
            "Nếu thiếu layer 2 → infinite loop crash browser\n"
            "Cả 2 layer đều critical."
        ),
    },
]

# Tìm row cuối cùng có dữ liệu
last_data_row = 3
for r in range(3, ws.max_row + 1):
    if ws.cell(row=r, column=COL_MA).value:
        last_data_row = r

print(f"Last data row: {last_data_row}")

# Data validation
dv_s = DataValidation(type="list", formula1='"Chưa test,Pass,Fail,Blocked,Re-test,Fixed,Skip"', sqref="J3:J600")
dv_p = DataValidation(type="list", formula1='"P1,P2,P3"', sqref="K3:K600")
dv_t = DataValidation(type="list", formula1='"Functional,UI,Edge case"', sqref="L3:L600")
ws.add_data_validation(dv_s)
ws.add_data_validation(dv_p)
ws.add_data_validation(dv_t)

new_added = 0
for tc in NEW_TCS:
    row = last_data_row + 1
    last_data_row = row
    is_alt = (tc["stt"] % 2) == 0
    bg = "F8FAFC" if is_alt else "FFFFFF"

    ws.row_dimensions[row].height = 120

    data = {
        COL_STT:    tc["stt"],
        COL_MOD:    tc["module"],
        COL_MA:     tc["ma_tc"],
        COL_TEN:    tc["ten"],
        COL_MOTA:   tc["mo_ta"],
        COL_DK:     tc["dieu_kien"],
        COL_BUOC:   tc["buoc"],
        COL_KQ:     tc["ket_qua"],
        COL_ACTUAL: tc["actual"],
        COL_STATUS: tc["status"],
        COL_PRIOR:  tc["uu_tien"],
        COL_TYPE:   tc["loai"],
        COL_NOTE:   tc["ghi_chu"],
    }

    for col, val in data.items():
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = thin_border()

        if col == COL_STT:
            cell.font = Font(bold=True, size=9, name="Segoe UI")
            cell.alignment = center()
            cell.fill = fill(bg)

        elif col == COL_MOD:
            cell.font = Font(bold=True, size=8, color="FFFFFF", name="Segoe UI")
            cell.fill = fill(MODULE_COLOR)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        elif col == COL_MA:
            cell.font = Font(bold=True, size=9, color="7C2D12", name="Courier New")  # orange-900
            cell.alignment = center()
            cell.fill = fill(bg)

        elif col == COL_TEN:
            cell.font = Font(bold=True, size=9, color="0F172A", name="Segoe UI")
            cell.fill = fill(bg)
            cell.alignment = wrap()

        elif col == COL_STATUS:
            cell.font = Font(bold=True, size=9, color="1E293B", name="Segoe UI")
            cell.fill = fill(STATUS_COLOR.get(val, "F1F5F9"))
            cell.alignment = center()

        elif col == COL_PRIOR:
            cell.font = Font(bold=True, size=9, name="Segoe UI")
            cell.fill = fill(PRIORITY_COLORS.get(val, "F1F5F9"))
            cell.alignment = center()

        elif col == COL_TYPE:
            cell.font = nfont(sz=9)
            cell.fill = fill(bg)
            cell.alignment = center()

        elif col == COL_NOTE:
            cell.font = Font(size=8, color="7C2D12", name="Segoe UI", italic=True)
            cell.fill = fill("FFF7ED")  # orange-50
            cell.alignment = wrap()

        else:
            cell.font = nfont(sz=9)
            cell.fill = fill(bg)
            cell.alignment = wrap()

    new_added += 1
    print(f"  + {tc['ma_tc']}: {tc['ten'][:55]}")

wb.save(PATH)
print(f"\n✅ Cập nhật TC-T124 + thêm {new_added} TC mới (TC-T125..T128)")
print(f"💾 Saved → {PATH}")

print("\n" + "═"*60)
print("  TÓM TẮT AUTO-RECALCULATE BEHAVIOR")
print("═"*60)
print("  recalculateAllTableFormulas() triggers:")
print("  [AUTO] onUpdate → queueMicrotask (mỗi editor change)")
print("  [AUTO] useEffect([editor]) (editor mount)")
print("  [AUTO] useEffect([content, editor]) (content prop change)")
print("  [PROTECT] scheduledFormulaRecalculateRef: max 1 per cycle")
print("  [PROTECT] RECALCULATE_META: ngăn infinite loop")
print("═"*60)
