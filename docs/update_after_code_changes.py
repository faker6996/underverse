#!/usr/bin/env python3
"""
Cập nhật test case sau khi dev team thực hiện các thay đổi từ code findings:

1. Xóa export thừa UEDITOR_TABLE_FORMULA_SYNC_META khỏi table-formula-commands.ts
   → Sửa note TC-T092 và TC-T124 (không còn "exported nhưng chưa dùng")

2. Thêm formatTableFormulaDisplayValue() và normalizeTableNumberFormat() vào table-formula.ts
   → computedValue lưu raw value; display content = format(computedValue, numberFormat)
   → numberFormat: "text" | "number" | "currency" | "percent" | "date"
   → Thêm TC-T129, TC-T130, TC-T131

3. Unit tests đã lock behavior (pass 16/16 sau khi thêm)
   → Cập nhật note một số TC để reference unit test

Ngày: 2026-06-07
Xác nhận: npm run typecheck pass, npm run test pass (46/46 + 16/16)
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

PATH = "/Users/tran_van_bach/Desktop/project/nextJs/underverse/docs/UEditor_Table_TestCase.xlsx"

def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def thin_border():
    s = Side(border_style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def nfont(bold=False, sz=9, col="374151", italic=False):
    return Font(bold=bold, size=sz, color=col, name="Segoe UI", italic=italic)

def wrap(h="left", v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)

def center():
    return Alignment(horizontal="center", vertical="center")

STATUS_COLOR = {
    "Pass":    "DCFCE7",
    "Fail":    "FEE2E2",
    "Blocked": "F3E8FF",
    "Skip":    "E2E8F0",
}
PRIORITY_COLORS = {"P1": "FEE2E2", "P2": "FEF9C3", "P3": "DCFCE7"}
MODULE_COLOR = "1C1917"

wb = load_workbook(PATH)
ws = wb["Table Test Cases"]

COL_STT=1; COL_MOD=2; COL_MA=3; COL_TEN=4; COL_MOTA=5
COL_DK=6; COL_BUOC=7; COL_KQ=8; COL_ACTUAL=9; COL_STATUS=10
COL_PRIOR=11; COL_TYPE=12; COL_NOTE=13

# ── Tìm row theo TC ID ────────────────────────────────────────────
def find_row(tc_id):
    for r in range(3, ws.max_row + 1):
        if ws.cell(row=r, column=COL_MA).value == tc_id:
            return r
    return None

# ──────────────────────────────────────────────────────────────────
# PHẦN 1 – CẬP NHẬT NOTE CHO TC-T092 (normalizeFormulaInput)
# SYNC_META đã bị xóa khỏi package source
# ──────────────────────────────────────────────────────────────────
r92 = find_row("TC-T092")
if r92:
    c = ws.cell(row=r92, column=COL_NOTE)
    c.value = (
        "✓ Normalize đúng.\n"
        "✅ FIX (2026-06-07): UEDITOR_TABLE_FORMULA_SYNC_META đã được xóa khỏi package.\n"
        "Trước đây: export const UEDITOR_TABLE_FORMULA_SYNC_META = 'ueditorTableFormulaSync' "
        "tồn tại trong table-formula-commands.ts nhưng không dùng trong package.\n"
        "Sau fix: export bị loại bỏ hoàn toàn. API surface gọn hơn.\n"
        "npm run typecheck: pass sau khi xóa."
    )
    c.font = Font(size=8, color="065F46", name="Segoe UI", italic=True)
    c.alignment = wrap()
    print(f"✓ Cập nhật TC-T092 (row {r92}): note SYNC_META đã bị xóa.")

# ──────────────────────────────────────────────────────────────────
# PHẦN 2 – CẬP NHẬT NOTE CHO TC-T124 (recalculateAllTableFormulas)
# SYNC_META không còn trong note nữa vì đã bị xóa
# ──────────────────────────────────────────────────────────────────
r124 = find_row("TC-T124")
if r124:
    c = ws.cell(row=r124, column=COL_NOTE)
    c.value = (
        "recalculateAllTableFormulas() phù hợp khi:\n"
        "- Load content (tất cả formulas cần recalc từ đầu)\n"
        "- Import/paste nhiều bảng cùng lúc\n"
        "- Collaboration: sync formulas sau conflict resolution\n"
        "Khác recalculateSelectedTable(): chỉ cần cursor trong bảng.\n\n"
        "✅ FIX (2026-06-07): UEDITOR_TABLE_FORMULA_SYNC_META đã bị xóa khỏi package.\n"
        "Chỉ còn UEDITOR_TABLE_FORMULA_RECALCULATE_META (dùng để ngăn infinite loop).\n"
        "Xem TC-T092 ghi chú chi tiết về việc xóa SYNC_META."
    )
    c.font = Font(size=8, color="065F46", name="Segoe UI", italic=True)
    c.alignment = wrap()
    print(f"✓ Cập nhật TC-T124 (row {r124}): note SYNC_META removed.")

# ──────────────────────────────────────────────────────────────────
# PHẦN 3 – TC MỚI (TC-T129, TC-T130, TC-T131)
# ──────────────────────────────────────────────────────────────────
# Tìm last data row
last_data_row = 3
for r in range(3, ws.max_row + 1):
    if ws.cell(row=r, column=COL_MA).value:
        last_data_row = r

print(f"\nLast data row: {last_data_row}")

NEW_TCS = [

    # ──────────────────────────────────────────────────────────────
    # TC-T129: formatTableFormulaDisplayValue — hiển thị có format
    # ──────────────────────────────────────────────────────────────
    {
        "stt": 129,
        "module": "11. Formula (Công thức)",
        "ma_tc": "TC-T129",
        "ten": "formatTableFormulaDisplayValue — hiển thị computed value theo numberFormat của cell",
        "mo_ta": (
            "Thêm mới (2026-06-07): formatTableFormulaDisplayValue(value, numberFormat) "
            "được gọi trong recalculateTableNode() để format display content của formula cell.\n\n"
            "Tách biệt hai khái niệm:\n"
            "• computedValue attr = raw numeric string (ví dụ '1234.5') — dùng cho formula dependency\n"
            "• Cell display content = formatTableFormulaDisplayValue(computedValue, numberFormat) "
            "— dùng để hiển thị cho user\n\n"
            "numberFormat: 'text' | 'number' | 'currency' | 'percent' | 'date'\n"
            "Default ('text'): không format, hiển thị raw value."
        ),
        "dieu_kien": (
            "Bảng 2×2. A1=1234.5 (số). B1 có formula '=A1'.\n"
            "Thử các giá trị numberFormat khác nhau trên B1 (qua attrs hoặc UI)."
        ),
        "buoc": (
            "1. Bảng: A1='1234.5', B1 formula '=A1'\n"
            "2. numberFormat='text' (mặc định) → quan sát B1\n"
            "3. numberFormat='number' → quan sát B1\n"
            "4. numberFormat='currency' → quan sát B1\n"
            "5. numberFormat='percent' → quan sát B1 (A1='0.125' để test)\n"
            "6. numberFormat='date' → quan sát B1 (A1='45658' để test)\n"
            "7. Formula trả về error (#INVALID-REFERENCE) + numberFormat='currency' → quan sát\n"
            "8. Kiểm tra data-computed-value trong DOM (phải là raw value)"
        ),
        "ket_qua": (
            "1. numberFormat='text': B1 hiển thị '1234.5' (raw)\n"
            "2. numberFormat='number': B1 hiển thị '1,234.5' (Intl.NumberFormat en-US)\n"
            "3. numberFormat='currency': B1 hiển thị '$1,234.50' (USD, 2 decimal)\n"
            "4. numberFormat='percent': A1='0.125' → B1 hiển thị '12.5%'\n"
            "5. numberFormat='date': A1='45658' → B1 hiển thị '01/01/2025'\n"
            "   (45658 ngày từ Excel epoch 1899-12-30 = 2025-01-01)\n"
            "6. Error pass-through: '#INVALID-REFERENCE' + currency → hiển thị '#INVALID-REFERENCE'\n"
            "   (stringValue.startsWith('#') → return as-is)\n"
            "7. DOM: data-computed-value='1234.5' (raw), cell text = '$1,234.50' (formatted)\n"
            "   → computedValue phục vụ formula dependency, display phục vụ UX"
        ),
        "status": "Pass",
        "actual": (
            "Code trace (table-formula.ts + table-formula-commands.ts):\n\n"
            "formatTableFormulaDisplayValue(value, numberFormat):\n"
            "  if startsWith('#') → return as-is (error pass-through)\n"
            "  if 'text' → return String(value)\n"
            "  numericValue = parseFloat(value)\n"
            "  if !isFinite → return String(value) (non-numeric fallback)\n"
            "  'number' → Intl.NumberFormat('en-US', {maxFractionDigits:6}).format(n)\n"
            "  'currency' → Intl.NumberFormat('en-US', {style:'currency',currency:'USD',maxFractionDigits:2})\n"
            "  'percent' → Intl.NumberFormat('en-US', {style:'percent',maxFractionDigits:2})\n"
            "  'date' → epoch = Date.UTC(1899,11,30) + n*86400000\n"
            "           → Intl.DateTimeFormat('en-US', {year,month,day,timeZone:'UTC'})\n\n"
            "Unit test (ueditor-table-formula.test.mjs) xác nhận:\n"
            "  formatTableFormulaDisplayValue('1234.5', 'number') = '1,234.5'\n"
            "  formatTableFormulaDisplayValue('1234.5', 'currency') = '$1,234.50'\n"
            "  formatTableFormulaDisplayValue('0.125', 'percent') = '12.5%'\n"
            "  formatTableFormulaDisplayValue('45658', 'date') = '01/01/2025'\n"
            "  formatTableFormulaDisplayValue('#INVALID-REFERENCE', 'currency') = '#INVALID-REFERENCE'"
        ),
        "loai": "Functional",
        "uu_tien": "P1",
        "ghi_chu": (
            "Thiết kế quan trọng:\n"
            "computedValue (raw) ≠ display content (formatted)\n"
            "→ Formula dependency dùng raw, người dùng thấy formatted.\n\n"
            "Unit test đã lock behavior trong:\n"
            "packages/underverse/tests/ueditor-table-formula.test.mjs\n"
            "'UEditor table formula utilities format computed display values'\n\n"
            "Locale: 'en-US' hardcoded → '$1,234.50' không phải '1.234,50'.\n"
            "Nếu cần i18n, cần cập nhật hàm để nhận locale."
        ),
    },

    # ──────────────────────────────────────────────────────────────
    # TC-T130: numberFormat attr — raw vs display tách biệt
    # ──────────────────────────────────────────────────────────────
    {
        "stt": 130,
        "module": "11. Formula (Công thức)",
        "ma_tc": "TC-T130",
        "ten": "computedValue attr giữ raw value — display content được format theo numberFormat",
        "mo_ta": (
            "Kiểm tra tính tách biệt giữa data layer và display layer:\n"
            "• computedValue attr (data-computed-value trong DOM) = raw numeric string\n"
            "• Cell content text = formatTableFormulaDisplayValue(computedValue, numberFormat)\n\n"
            "Khi cell B2 tham chiếu B1 (formula cell có numberFormat='currency'):\n"
            "B2 phải dùng raw value từ computedValue của B1, không phải display '$1,234.50'."
        ),
        "dieu_kien": (
            "Bảng 3 hàng. A1=1000. "
            "B1 formula '=A1', numberFormat='currency'. "
            "B2 formula '=B1*2' (tham chiếu B1 formatted)."
        ),
        "buoc": (
            "1. A1=1000, B1='=A1' với numberFormat='currency'\n"
            "2. Verify B1 display = '$1,000.00', data-computed-value='1000'\n"
            "3. B2='=B1*2'\n"
            "4. Quan sát B2\n"
            "5. Kiểm tra DOM: data-computed-value của B2\n"
            "6. Đổi A1=2000 → quan sát B1 và B2 update đồng bộ"
        ),
        "ket_qua": (
            "- B1 display = '$1,000.00', data-computed-value = '1000'\n"
            "- B2 = '=B1*2': getCellValue('B1') trả '1000' (từ computedValue raw)\n"
            "  → 1000 * 2 = 2000 → B2 data-computed-value = '2000'\n"
            "  → B2 display (numberFormat='text' mặc định) = '2000'\n"
            "- Sau A1=2000: B1 = '$2,000.00', B2 = '4000'\n"
            "✓ Formula chain dùng raw values, không bị lẫn formatted strings.\n"
            "✓ '$1,000.00' không được parseFloat là 1000 (dấu '$' gây NaN),\n"
            "  nhưng không sao vì getCellValue() lấy từ computedValue attr = '1000'."
        ),
        "status": "Pass",
        "actual": (
            "Code trace (recalculateTableNode + buildTableValueMap):\n\n"
            "buildTableValueMap() (dùng cho getCellValue):\n"
            "  const computedValue = entry.node.attrs.computedValue;\n"
            "  values.set(label, computedValue.trim() ? computedValue : getCellText(entry.node));\n"
            "  → Luôn dùng raw computedValue, không bao giờ dùng display text.\n\n"
            "recalculateTableNode():\n"
            "  computedValue = String(result.value)          // raw '1000'\n"
            "  displayValue = formatTableFormulaDisplayValue(computedValue, numberFormat) // '$1,000.00'\n"
            "  cell.attrs.computedValue = computedValue      // raw\n"
            "  cell content = createCellDisplayContent(displayValue) // formatted\n\n"
            "→ data layer (computedValue) và display layer tách biệt hoàn toàn."
        ),
        "loai": "Functional",
        "uu_tien": "P1",
        "ghi_chu": (
            "Invariant quan trọng:\n"
            "getCellValue() luôn đọc attrs.computedValue (raw), không đọc DOM text.\n"
            "→ Formatting không ảnh hưởng đến formula dependency chain.\n"
            "Nếu vi phạm invariant này: SUM($1,000.00:$2,000.00) → parseFloat('$1,000.00') = NaN → #INVALID-REFERENCE.\n"
            "Hiện tại code đúng: buildTableValueMap dùng attrs.computedValue."
        ),
    },

    # ──────────────────────────────────────────────────────────────
    # TC-T131: normalizeTableNumberFormat — validation
    # ──────────────────────────────────────────────────────────────
    {
        "stt": 131,
        "module": "11. Formula (Công thức)",
        "ma_tc": "TC-T131",
        "ten": "normalizeTableNumberFormat — validate và fallback về 'text' cho format không hợp lệ",
        "mo_ta": (
            "normalizeTableNumberFormat(format) validate giá trị numberFormat attr.\n"
            "Chỉ chấp nhận: 'text' | 'number' | 'currency' | 'percent' | 'date'.\n"
            "Mọi giá trị khác (null, undefined, '', 'foo', số) → fallback về 'text'.\n\n"
            "Hàm này được gọi trong formatTableFormulaDisplayValue() trước khi format."
        ),
        "dieu_kien": "Formula cell với numberFormat attr được set các giá trị khác nhau.",
        "buoc": (
            "1. numberFormat='text' → normalizeTableNumberFormat trả 'text'\n"
            "2. numberFormat='number' → trả 'number'\n"
            "3. numberFormat='currency' → trả 'currency'\n"
            "4. numberFormat='percent' → trả 'percent'\n"
            "5. numberFormat='date' → trả 'date'\n"
            "6. numberFormat=null → trả 'text' (fallback)\n"
            "7. numberFormat=undefined → trả 'text'\n"
            "8. numberFormat='invalid' → trả 'text'\n"
            "9. numberFormat=123 (số) → trả 'text'\n"
            "10. numberFormat='' → trả 'text'"
        ),
        "ket_qua": (
            "- Các giá trị hợp lệ ('text','number','currency','percent','date') → trả đúng\n"
            "- Mọi giá trị không hợp lệ → fallback 'text' → display raw value\n"
            "- Không throw error với bất kỳ input nào\n"
            "- formatTableFormulaDisplayValue với format không hợp lệ: display raw string\n"
            "✓ Defensive programming: cell attrs corrupt không gây crash"
        ),
        "status": "Pass",
        "actual": (
            "Code trace (table-formula.ts):\n\n"
            "export type TableNumberFormat = 'text' | 'number' | 'currency' | 'percent' | 'date';\n\n"
            "export function normalizeTableNumberFormat(format: unknown): TableNumberFormat {\n"
            "  return format === 'text' || format === 'number' || format === 'currency'\n"
            "      || format === 'percent' || format === 'date'\n"
            "    ? format\n"
            "    : 'text';  // fallback cho mọi giá trị khác\n"
            "}\n\n"
            "Trong formatTableFormulaDisplayValue():\n"
            "  const normalizedFormat = normalizeTableNumberFormat(numberFormat);\n"
            "  if (normalizedFormat === 'text') return stringValue;\n"
            "  // ... format theo normalized value\n\n"
            "Unit tests pass xác nhận behavior đúng."
        ),
        "loai": "Functional",
        "uu_tien": "P3",
        "ghi_chu": (
            "Defensive: attrs từ paste/import HTML có thể có giá trị bất kỳ.\n"
            "normalizeTableNumberFormat đảm bảo không crash với input bất thường.\n"
            "Default 'text' là safe choice: hiển thị raw value, không mất thông tin."
        ),
    },
]

# Data validation
dv_s = DataValidation(type="list", formula1='"Chưa test,Pass,Fail,Blocked,Re-test,Fixed,Skip"', sqref="J3:J700")
dv_p = DataValidation(type="list", formula1='"P1,P2,P3"', sqref="K3:K700")
dv_t = DataValidation(type="list", formula1='"Functional,UI,Edge case"', sqref="L3:L700")
ws.add_data_validation(dv_s)
ws.add_data_validation(dv_p)
ws.add_data_validation(dv_t)

new_added = 0
for tc in NEW_TCS:
    row = last_data_row + 1
    last_data_row = row
    bg = "F8FAFC" if (tc["stt"] % 2 == 0) else "FFFFFF"

    ws.row_dimensions[row].height = 115

    data = {
        COL_STT:    tc["stt"],
        COL_MOD:    tc["module"],
        COL_MA:     tc["ma_tc"],
        COL_TEN:    tc["ten"],
        COL_MOTA:   tc["mo_ta"],
        COL_DK:     tc["dieu_kien"],
        COL_BUOC:   tc["buoc"],
        COL_KQ:     tc["ket_qua"],
        COL_ACTUAL: tc["actual"],
        COL_STATUS: tc["status"],
        COL_PRIOR:  tc["uu_tien"],
        COL_TYPE:   tc["loai"],
        COL_NOTE:   tc["ghi_chu"],
    }

    for col, val in data.items():
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = thin_border()

        if col == COL_STT:
            cell.font = Font(bold=True, size=9, name="Segoe UI")
            cell.alignment = center()
            cell.fill = fill(bg)
        elif col == COL_MOD:
            cell.font = Font(bold=True, size=8, color="FFFFFF", name="Segoe UI")
            cell.fill = fill(MODULE_COLOR)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        elif col == COL_MA:
            cell.font = Font(bold=True, size=9, color="065F46", name="Courier New")  # green = new after fix
            cell.alignment = center()
            cell.fill = fill(bg)
        elif col == COL_TEN:
            cell.font = Font(bold=True, size=9, color="0F172A", name="Segoe UI")
            cell.fill = fill(bg)
            cell.alignment = wrap()
        elif col == COL_STATUS:
            cell.font = Font(bold=True, size=9, color="1E293B", name="Segoe UI")
            cell.fill = fill(STATUS_COLOR.get(val, "F1F5F9"))
            cell.alignment = center()
        elif col == COL_PRIOR:
            cell.font = Font(bold=True, size=9, name="Segoe UI")
            cell.fill = fill(PRIORITY_COLORS.get(val, "F1F5F9"))
            cell.alignment = center()
        elif col == COL_TYPE:
            cell.font = nfont(sz=9)
            cell.fill = fill(bg)
            cell.alignment = center()
        elif col == COL_NOTE:
            cell.font = Font(size=8, color="065F46", name="Segoe UI", italic=True)
            cell.fill = fill("F0FDF4")  # green-50 = post-fix TC
            cell.alignment = wrap()
        else:
            cell.font = nfont(sz=9)
            cell.fill = fill(bg)
            cell.alignment = wrap()

    new_added += 1
    print(f"  + {tc['ma_tc']}: {tc['ten'][:58]}")

wb.save(PATH)
print(f"\n✅ Cập nhật note TC-T092, TC-T124")
print(f"✅ Thêm {new_added} TC mới (TC-T129, TC-T130, TC-T131)")
print(f"💾 Saved → {PATH}")

print("\n" + "═"*60)
print("  TÓM TẮT CẬP NHẬT SAU CODE CHANGES")
print("═"*60)
print("  [XÓA] UEDITOR_TABLE_FORMULA_SYNC_META removed from package")
print("  [MỚI] formatTableFormulaDisplayValue(value, numberFormat)")
print("         text    → raw string")
print("         number  → 1,234.5 (Intl.NumberFormat en-US)")
print("         currency→ $1,234.50 (USD)")
print("         percent → 12.5%")
print("         date    → 01/01/2025 (Excel epoch)")
print("  [MỚI] normalizeTableNumberFormat() → fallback về 'text'")
print("  [UNIT] ueditor-table-formula.test.mjs pass 16/16")
print("  [LOCK] behavior text ref → #INVALID-REFERENCE locked by test")
print("═"*60)
