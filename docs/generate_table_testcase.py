#!/usr/bin/env python3
"""
Script tạo file test case Excel chuyên sâu cho Table trong UEditor
Chạy: python3 docs/generate_table_testcase.py
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from collections import Counter

# ──────────────────────────────────────────────
# COLOR PALETTE
# ──────────────────────────────────────────────
C_HEADER_BG  = "0F172A"
C_HEADER_FG  = "FFFFFF"
C_WHITE      = "FFFFFF"
C_ALT_ROW    = "F8FAFC"
C_BORDER     = "CBD5E1"

MODULE_COLORS = {
    "01. Tạo bảng":               "1E3A5F",
    "02. Điều hướng":              "14532D",
    "03. Thao tác Hàng":          "4C1D95",
    "04. Thao tác Cột":           "7C2D12",
    "05. Add Rails (Mở rộng)":    "065F46",
    "06. Merge & Split Ô":        "1E40AF",
    "07. Resize Cột":             "92400E",
    "08. Header Row / Column":    "831843",
    "09. Căn lề Bảng":            "134E4A",
    "10. Màu & Style Ô":          "312E81",
    "11. Formula (Công thức)":    "1C1917",
    "12. Clipboard - HTML Table": "0C4A6E",
    "13. Clipboard - TSV":        "3F3F46",
    "14. Keyboard Navigation":    "1F2937",
    "15. Undo / Redo":            "374151",
    "16. Read-only Mode":         "4B5563",
    "17. prepareContentForSave":  "374151",
    "18. Edge Cases":             "1E293B",
}

STATUS_COLORS = {
    "Chưa test": "F1F5F9",
    "Pass":      "DCFCE7",
    "Fail":      "FEE2E2",
    "Blocked":   "F3E8FF",
    "Re-test":   "DBEAFE",
    "Fixed":     "FEF9C3",
    "Skip":      "F1F5F9",
}
PRIORITY_COLORS = {"P1": "FEE2E2", "P2": "FEF9C3", "P3": "DCFCE7"}

def fill(hex_c): return PatternFill("solid", fgColor=hex_c)
def bfont(sz=9, col="000000"): return Font(bold=True, size=sz, color=col, name="Segoe UI")
def nfont(sz=9, col="1E293B"): return Font(size=sz, color=col, name="Segoe UI")
def border():
    s = Side(border_style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)
def wrap(h="left", v="top"): return Alignment(horizontal=h, vertical=v, wrap_text=True)

# ──────────────────────────────────────────────
# COLUMNS
# ──────────────────────────────────────────────
COLS = [
    ("STT",                   4),
    ("Module",               18),
    ("Mã TC",                10),
    ("Tên Test Case",        36),
    ("Mô tả",                42),
    ("Điều kiện\ntiên quyết", 28),
    ("Các bước thực hiện",   58),
    ("Kết quả mong muốn",    54),
    ("Kết quả thực tế",      32),
    ("Status",               12),
    ("Ưu tiên",               9),
    ("Loại test",            13),
    ("Ghi chú",              22),
]

# ──────────────────────────────────────────────
# TEST CASES  (stt, module, ma_tc, ten, mo_ta, dieu_kien, buoc, ket_qua, loai, uu_tien)
# ──────────────────────────────────────────────
TCS = [

  # ══════════════════════════════════════════════
  # 01. TẠO BẢNG
  # ══════════════════════════════════════════════
  (1, "01. Tạo bảng", "TC-T001",
   "Tạo bảng 3×3 qua TableInsertGrid với header row",
   "Kiểm tra insert bảng cơ bản qua visual grid picker trên toolbar, đảm bảo header row được tạo.",
   "Editor đang focus, toolbar hiển thị.",
   "1. Click nút Table (⊞) trên toolbar\n2. Trong grid picker di chuột đến ô hàng 3 cột 3\n3. Click chuột trái",
   "- Bảng 3 hàng × 3 cột xuất hiện trong editor\n- Hàng đầu tiên là <th> (header)\n- 2 hàng còn lại là <td> (body)\n- HTML: <table class='border-collapse my-4'>...\n- Cursor tự động đặt vào ô đầu tiên (A1)",
   "Functional", "P1"),

  (2, "01. Tạo bảng", "TC-T002",
   "Tạo bảng kích thước tối đa 8×8 qua InsertGrid",
   "Kiểm tra giới hạn grid picker là 8×8.",
   "Toolbar Table dropdown đang mở.",
   "1. Di chuột đến ô hàng 8 cột 8 trong grid picker\n2. Click",
   "- Grid picker cho phép chọn tối đa 8 hàng × 8 cột\n- Bảng 8×8 được tạo thành công\n- Không có ô nào vượt ngoài lưới 8×8",
   "Functional", "P2"),

  (3, "01. Tạo bảng", "TC-T003",
   "Grid picker highlight đúng số hàng/cột khi hover",
   "Kiểm tra visual feedback của InsertGrid khi di chuột.",
   "Grid picker dropdown đang mở.",
   "1. Di chuột từng ô trong grid (ví dụ 2×4, 5×3)\n2. Quan sát màu highlight và label preview",
   "- Các ô từ (1,1) đến (row,col) được highlight màu primary/20\n- Label phía trên hiển thị đúng: '3 hàng × 4 cột' (theo ngôn ngữ hiện tại)\n- Highlight cập nhật real-time theo vị trí chuột",
   "UI", "P2"),

  (4, "01. Tạo bảng", "TC-T004",
   "Tạo bảng qua Slash Command /table",
   "Kiểm tra insert bảng 3×3 qua slash command.",
   "Editor focus, dòng trống.",
   "1. Gõ '/table'\n2. Chọn 'Table' trong danh sách\n3. Nhấn Enter",
   "- Bảng 3×3 với withHeaderRow=true được tạo\n- Text '/table' bị xóa\n- Menu slash command đóng lại\n- Cursor vào ô đầu tiên",
   "Functional", "P1"),

  (5, "01. Tạo bảng", "TC-T005",
   "Tạo bảng qua Menu Bar (Insert > Table)",
   "Kiểm tra insert bảng từ menu bar.",
   "showMenuBar=true, editor focus.",
   "1. Click menu 'Insert'\n2. Click 'Table'\n3. Quan sát kết quả",
   "- Bảng được tạo trong editor\n- Hoặc: mở dialog cho phép chọn kích thước trước khi tạo",
   "Functional", "P2"),

  (6, "01. Tạo bảng", "TC-T006",
   "Tạo nhiều bảng trong cùng document",
   "Kiểm tra document có thể chứa nhiều bảng độc lập.",
   "Editor trống.",
   "1. Tạo bảng 2×2\n2. Nhấn Enter để ra ngoài bảng\n3. Tạo thêm bảng 3×3\n4. Kiểm tra 2 bảng không ảnh hưởng nhau",
   "- 2 bảng tồn tại độc lập trong document\n- Click vào bảng 1: controls chỉ hiện cho bảng 1\n- Click vào bảng 2: controls chỉ hiện cho bảng 2\n- HTML output chứa 2 thẻ <table> riêng biệt",
   "Functional", "P2"),

  (7, "01. Tạo bảng", "TC-T007",
   "Không tạo bảng trong chế độ read-only",
   "Kiểm tra editable=false ngăn tạo bảng.",
   "editable=false.",
   "1. Mount <UEditor editable={false} />\n2. Kiểm tra toolbar không hiển thị\n3. Gõ '/table'\n4. Quan sát kết quả",
   "- Toolbar bị ẩn, không có nút Table\n- Slash command không hoạt động (editor không editable)\n- Không có bảng nào được tạo",
   "Functional", "P1"),

  # ══════════════════════════════════════════════
  # 02. ĐIỀU HƯỚNG
  # ══════════════════════════════════════════════
  (8, "02. Điều hướng", "TC-T008",
   "Tab di chuyển từ ô này sang ô kế tiếp",
   "Kiểm tra Tab navigation trong bảng.",
   "Bảng đã tạo, cursor đang trong ô A1.",
   "1. Nhấn Tab từ ô A1\n2. Nhấn Tab từ ô A2\n3. Tiếp tục nhấn Tab",
   "- Tab: di chuyển sang ô tiếp theo theo chiều ngang\n- Tab ở ô cuối hàng: xuống hàng tiếp theo, ô đầu tiên\n- Cursor luôn visible trong ô được focus",
   "Functional", "P1"),

  (9, "02. Điều hướng", "TC-T009",
   "Shift+Tab di chuyển ngược về ô trước",
   "Kiểm tra Shift+Tab navigation.",
   "Cursor trong ô bảng.",
   "1. Nhấn Shift+Tab từ ô B2\n2. Tiếp tục nhấn Shift+Tab",
   "- Shift+Tab: di chuyển về ô phía trước\n- Shift+Tab ở ô đầu hàng: lên hàng trước, ô cuối cùng",
   "Functional", "P2"),

  (10, "02. Điều hướng", "TC-T010",
   "Tab từ ô cuối cùng tự động tạo hàng mới",
   "Kiểm tra auto-expand bảng khi Tab ở ô cuối.",
   "Bảng 2×2, cursor ở ô cuối cùng (B2).",
   "1. Nhấn Tab từ ô cuối cùng của bảng",
   "- Hàng mới được thêm tự động\n- Cursor di chuyển vào ô đầu tiên của hàng mới\n- Số hàng tăng thêm 1",
   "Functional", "P1"),

  (11, "02. Điều hướng", "TC-T011",
   "Phím mũi tên không thoát khỏi bảng",
   "Kiểm tra arrow keys bị stopPropagation trong bảng.",
   "Cursor trong ô bảng.",
   "1. Nhấn ArrowLeft, ArrowRight, ArrowUp, ArrowDown\n2. Quan sát hành vi",
   "- Mũi tên di chuyển trong nội dung ô bình thường\n- Không bubble lên document (stopPropagation được gọi)\n- Event không gây scroll page không mong muốn",
   "Functional", "P2"),

  (12, "02. Điều hướng", "TC-T012",
   "Click vào ô để focus và hiện table controls",
   "Kiểm tra mouse click trigger controls.",
   "Bảng đã tạo.",
   "1. Click chuột vào một ô bất kỳ\n2. Quan sát toolbar và controls",
   "- Cursor đặt vào trong ô được click\n- Toolbar hiện thêm nhóm nút table (←↑↓→ và merge/split)\n- Table control menu (⊞) xuất hiện góc trên trái bảng\n- Row/column handles hiện ở rìa bảng",
   "UI", "P1"),

  (13, "02. Điều hướng", "TC-T013",
   "Hover chuột vào ô highlight row/column handle",
   "Kiểm tra hover state highlight.",
   "Bảng đã tạo, không có ô nào đang focus.",
   "1. Di chuột vào một ô trong bảng\n2. Quan sát row handle và column handle",
   "- Row handle tương ứng với hàng đang hover được highlight\n- Column handle tương ứng với cột đang hover được highlight\n- Khi rời chuột: highlight biến mất",
   "UI", "P2"),

  # ══════════════════════════════════════════════
  # 03. THAO TÁC HÀNG
  # ══════════════════════════════════════════════
  (14, "03. Thao tác Hàng", "TC-T014",
   "Thêm hàng trước (addRowBefore) qua toolbar dropdown",
   "Kiểm tra insert row before qua toolbar.",
   "Cursor trong hàng giữa của bảng 3×3.",
   "1. Click vào ô ở hàng 2\n2. Mở Table dropdown\n3. Click 'Thêm hàng trước'",
   "- Hàng mới rỗng xuất hiện phía TRÊN hàng hiện tại\n- Tổng số hàng tăng từ 3 → 4\n- Dữ liệu các hàng khác không thay đổi\n- Nút disabled khi không có cursor trong bảng",
   "Functional", "P1"),

  (15, "03. Thao tác Hàng", "TC-T015",
   "Thêm hàng sau (addRowAfter) qua toolbar dropdown",
   "Kiểm tra insert row after qua toolbar.",
   "Cursor trong hàng giữa của bảng.",
   "1. Click ô ở hàng 2\n2. Mở Table dropdown\n3. Click 'Thêm hàng sau'",
   "- Hàng mới rỗng xuất hiện phía DƯỚI hàng hiện tại\n- Tổng số hàng tăng lên 1",
   "Functional", "P1"),

  (16, "03. Thao tác Hàng", "TC-T016",
   "Thêm hàng trước qua Table Control Menu (⊞)",
   "Kiểm tra addRowBefore từ floating table menu.",
   "Cursor trong bảng, menu ⊞ đang hiển thị.",
   "1. Click nút ⊞ góc trên trái bảng\n2. Click 'Thêm hàng trước'",
   "- Hàng mới được thêm phía trên hàng hiện tại\n- Menu đóng sau khi thực hiện",
   "Functional", "P2"),

  (17, "03. Thao tác Hàng", "TC-T017",
   "Thêm hàng sau qua Table Control Menu",
   "Kiểm tra addRowAfter từ floating table menu.",
   "Cursor trong bảng.",
   "1. Click ⊞ > 'Thêm hàng sau'",
   "- Hàng mới được thêm phía dưới hàng hiện tại",
   "Functional", "P2"),

  (18, "03. Thao tác Hàng", "TC-T018",
   "Thêm hàng trước/sau qua Row Handle Menu",
   "Kiểm tra context menu từ row handle.",
   "Bảng đã tạo, row handle hiển thị ở trái.",
   "1. Hover vào row handle (vạch dọc trái bảng) của hàng 2\n2. Click icon ⋮ hoặc right-click\n3. Chọn 'Thêm hàng trước'\n4. Chọn 'Thêm hàng sau'",
   "- Hàng mới được thêm tại đúng vị trí\n- Context menu đóng sau khi chọn\n- Cursor đặt vào hàng mới",
   "Functional", "P1"),

  (19, "03. Thao tác Hàng", "TC-T019",
   "Xóa hàng (deleteRow) qua toolbar dropdown",
   "Kiểm tra delete row qua toolbar.",
   "Bảng ≥ 2 hàng body (không tính header), cursor trong bảng.",
   "1. Click vào ô ở hàng 2\n2. Mở Table dropdown\n3. Click 'Xóa hàng'",
   "- Hàng hiện tại bị xóa hoàn toàn\n- Nội dung hàng đó mất\n- Tổng số hàng giảm đi 1\n- Cursor chuyển sang hàng kề",
   "Functional", "P1"),

  (20, "03. Thao tác Hàng", "TC-T020",
   "Xóa hàng qua Row Handle Menu",
   "Kiểm tra delete row từ row handle context menu.",
   "Bảng ≥ 2 hàng.",
   "1. Hover vào row handle\n2. Mở context menu\n3. Click 'Xóa hàng' (màu đỏ/destructive)",
   "- Hàng được xóa\n- Nút 'Xóa hàng' có style destructive (màu đỏ)\n- Confirm không cần thiết nhưng undo vẫn được",
   "Functional", "P1"),

  (21, "03. Thao tác Hàng", "TC-T021",
   "Duplicate hàng qua Row Handle Menu",
   "Kiểm tra sao chép hàng.",
   "Bảng có dữ liệu.",
   "1. Hover vào row handle của hàng có data\n2. Mở context menu\n3. Click 'Nhân đôi hàng' (Duplicate Row)",
   "- Một hàng mới giống hệt hàng gốc được thêm phía DƯỚI\n- Nội dung (text, style) được copy chính xác\n- Không ảnh hưởng các hàng khác",
   "Functional", "P1"),

  (22, "03. Thao tác Hàng", "TC-T022",
   "Clear nội dung hàng qua Row Handle Menu",
   "Kiểm tra xóa nội dung mà không xóa hàng.",
   "Hàng có dữ liệu trong các ô.",
   "1. Mở row handle context menu\n2. Click 'Xóa nội dung hàng' (Clear Row Contents)",
   "- Tất cả ô trong hàng trở nên rỗng\n- Hàng vẫn tồn tại (không bị xóa)\n- Style/attrs ô vẫn giữ nguyên",
   "Functional", "P2"),

  (23, "03. Thao tác Hàng", "TC-T023",
   "Kéo thả di chuyển hàng (Drag Row)",
   "Kiểm tra reorder hàng bằng drag & drop.",
   "Bảng ≥ 3 hàng.",
   "1. Hover vào row handle để hiện drag handle\n2. Mousedown để bắt đầu drag\n3. Kéo lên/xuống\n4. Thả chuột ở vị trí mới",
   "- Drag preview hiển thị hàng đang được kéo\n- Cursor thay đổi thành 'grabbing'\n- Sau khi thả: hàng dịch chuyển đúng vị trí target\n- moveTableRow được gọi với from/to đúng",
   "Functional", "P1"),

  (24, "03. Thao tác Hàng", "TC-T024",
   "Drag row về đúng vị trí khi cancel",
   "Kiểm tra drag cancel khi thả ở vị trí gốc.",
   "Bảng đang trong trạng thái drag hàng.",
   "1. Bắt đầu drag hàng 2\n2. Kéo sang hàng 2 (vị trí gốc)\n3. Thả chuột",
   "- Thứ tự hàng không thay đổi (originIndex === targetIndex)\n- moveTableRow không được gọi\n- Preview biến mất",
   "Functional", "P2"),

  (25, "03. Thao tác Hàng", "TC-T025",
   "Thêm nhiều hàng liên tiếp để kiểm tra index",
   "Kiểm tra thứ tự hàng đúng khi thêm nhiều lần.",
   "Bảng 2×2 (1 header + 1 body).",
   "1. Thêm hàng sau (addRowAfter) 3 lần liên tiếp\n2. Kiểm tra thứ tự các hàng",
   "- 4 hàng body được tạo đúng thứ tự\n- Không có hàng bị skip hay duplicate\n- Cursor không bị reset về đầu",
   "Functional", "P2"),

  (26, "03. Thao tác Hàng", "TC-T026",
   "Disabled deleteRow khi bảng chỉ có 1 hàng",
   "Kiểm tra nút xóa bị disabled đúng.",
   "Bảng chỉ có header row (đã xóa hết body rows).",
   "1. Tạo bảng 1×3 (chỉ header)\n2. Thử click 'Xóa hàng'",
   "- Nút 'Xóa hàng' bị disabled (editor.can().deleteRow() = false)\n- Không có hàng nào bị xóa",
   "Functional", "P2"),

  # ══════════════════════════════════════════════
  # 04. THAO TÁC CỘT
  # ══════════════════════════════════════════════
  (27, "04. Thao tác Cột", "TC-T027",
   "Thêm cột trước (addColumnBefore) qua toolbar",
   "Kiểm tra insert column before.",
   "Cursor trong cột giữa của bảng 3×3.",
   "1. Click ô cột B\n2. Mở Table dropdown\n3. Click 'Thêm cột trước'",
   "- Cột mới rỗng xuất hiện TRƯỚC cột hiện tại (bên trái)\n- Tổng số cột tăng từ 3 → 4\n- Dữ liệu các cột khác không đổi",
   "Functional", "P1"),

  (28, "04. Thao tác Cột", "TC-T028",
   "Thêm cột sau (addColumnAfter) qua toolbar",
   "Kiểm tra insert column after.",
   "Cursor trong bảng.",
   "1. Click ô cột B\n2. Mở Table dropdown\n3. Click 'Thêm cột sau'",
   "- Cột mới xuất hiện SAU cột hiện tại (bên phải)\n- Tổng số cột tăng thêm 1",
   "Functional", "P1"),

  (29, "04. Thao tác Cột", "TC-T029",
   "Thêm cột trước/sau qua Column Handle Menu",
   "Kiểm tra context menu từ column handle.",
   "Bảng đã tạo, column handle hiển thị phía trên.",
   "1. Hover vào column handle phía trên cột B\n2. Mở context menu\n3. Chọn 'Thêm cột trước'\n4. Chọn 'Thêm cột sau'",
   "- Cột mới được thêm tại đúng vị trí tương ứng\n- Context menu đóng sau khi chọn",
   "Functional", "P1"),

  (30, "04. Thao tác Cột", "TC-T030",
   "Xóa cột (deleteColumn) qua toolbar",
   "Kiểm tra delete column qua toolbar.",
   "Bảng ≥ 2 cột, cursor trong bảng.",
   "1. Click ô cột B\n2. Mở Table dropdown\n3. Click 'Xóa cột'",
   "- Cột B bị xóa hoàn toàn\n- Tất cả ô trong cột đó mất\n- Số cột giảm đi 1\n- Cursor chuyển sang cột kề",
   "Functional", "P1"),

  (31, "04. Thao tác Cột", "TC-T031",
   "Xóa cột qua Column Handle Menu",
   "Kiểm tra deleteColumn từ column handle.",
   "Column handle hiển thị.",
   "1. Mở column handle context menu\n2. Click 'Xóa cột' (destructive)",
   "- Cột được xóa\n- 'Xóa cột' có style destructive",
   "Functional", "P1"),

  (32, "04. Thao tác Cột", "TC-T032",
   "Duplicate cột qua Column Handle Menu",
   "Kiểm tra sao chép cột.",
   "Cột có dữ liệu.",
   "1. Hover vào column handle của cột có data\n2. Mở context menu\n3. Click 'Nhân đôi cột'",
   "- Cột mới giống hệt cột gốc được thêm sang phải\n- Text content và attrs được copy\n- Không dùng createEmptyCellNode mà dùng createCellCopyForColumnDuplicate",
   "Functional", "P1"),

  (33, "04. Thao tác Cột", "TC-T033",
   "Clear nội dung cột qua Column Handle Menu",
   "Kiểm tra xóa data trong cột.",
   "Cột có dữ liệu.",
   "1. Mở column handle context menu\n2. Click 'Xóa nội dung cột'",
   "- Tất cả ô trong cột trở nên rỗng\n- Cột vẫn tồn tại\n- Attrs (backgroundColor...) vẫn giữ nguyên",
   "Functional", "P2"),

  (34, "04. Thao tác Cột", "TC-T034",
   "Kéo thả di chuyển cột (Drag Column)",
   "Kiểm tra reorder cột bằng drag & drop.",
   "Bảng ≥ 3 cột.",
   "1. Hover vào column handle để hiện drag\n2. Mousedown bắt đầu drag\n3. Kéo ngang sang cột khác\n4. Thả chuột",
   "- Drag preview hiển thị cột đang được kéo (dạng dọc)\n- Cursor 'grabbing' trong khi drag\n- Sau khi thả: cột di chuyển đúng vị trí\n- moveTableColumn được gọi",
   "Functional", "P1"),

  (35, "04. Thao tác Cột", "TC-T035",
   "Thêm cột qua Quick Action Buttons trên toolbar context",
   "Kiểm tra quick buttons ← → trong toolbar khi focus vào bảng.",
   "Cursor trong bảng.",
   "1. Click vào ô trong bảng\n2. Quan sát toolbar có thêm nhóm nút ← → ↑ ↓\n3. Click nút ← (Thêm cột trước)\n4. Click nút → (Thêm cột sau)",
   "- Nhóm nút context xuất hiện khi hasTableContext=true\n- Click ← thêm cột trước đúng\n- Click → thêm cột sau đúng",
   "UI", "P1"),

  (36, "04. Thao tác Cột", "TC-T036",
   "Disabled addColumnBefore khi không trong bảng",
   "Kiểm tra nút disabled đúng.",
   "Cursor KHÔNG trong bảng (đang ở paragraph thường).",
   "1. Click vào vùng text thường\n2. Quan sát toolbar table buttons",
   "- Nút 'Thêm cột trước/sau' bị disabled hoặc ẩn\n- editor.can().addColumnBefore() = false",
   "Functional", "P2"),

  # ══════════════════════════════════════════════
  # 05. ADD RAILS (MỞ RỘNG)
  # ══════════════════════════════════════════════
  (37, "05. Add Rails (Mở rộng)", "TC-T037",
   "Hover phía dưới bảng hiện nút '+' thêm hàng",
   "Kiểm tra Add Row Rail hiển thị khi hover.",
   "Bảng đã tạo, cursor đang trong bảng.",
   "1. Di chuột xuống phía dưới bảng (sát border cuối)\n2. Quan sát UI",
   "- Thanh ngang '+' xuất hiện bên dưới bảng\n- Tooltip 'Thêm hàng' (quickAddRowLabel)\n- Width = chiều rộng bảng",
   "UI", "P2"),

  (38, "05. Add Rails (Mở rộng)", "TC-T038",
   "Click nút '+' dưới bảng thêm 1 hàng mới",
   "Kiểm tra quick add row bằng 1 click.",
   "Add row rail đang hiển thị.",
   "1. Click nút '+' dưới bảng",
   "- 1 hàng mới rỗng được thêm vào cuối bảng\n- expandTableFromCell được gọi với rows=1, cols=0\n- Cursor không bị dịch chuyển không mong muốn",
   "Functional", "P1"),

  (39, "05. Add Rails (Mở rộng)", "TC-T039",
   "Kéo nút '+' để thêm nhiều hàng cùng lúc",
   "Kiểm tra drag add-row để chọn số hàng.",
   "Add row rail hiển thị.",
   "1. Mousedown trên nút '+' dưới bảng\n2. Kéo xuống thêm 3 hàng nữa\n3. Thả chuột",
   "- Trong khi drag: preview số hàng được thêm thay đổi real-time\n- Cursor 'ns-resize' trong khi drag\n- Sau khi thả: số hàng đã chọn được thêm vào bảng\n- previewRows tính đúng theo delta Y",
   "Functional", "P2"),

  (40, "05. Add Rails (Mở rộng)", "TC-T040",
   "Hover phía phải bảng hiện nút '+' thêm cột",
   "Kiểm tra Add Column Rail hiển thị.",
   "Bảng đang trong trạng thái active.",
   "1. Di chuột sang phải bảng (sát border phải)\n2. Quan sát UI",
   "- Thanh dọc '+' xuất hiện bên phải bảng\n- Tooltip 'Thêm cột'\n- Height = chiều cao vùng visible của bảng",
   "UI", "P2"),

  (41, "05. Add Rails (Mở rộng)", "TC-T041",
   "Click nút '+' phải bảng thêm 1 cột mới",
   "Kiểm tra quick add column.",
   "Add column rail đang hiển thị.",
   "1. Click nút '+' bên phải bảng",
   "- 1 cột mới rỗng được thêm vào cuối bảng\n- expandTableFromCell với rows=0, cols=1",
   "Functional", "P1"),

  (42, "05. Add Rails (Mở rộng)", "TC-T042",
   "Kéo nút '+' phải bảng để thêm nhiều cột",
   "Kiểm tra drag add-column.",
   "Add column rail hiển thị.",
   "1. Mousedown nút '+' phải bảng\n2. Kéo ngang sang phải thêm 2 cột\n3. Thả chuột",
   "- Preview số cột tăng theo delta X\n- Cursor 'ew-resize'\n- Sau thả: số cột chọn được thêm\n- previewCols = 1 + floor(deltaX / avgColumnWidth)",
   "Functional", "P2"),

  (43, "05. Add Rails (Mở rộng)", "TC-T043",
   "Rails ẩn khi chuột rời khỏi bảng",
   "Kiểm tra cleanup khi mouseleave.",
   "Add rails đang hiển thị.",
   "1. Di chuột ra ngoài vùng bảng hoàn toàn",
   "- Nút '+' hàng và '+' cột đều ẩn (opacity=0)\n- pointerEvents='none'\n- Không có layout memory sai",
   "UI", "P3"),

  # ══════════════════════════════════════════════
  # 06. MERGE & SPLIT Ô
  # ══════════════════════════════════════════════
  (44, "06. Merge & Split Ô", "TC-T044",
   "Merge 2 ô ngang hàng (colspan=2)",
   "Kiểm tra gộp 2 ô cạnh nhau tạo colspan.",
   "Bảng 3×3, 2 ô liền kề được chọn.",
   "1. Click ô A2\n2. Shift+click ô B2 (hoặc kéo chuột)\n3. Click nút Merge (⊞↔) hoặc Table > 'Gộp ô'",
   "- Hai ô gộp thành 1 với colspan=2\n- Nội dung 2 ô được gộp lại\n- HTML: <td colspan='2'>...\n- Ô bên phải biến mất khỏi DOM",
   "Functional", "P1"),

  (45, "06. Merge & Split Ô", "TC-T045",
   "Merge ô dọc (rowspan=2)",
   "Kiểm tra gộp 2 ô theo chiều dọc.",
   "Bảng 3×3.",
   "1. Chọn ô A1 và A2 (cùng cột, khác hàng)\n2. Merge cells",
   "- Ô gộp với rowspan=2\n- HTML: <td rowspan='2'>...\n- Ô A2 biến mất",
   "Functional", "P1"),

  (46, "06. Merge & Split Ô", "TC-T046",
   "Merge ô theo khối (colspan và rowspan)",
   "Kiểm tra merge ô 2×2.",
   "Bảng 4×4.",
   "1. Chọn vùng 2×2 (A1, B1, A2, B2)\n2. Merge cells",
   "- Ô gộp với colspan=2, rowspan=2\n- 3 ô còn lại biến mất\n- HTML đúng cấu trúc",
   "Functional", "P1"),

  (47, "06. Merge & Split Ô", "TC-T047",
   "mergeTableCellsPreservingColumnWidths",
   "Kiểm tra column widths được giữ nguyên sau merge.",
   "Bảng có column widths đã được set.",
   "1. Resize cột A và B có width khác nhau\n2. Chọn A1 và B1\n3. Merge cells",
   "- Sau merge: colwidth attribute của ô gộp chứa widths của cả 2 cột gốc\n- Bảng không bị co giãn bất thường\n- dispatchTableLayoutChange được gọi",
   "Functional", "P1"),

  (48, "06. Merge & Split Ô", "TC-T048",
   "Split ô đã merge (colspan → các ô riêng)",
   "Kiểm tra tách ô colspan.",
   "Ô đang có colspan=2.",
   "1. Click vào ô đã merge\n2. Click nút Split hoặc Table > 'Tách ô'",
   "- Ô được tách thành 2 ô riêng lẻ\n- Ô bên phải được tạo rỗng\n- Tổng cấu trúc bảng khôi phục đúng",
   "Functional", "P1"),

  (49, "06. Merge & Split Ô", "TC-T049",
   "Split ô đã merge (rowspan → các ô riêng)",
   "Kiểm tra tách ô rowspan.",
   "Ô đang có rowspan=2.",
   "1. Click vào ô đã merge (rowspan)\n2. Split cell",
   "- Ô tách thành 2 ô theo chiều dọc\n- Ô bên dưới được tạo rỗng",
   "Functional", "P1"),

  (50, "06. Merge & Split Ô", "TC-T050",
   "Toggle Merge/Split button theo trạng thái ô",
   "Kiểm tra nút thay đổi giữa Merge và Split.",
   "Bảng đã tạo.",
   "1. Chọn 2 ô → nút hiển thị 'Merge'\n2. Merge → nút chuyển thành 'Split'\n3. Split → nút chuyển lại 'Merge'",
   "- canMergeCells=true khi có nhiều ô chọn → nút Merge\n- canSplitCell=true khi ô đã merge → nút Split\n- Nút disabled khi không merge được và không split được",
   "UI", "P1"),

  (51, "06. Merge & Split Ô", "TC-T051",
   "Disabled merge khi chỉ có 1 ô chọn",
   "Kiểm tra disabled state.",
   "Chỉ 1 ô đang được focus, ô chưa merge.",
   "1. Click đơn 1 ô\n2. Quan sát nút Merge/Split",
   "- canMergeCells = false (không chọn nhiều ô)\n- canSplitCell = false (ô chưa merge)\n- Nút disabled",
   "Functional", "P2"),

  (52, "06. Merge & Split Ô", "TC-T052",
   "Merge cell giữ nội dung từ tất cả ô",
   "Kiểm tra nội dung không bị mất khi merge.",
   "Ô A1 có 'Hello', ô B1 có 'World'.",
   "1. Chọn A1 và B1\n2. Merge cells\n3. Quan sát nội dung",
   "- Ô gộp chứa nội dung của cả 2 ô\n- Thứ tự: nội dung ô đầu trước, ô sau tiếp theo\n- Không mất dữ liệu",
   "Functional", "P2"),

  (53, "06. Merge & Split Ô", "TC-T053",
   "Merge cells hoạt động qua toolbar context buttons",
   "Kiểm tra nút ⊞ trong toolbar hiện khi chọn nhiều ô.",
   "Đang chọn nhiều ô trong bảng.",
   "1. Chọn nhiều ô\n2. Quan sát toolbar\n3. Click nút ⊞ (TableCellsMerge icon)",
   "- Nút ⊞ xuất hiện trong toolbar khi hasTableContext=true\n- Click nút gọi mergeTableCellsPreservingColumnWidths\n- Hoặc splitCell nếu canSplitCell=true",
   "UI", "P1"),

  # ══════════════════════════════════════════════
  # 07. RESIZE CỘT
  # ══════════════════════════════════════════════
  (54, "07. Resize Cột", "TC-T054",
   "Kéo thả border giữa 2 cột để resize",
   "Kiểm tra column resize tương tác.",
   "Bảng resizable=true (editable=true), bảng ≥ 2 cột.",
   "1. Hover chuột vào border giữa cột A và B\n2. Cursor thay đổi thành ↔\n3. Mousedown và kéo sang phải 50px\n4. Thả chuột",
   "- Cột A rộng hơn, cột B hẹp lại (hoặc ngược lại)\n- colwidth attribute được cập nhật\n- UEDITOR_TABLE_LAYOUT_CHANGE_EVENT được dispatch",
   "Functional", "P1"),

  (55, "07. Resize Cột", "TC-T055",
   "Column width được lưu trong colwidth attribute",
   "Kiểm tra persistence của column width.",
   "Đã resize một cột.",
   "1. Resize cột A\n2. Kiểm tra DOM attribute\n3. Gọi editor.getHTML() và kiểm tra output",
   "- HTML output chứa colwidth trong data attribute hoặc style\n- Reload content giữ nguyên width",
   "Functional", "P2"),

  (56, "07. Resize Cột", "TC-T056",
   "Resize cột đầu tiên (cột A)",
   "Kiểm tra resize cột ngoài cùng trái.",
   "Bảng ≥ 2 cột.",
   "1. Kéo border phải của cột A\n2. Quan sát các cột khác",
   "- Cột A thay đổi width\n- Cột B bù đắp width ngược lại\n- Bảng không bị vỡ layout",
   "Functional", "P2"),

  (57, "07. Resize Cột", "TC-T057",
   "Resize cột cuối cùng",
   "Kiểm tra resize cột ngoài cùng phải.",
   "Bảng ≥ 2 cột.",
   "1. Kéo border phải của cột cuối\n2. Quan sát layout",
   "- Cột cuối thay đổi width\n- Chiều rộng tổng bảng có thể thay đổi",
   "Functional", "P2"),

  (58, "07. Resize Cột", "TC-T058",
   "Resize bị disabled trong editable=false",
   "Kiểm tra không resize được trong read-only.",
   "editable=false.",
   "1. Mount UEditor với editable={false} và content có bảng\n2. Hover vào border cột\n3. Thử kéo resize",
   "- Cursor không thay đổi (không thành ↔)\n- Không thể kéo resize\n- UEditorTable.configure({ resizable: editable }) → resizable=false",
   "Functional", "P1"),

  (59, "07. Resize Cột", "TC-T059",
   "Column resize không crash khi column handle không tìm được DOM",
   "Kiểm tra graceful fallback khi DOM bất thường.",
   "Bảng đang render.",
   "1. Force resize trong điều kiện edge: bảng với merged cells\n2. Resize cột có merged cell",
   "- Không throw error\n- Width được tính đúng qua getDomColumnWidths hoặc getNodeColumnWidths fallback",
   "Functional", "P2"),

  # ══════════════════════════════════════════════
  # 08. HEADER ROW / COLUMN
  # ══════════════════════════════════════════════
  (60, "08. Header Row / Column", "TC-T060",
   "Bảng tạo mới có header row mặc định",
   "Kiểm tra withHeaderRow=true khi insert bảng.",
   "Bảng vừa được tạo.",
   "1. Tạo bảng qua InsertGrid\n2. Inspect DOM hàng đầu tiên",
   "- Hàng đầu tiên dùng <th> (tableHeader)\n- CSS class: 'border border-border p-2 bg-muted font-semibold min-w-25'\n- Các hàng body dùng <td>",
   "Functional", "P1"),

  (61, "08. Header Row / Column", "TC-T061",
   "Toggle Header Row - chuyển th ↔ td",
   "Kiểm tra toggleHeaderRow.",
   "Bảng có header row.",
   "1. Click ô trong header row\n2. Table menu > 'Toggle Header Row'\n3. Toggle lại",
   "- Lần 1: <th> chuyển thành <td>, style thay đổi (mất bg-muted)\n- Lần 2: <td> chuyển lại <th>, style khôi phục\n- Không ảnh hưởng đến hàng body",
   "Functional", "P1"),

  (62, "08. Header Row / Column", "TC-T062",
   "Toggle Header Column - chuyển cột đầu thành th",
   "Kiểm tra toggleHeaderColumn.",
   "Bảng 3×3.",
   "1. Click ô trong cột A\n2. Table menu > 'Toggle Header Column'",
   "- Cột A: tất cả ô chuyển thành <th>\n- Cột A có style header (font-semibold, bg-muted)\n- Các cột khác không đổi",
   "Functional", "P2"),

  (63, "08. Header Row / Column", "TC-T063",
   "Header style khác biệt rõ ràng về visual",
   "Kiểm tra UI phân biệt header và body.",
   "Bảng có header row và body rows.",
   "1. Quan sát bảng\n2. So sánh header row và body rows",
   "- Header row: nền xám (bg-muted), chữ đậm (font-semibold)\n- Body rows: nền trắng, chữ thường\n- Phân biệt rõ ràng bằng mắt thường",
   "UI", "P2"),

  (64, "08. Header Row / Column", "TC-T064",
   "Disabled toggle header khi không focus trong bảng",
   "Kiểm tra disabled state.",
   "Cursor KHÔNG trong bảng.",
   "1. Click ra ngoài bảng\n2. Kiểm tra nút Toggle Header",
   "- Nút 'Toggle Header Row/Column' bị disabled\n- editor.can().toggleHeaderRow() = false",
   "Functional", "P2"),

  # ══════════════════════════════════════════════
  # 09. CĂN LỀ BẢNG
  # ══════════════════════════════════════════════
  (65, "09. Căn lề Bảng", "TC-T065",
   "Căn lề bảng sang trái (align left)",
   "Kiểm tra table alignment left.",
   "Bảng đã tạo, cursor trong bảng.",
   "1. Mở Table dropdown trên toolbar\n2. Click 'Căn trái'",
   "- Bảng căn về phía trái container\n- data-table-align='left' hoặc textAlign='left'\n- Nút 'Căn trái' ở trạng thái active",
   "Functional", "P2"),

  (66, "09. Căn lề Bảng", "TC-T066",
   "Căn giữa bảng (align center)",
   "Kiểm tra table alignment center.",
   "Bảng đã tạo.",
   "1. Table dropdown > 'Căn giữa'",
   "- Bảng căn giữa (margin: auto)\n- data-table-align='center'\n- Nút 'Căn giữa' active",
   "Functional", "P2"),

  (67, "09. Căn lề Bảng", "TC-T067",
   "Căn lề phải bảng (align right)",
   "Kiểm tra table alignment right.",
   "Bảng đã tạo.",
   "1. Table dropdown > 'Căn phải'",
   "- Bảng căn về phía phải container\n- data-table-align='right'",
   "Functional", "P2"),

  (68, "09. Căn lề Bảng", "TC-T068",
   "Căn lề qua Table Control Menu (⊞)",
   "Kiểm tra table alignment từ floating menu.",
   "Bảng active, menu ⊞ hiển thị.",
   "1. Click ⊞\n2. Chọn 'Căn giữa'",
   "- Bảng căn giữa tương tự toolbar\n- applyTableAlignment được gọi với đúng cellPos",
   "Functional", "P2"),

  (69, "09. Căn lề Bảng", "TC-T069",
   "tableAlign attribute được persist trong HTML",
   "Kiểm tra lưu trạng thái căn lề.",
   "Đã set table align center.",
   "1. Set table align center\n2. Gọi editor.getHTML()\n3. Set content lại từ HTML đó\n4. Kiểm tra bảng vẫn căn giữa",
   "- HTML output: data-text-align='center' hoặc tableAlign attr\n- Reload: bảng vẫn giữ căn lề\n- parseHTML đọc đúng attribute",
   "Functional", "P2"),

  (70, "09. Căn lề Bảng", "TC-T070",
   "Active state indicator căn lề đúng",
   "Kiểm tra nút active phản ánh trạng thái hiện tại.",
   "Bảng đang align center.",
   "1. Click vào bảng đã align center\n2. Mở Table dropdown\n3. Quan sát nút nào active",
   "- Chỉ nút 'Căn giữa' ở trạng thái active (highlight)\n- Nút 'Căn trái' và 'Căn phải' ở trạng thái inactive\n- currentTableAlign = 'center'",
   "UI", "P3"),

  (71, "09. Căn lề Bảng", "TC-T071",
   "Disabled table alignment khi không có bảng",
   "Kiểm tra disabled state alignment.",
   "Cursor không trong bảng.",
   "1. Click ngoài bảng\n2. Mở Table dropdown\n3. Quan sát nút align",
   "- Các nút 'Căn trái/giữa/phải' bị disabled\n- hasTableContext = false",
   "Functional", "P2"),

  # ══════════════════════════════════════════════
  # 10. MÀU & STYLE Ô
  # ══════════════════════════════════════════════
  (72, "10. Màu & Style Ô", "TC-T072",
   "Đổi màu nền ô (Cell Background Color)",
   "Kiểm tra set backgroundColor cho ô bảng.",
   "Cursor trong ô bảng, bubble menu hiển thị.",
   "1. Click vào ô bảng\n2. Trong bubble menu chọn icon màu nền ô (CellBgColorIcon)\n3. Chọn màu xanh",
   "- Ô có nền màu xanh\n- data-background-color='#...' được set\n- style='background-color: #...' trong HTML output\n- Màu hiện thị ngay lập tức",
   "Functional", "P1"),

  (73, "10. Màu & Style Ô", "TC-T073",
   "Màu nền ô lưu trong data attribute và style",
   "Kiểm tra render HTML đúng cho backgroundColor.",
   "Ô đang có backgroundColor.",
   "1. Set backgroundColor cho ô\n2. Gọi editor.getHTML()\n3. Kiểm tra HTML",
   "- HTML: <td data-background-color='#color' style='background-color: #color'>\n- renderHTML của CustomTableCell merge style đúng\n- Không bị mất khi reload content",
   "Functional", "P1"),

  (74, "10. Màu & Style Ô", "TC-T074",
   "Đổi màu border ô (Cell Border Color)",
   "Kiểm tra set borderColor.",
   "Cursor trong ô bảng, bubble menu hiển thị.",
   "1. Click bubble menu icon CellBorderIcon\n2. Chọn màu border",
   "- Border ô thay đổi màu\n- data-border-color được set\n- style='border-color: ...' trong HTML",
   "Functional", "P2"),

  (75, "10. Màu & Style Ô", "TC-T075",
   "Đổi border style (solid/dashed/dotted)",
   "Kiểm tra set borderStyle.",
   "Cursor trong ô bảng.",
   "1. Tìm cách set borderStyle cho ô\n2. Chọn 'dashed'",
   "- Border ô hiển thị kiểu dashed\n- data-border-style='dashed'\n- style='border-style: dashed'",
   "Functional", "P3"),

  (76, "10. Màu & Style Ô", "TC-T076",
   "Reset màu nền ô về mặc định",
   "Kiểm tra xóa backgroundColor.",
   "Ô đang có backgroundColor.",
   "1. Mở color picker cho ô đã có màu\n2. Chọn 'Mặc định' hoặc ô trống",
   "- data-background-color bị xóa\n- style không còn background-color\n- Ô trở về màu mặc định",
   "Functional", "P2"),

  (77, "10. Màu & Style Ô", "TC-T077",
   "cellId attribute được parse và render",
   "Kiểm tra cellId attribute persistence.",
   "Ô có data-cell-id.",
   "1. Set content có <td data-cell-id='cell-1'>...\n2. Kiểm tra editor parse đúng\n3. Gọi getHTML()",
   "- data-cell-id được parse vào node.attrs.cellId\n- getHTML() render lại data-cell-id đúng\n- Không bị mất qua serialize/deserialize",
   "Functional", "P2"),

  (78, "10. Màu & Style Ô", "TC-T078",
   "numberFormat attribute lưu trữ format số",
   "Kiểm tra numberFormat attribute.",
   "Ô có data-number-format.",
   "1. Set content có <td data-number-format='0.00'>...\n2. Kiểm tra parse/render",
   "- data-number-format được parse đúng vào attrs\n- Render lại đúng trong HTML output",
   "Functional", "P3"),

  # ══════════════════════════════════════════════
  # 11. FORMULA (CÔNG THỨC)
  # ══════════════════════════════════════════════
  (79, "11. Formula (Công thức)", "TC-T079",
   "SUM - tính tổng vùng ô",
   "Kiểm tra hàm SUM tính đúng.",
   "Bảng có dữ liệu số: A1=10, A2=20, A3=30.",
   "1. Click vào ô A4\n2. Nhập công thức =SUM(A1:A3)\n3. Xem kết quả",
   "- computedValue = '60'\n- Ô hiển thị 60\n- data-formula='=SUM(A1:A3)', data-computed-value='60'",
   "Functional", "P1"),

  (80, "11. Formula (Công thức)", "TC-T080",
   "AVG - tính giá trị trung bình",
   "Kiểm tra hàm AVG.",
   "Bảng: A1=10, A2=20, A3=30.",
   "1. Nhập =AVG(A1:A3) vào ô A4",
   "- Kết quả: 20 ((10+20+30)/3)\n- computedValue = '20'",
   "Functional", "P1"),

  (81, "11. Formula (Công thức)", "TC-T081",
   "MIN - lấy giá trị nhỏ nhất",
   "Kiểm tra hàm MIN.",
   "Bảng: A1=5, A2=15, A3=10.",
   "1. Nhập =MIN(A1:A3)",
   "- Kết quả: 5\n- computedValue = '5'",
   "Functional", "P1"),

  (82, "11. Formula (Công thức)", "TC-T082",
   "MAX - lấy giá trị lớn nhất",
   "Kiểm tra hàm MAX.",
   "Bảng: A1=5, A2=15, A3=10.",
   "1. Nhập =MAX(A1:A3)",
   "- Kết quả: 15\n- computedValue = '15'",
   "Functional", "P1"),

  (83, "11. Formula (Công thức)", "TC-T083",
   "COUNT - đếm số ô có giá trị số",
   "Kiểm tra hàm COUNT.",
   "Bảng: A1=10, A2=20, A3=30.",
   "1. Nhập =COUNT(A1:A3)",
   "- Kết quả: 3\n- computedValue = '3'",
   "Functional", "P2"),

  (84, "11. Formula (Công thức)", "TC-T084",
   "Phép tính số học đơn giản (A1+B1)",
   "Kiểm tra biểu thức cộng.",
   "Bảng: A1=5, B1=3.",
   "1. Nhập =A1+B1 vào C1",
   "- Kết quả: 8\n- Hỗ trợ operator: +",
   "Functional", "P1"),

  (85, "11. Formula (Công thức)", "TC-T085",
   "Phép tính nhân và chia (A1*B1/C1)",
   "Kiểm tra operator * và /.",
   "Bảng: A1=10, B1=3, C1=2.",
   "1. Nhập =A1*B1/C1 vào D1",
   "- Kết quả: 15 (10*3/2)\n- Operator precedence đúng (* và / trước + và -)",
   "Functional", "P1"),

  (86, "11. Formula (Công thức)", "TC-T086",
   "Biểu thức phức tạp với ngoặc =(A1+B1)*C1",
   "Kiểm tra ưu tiên phép tính với ngoặc.",
   "Bảng: A1=2, B1=3, C1=4.",
   "1. Nhập =(A1+B1)*C1",
   "- Kết quả: 20 ((2+3)*4)\n- Ngoặc được xử lý đúng bởi parseFactor",
   "Functional", "P2"),

  (87, "11. Formula (Công thức)", "TC-T087",
   "Lỗi division-by-zero (#DIVISION-BY-ZERO)",
   "Kiểm tra xử lý chia cho 0.",
   "Bảng: A1=0.",
   "1. Nhập =10/A1",
   "- computedValue = '#DIVISION-BY-ZERO'\n- Ô hiển thị lỗi rõ ràng\n- Không crash",
   "Functional", "P1"),

  (88, "11. Formula (Công thức)", "TC-T088",
   "Lỗi invalid-reference (#INVALID-REFERENCE)",
   "Kiểm tra tham chiếu ô không tồn tại.",
   "Bảng 2×2, không có ô Z99.",
   "1. Nhập =Z99 (ô không tồn tại trong bảng)",
   "- computedValue = '#INVALID-REFERENCE'\n- Ô hiển thị lỗi",
   "Functional", "P1"),

  (89, "11. Formula (Công thức)", "TC-T089",
   "Lỗi invalid-formula (#INVALID-FORMULA)",
   "Kiểm tra cú pháp sai.",
   "Ô trống.",
   "1. Nhập =SUM( (ngoặc không đóng)\n2. Nhập =+++ (syntax lỗi)\n3. Nhập =UNKNOWN_FUNC(A1)",
   "- computedValue = '#INVALID-FORMULA'\n- Tokenizer trả về null\n- Không crash",
   "Functional", "P1"),

  (90, "11. Formula (Công thức)", "TC-T090",
   "Circular reference detection",
   "Kiểm tra phát hiện tham chiếu vòng tròn.",
   "Bảng 2×2.",
   "1. Nhập =B1 vào A1\n2. Nhập =A1 vào B1\n3. Gọi recalculateSelectedTable",
   "- Các ô trong circular reference hiển thị '#CIRCULAR-REFERENCE'\n- buildTableFormulaDependencyGraph và getTableFormulaRecalculationOrder phát hiện đúng\n- Không loop vô hạn",
   "Functional", "P1"),

  (91, "11. Formula (Công thức)", "TC-T091",
   "Formula tự cập nhật khi ô nguồn thay đổi",
   "Kiểm tra recalculation khi giá trị thay đổi.",
   "Ô A4 có công thức =SUM(A1:A3), A1=10, A2=20, A3=30.",
   "1. Thay đổi A1 từ 10 → 50\n2. Trigger recalculate\n3. Quan sát A4",
   "- A4 cập nhật thành 100 (50+20+30)\n- recalculateSelectedTable chạy lại toàn bộ bảng\n- buildTableValueGetter lấy computedValue mới nhất",
   "Functional", "P1"),

  (92, "11. Formula (Công thức)", "TC-T092",
   "normalizeFormulaInput - thêm = nếu thiếu",
   "Kiểm tra normalize formula input.",
   "Ô formula input.",
   "1. Nhập 'SUM(A1:A3)' (không có =)\n2. Nhập '=SUM(A1:A3)' (đã có =)",
   "- 'SUM(A1:A3)' → '=SUM(A1:A3)'\n- '=SUM(A1:A3)' → '=SUM(A1:A3)' (giữ nguyên)\n- Chuỗi rỗng → '' (formula bị clear)",
   "Functional", "P2"),

  (93, "11. Formula (Công thức)", "TC-T093",
   "clearSelectedTableCellFormula - xóa formula",
   "Kiểm tra xóa công thức.",
   "Ô đang có formula.",
   "1. Click ô có formula\n2. Gọi clearSelectedTableCellFormula\n3. Kiểm tra attrs",
   "- formula attr = null\n- computedValue attr = null\n- Ô hiển thị nội dung text bình thường",
   "Functional", "P2"),

  # ══════════════════════════════════════════════
  # 12. CLIPBOARD - HTML TABLE
  # ══════════════════════════════════════════════
  (94, "12. Clipboard - HTML Table", "TC-T094",
   "Paste HTML table từ Google Sheets",
   "Kiểm tra paste bảng từ Sheets giữ màu nền.",
   "Clipboard có HTML table từ Google Sheets với màu nền ô.",
   "1. Copy vùng data từ Google Sheets (có màu)\n2. Paste vào editor (Ctrl+V)",
   "- Bảng HTML được tạo trong editor\n- Màu nền ô được giữ nguyên (backgroundColor attr)\n- Text content đúng\n- Cấu trúc hàng/cột đúng",
   "Functional", "P1"),

  (95, "12. Clipboard - HTML Table", "TC-T095",
   "Paste HTML table từ Microsoft Excel",
   "Kiểm tra paste từ Excel giữ border và font style.",
   "Clipboard có HTML table từ Excel.",
   "1. Copy vùng data từ Excel\n2. Paste vào editor",
   "- Border color/style/width được giữ nếu có\n- Bold/italic text trong ô được giữ\n- Không crash khi parse HTML phức tạp của Excel",
   "Functional", "P1"),

  (96, "12. Clipboard - HTML Table", "TC-T096",
   "Paste HTML table có colspan/rowspan",
   "Kiểm tra merge cells từ clipboard.",
   "HTML table clipboard có colspan=2 và rowspan=2.",
   "1. Copy table có merged cells\n2. Paste vào editor",
   "- colspan/rowspan được parse đúng (parsePositiveInteger, max=100)\n- Bảng render đúng cấu trúc\n- normalizeTableRows xử lý đúng rowspan conflicts",
   "Functional", "P1"),

  (97, "12. Clipboard - HTML Table", "TC-T097",
   "Paste HTML table - style được áp dụng qua CSS class",
   "Kiểm tra parse CSS class styles từ clipboard.",
   "HTML table clipboard có <style> block định nghĩa class.",
   "1. Copy bảng từ nguồn có CSS class styling\n2. Paste vào editor",
   "- parseClipboardCssClassStyles đọc được styles từ <style> block\n- Inline styles được merge với class styles đúng\n- mergeStyleDeclarations ưu tiên inline style",
   "Functional", "P2"),

  (98, "12. Clipboard - HTML Table", "TC-T098",
   "Paste bảng - màu chữ sáng được làm tối để đọc được",
   "Kiểm tra ensureReadableSpreadsheetSegments.",
   "Clipboard có bảng với chữ màu trắng trên nền trắng.",
   "1. Copy bảng có text-color trắng trên background trắng\n2. Paste vào editor",
   "- isLightTextColor('#ffffff') = true\n- Không có dark background → text color được thay bằng DEFAULT (#000000)\n- Ô đọc được sau khi paste",
   "Functional", "P2"),

  # ══════════════════════════════════════════════
  # 13. CLIPBOARD - TSV
  # ══════════════════════════════════════════════
  (99, "13. Clipboard - TSV", "TC-T099",
   "Paste TSV data từ Excel (tab-separated)",
   "Kiểm tra parse TSV tạo bảng.",
   "Clipboard có text/plain với tab-separated values từ Excel.",
   "1. Copy vùng data 3×4 từ Excel\n2. Paste vào editor trống",
   "- Bảng 3 hàng × 4 cột được tạo\n- Dữ liệu mỗi ô đúng với cell gốc\n- Tất cả ô là <td> (không phải <th>)\n- Không tạo bảng nếu minColumnCount không đạt (< 2 cột)",
   "Functional", "P1"),

  (100, "13. Clipboard - TSV", "TC-T100",
   "TSV data nhiều hàng - số hàng đúng",
   "Kiểm tra TSV parser tạo đúng số hàng.",
   "Clipboard: 5 dòng, mỗi dòng 3 giá trị tab-separated.",
   "1. Paste data 5 hàng từ Excel",
   "- Bảng 5×3 được tạo\n- Dòng trống cuối cùng bị loại bỏ (parseClipboardTsvRows trim trailing empty rows)\n- Không thêm hàng thừa",
   "Functional", "P2"),

  (101, "13. Clipboard - TSV", "TC-T101",
   "TSV data có quoted fields với dấu nháy đôi",
   "Kiểm tra TSV parser xử lý quoted fields.",
   "Clipboard: '\"Hello, World\"\t42\n\"Say \\\"hi\\\"\"\t99'",
   "1. Paste TSV với trường được quote\n2. Kiểm tra nội dung ô",
   "- 'Hello, World' được parse đúng (dấu phẩy bên trong không tách ô)\n- 'Say \"hi\"' (escaped quote) được xử lý đúng\n- parseClipboardTsvRows handles double-quote escaping",
   "Functional", "P2"),

  (102, "13. Clipboard - TSV", "TC-T102",
   "Paste text plain không có tab - KHÔNG tạo bảng",
   "Kiểm tra fallback khi không phải TSV.",
   "Clipboard có text thuần không có ký tự tab.",
   "1. Copy text bình thường 'Hello World'\n2. Paste vào editor",
   "- Không tạo bảng\n- Text được paste bình thường\n- getClipboardTsvTableContent: text.includes('\\t') = false → return null",
   "Functional", "P1"),

  # ══════════════════════════════════════════════
  # 14. KEYBOARD NAVIGATION
  # ══════════════════════════════════════════════
  (103, "14. Keyboard Navigation", "TC-T103",
   "Enter trong ô không thoát ra ngoài bảng",
   "Kiểm tra Enter tạo dòng mới trong ô.",
   "Cursor đang trong ô bảng.",
   "1. Nhấn Enter trong ô bảng",
   "- Tạo dòng mới bên trong ô (paragraph mới)\n- Không thoát ra ngoài bảng\n- Ô bảng có thể chứa nhiều paragraph",
   "Functional", "P1"),

  (104, "14. Keyboard Navigation", "TC-T104",
   "Ctrl+A trong ô chọn nội dung ô (không chọn toàn bảng)",
   "Kiểm tra Ctrl+A scope.",
   "Cursor trong ô có text.",
   "1. Nhấn Ctrl+A trong ô bảng",
   "- Chỉ nội dung ô được chọn\n- Không chọn toàn bộ editor\n- Hành vi giống text area thông thường",
   "Functional", "P2"),

  (105, "14. Keyboard Navigation", "TC-T105",
   "ArrowLeft/Right không escape khỏi bảng",
   "Kiểm tra stopPropagation cho arrow keys.",
   "Cursor trong ô, đang ở đầu hoặc cuối text.",
   "1. Cursor ở đầu ô A1, nhấn ArrowLeft\n2. Cursor ở cuối ô, nhấn ArrowRight",
   "- Cursor không nhảy ra ngoài bảng\n- ArrowLeft tại đầu ô: không làm gì (hoặc di chuyển sang ô trước tùy ProseMirror)\n- event.stopPropagation() được gọi trong handleDOMEvents",
   "Functional", "P2"),

  (106, "14. Keyboard Navigation", "TC-T106",
   "Backspace ở đầu ô không xóa cấu trúc bảng",
   "Kiểm tra Backspace không phá vỡ bảng.",
   "Cursor ở đầu ô rỗng.",
   "1. Nhấn Backspace khi ô đang rỗng",
   "- Ô không bị xóa\n- Cấu trúc bảng nguyên vẹn\n- Tiptap Table extension xử lý đúng",
   "Functional", "P2"),

  # ══════════════════════════════════════════════
  # 15. UNDO / REDO
  # ══════════════════════════════════════════════
  (107, "15. Undo / Redo", "TC-T107",
   "Undo sau khi thêm hàng",
   "Kiểm tra history ghi nhận addRowAfter.",
   "Bảng 3×3.",
   "1. Thêm 1 hàng sau\n2. Nhấn Ctrl+Z (Undo)",
   "- Hàng vừa thêm biến mất\n- Bảng trở về 3 hàng\n- Cursor trở về vị trí trước khi thêm",
   "Functional", "P1"),

  (108, "15. Undo / Redo", "TC-T108",
   "Undo sau khi xóa cột",
   "Kiểm tra history ghi nhận deleteColumn.",
   "Bảng 3×3, đã xóa cột B.",
   "1. Xóa cột B\n2. Undo",
   "- Cột B được khôi phục với đầy đủ nội dung\n- Bảng trở về 3 cột",
   "Functional", "P1"),

  (109, "15. Undo / Redo", "TC-T109",
   "Undo sau khi merge cells",
   "Kiểm tra history ghi nhận merge.",
   "Đã merge 2 ô.",
   "1. Merge A1 và B1\n2. Undo",
   "- Hai ô được tách trở lại\n- colspan biến mất\n- Nội dung khôi phục",
   "Functional", "P1"),

  (110, "15. Undo / Redo", "TC-T110",
   "Redo sau khi Undo thao tác bảng",
   "Kiểm tra Redo khôi phục thao tác.",
   "Đã Undo 1 lần.",
   "1. Undo thêm hàng\n2. Redo (Ctrl+Y)",
   "- Hàng được thêm trở lại\n- Trạng thái bảng đúng như trước khi Undo",
   "Functional", "P1"),

  # ══════════════════════════════════════════════
  # 16. READ-ONLY MODE
  # ══════════════════════════════════════════════
  (111, "16. Read-only Mode", "TC-T111",
   "Bảng không resize được trong editable=false",
   "Kiểm tra resizable=false khi read-only.",
   "editable=false, content có bảng.",
   "1. Mount với editable={false}\n2. Hover vào border cột\n3. Thử resize",
   "- Không có resize handle/cursor\n- UEditorTable.configure({ resizable: false })\n- Bảng hiển thị đúng nhưng không tương tác",
   "Functional", "P1"),

  (112, "16. Read-only Mode", "TC-T112",
   "Row/column handles không xuất hiện trong read-only",
   "Kiểm tra TableControls ẩn trong read-only.",
   "editable=false.",
   "1. Click vào ô trong bảng read-only\n2. Quan sát UI",
   "- Row handles không hiện\n- Column handles không hiện\n- TableControlMenu (⊞) không hiển thị\n- <TableControls> chỉ render khi editable=true",
   "UI", "P1"),

  (113, "16. Read-only Mode", "TC-T113",
   "Bảng render đúng style và data trong read-only",
   "Kiểm tra visual output đúng khi không editable.",
   "Content có bảng với màu nền ô, colspan.",
   "1. Mount với editable={false} và content phức tạp\n2. Quan sát bảng",
   "- Màu nền ô hiển thị đúng\n- colspan/rowspan render đúng\n- Style CSS đúng\n- Không có toolbar, handles hay menu",
   "UI", "P1"),

  # ══════════════════════════════════════════════
  # 17. prepareContentForSave
  # ══════════════════════════════════════════════
  (114, "17. prepareContentForSave", "TC-T114",
   "HTML output bảng giữ nguyên cấu trúc sau save",
   "Kiểm tra bảng không bị thay đổi qua save pipeline.",
   "Bảng có header, data, và styled cells.",
   "1. Tạo bảng với dữ liệu\n2. Gọi ref.current.prepareContentForSave()\n3. Kiểm tra result.html",
   "- Cấu trúc bảng trong HTML giữ nguyên\n- Số hàng, cột đúng\n- Không có ô nào bị mất",
   "Functional", "P1"),

  (115, "17. prepareContentForSave", "TC-T115",
   "data-* attributes của ô được giữ trong HTML output",
   "Kiểm tra cell attrs không bị mất qua save.",
   "Ô bảng có data-background-color, data-formula, data-cell-id.",
   "1. Set các attributes cho ô\n2. Gọi prepareContentForSave()\n3. Parse result.html",
   "- data-background-color còn trong HTML output\n- data-formula còn trong output\n- data-computed-value còn trong output\n- data-cell-id còn trong output",
   "Functional", "P1"),

  (116, "17. prepareContentForSave", "TC-T116",
   "Bảng có ảnh trong ô - ảnh được upload đúng",
   "Kiểm tra ảnh base64 trong ô bảng được upload.",
   "Ô bảng có chứa ảnh base64.",
   "1. Paste ảnh vào ô bảng\n2. Gọi prepareContentForSave()\n3. Kiểm tra result",
   "- uploadImageForSave được gọi cho ảnh trong ô bảng\n- result.html: src ảnh thay bằng URL\n- Cấu trúc bảng không bị phá vỡ",
   "Functional", "P2"),

  # ══════════════════════════════════════════════
  # 18. EDGE CASES
  # ══════════════════════════════════════════════
  (117, "18. Edge Cases", "TC-T117",
   "Bảng 1×1 (một ô duy nhất)",
   "Kiểm tra bảng tối giản.",
   "Insert bảng 1×1.",
   "1. Tạo bảng 1 hàng × 1 cột\n2. Nhập text\n3. Thử thêm hàng, thêm cột\n4. Thử merge",
   "- Bảng 1×1 tạo và hiển thị đúng\n- Thêm hàng/cột hoạt động bình thường\n- Merge không thể (chỉ 1 ô)\n- Xóa hàng duy nhất: xóa cả bảng",
   "Functional", "P2"),

  (118, "18. Edge Cases", "TC-T118",
   "Bảng với tất cả ô rỗng",
   "Kiểm tra rendering bảng không có nội dung.",
   "Bảng mới tạo chưa nhập gì.",
   "1. Tạo bảng 3×3\n2. Không nhập gì\n3. Click ra ngoài\n4. Gọi getHTML()",
   "- Bảng vẫn render đúng với đủ hàng/cột\n- HTML chứa cấu trúc <tr><td></td></tr> đúng\n- Không bị collapse hay mất cấu trúc",
   "Functional", "P2"),

  (119, "18. Edge Cases", "TC-T119",
   "Bảng với nội dung rất dài trong 1 ô",
   "Kiểm tra overflow handling.",
   "Bảng đã tạo.",
   "1. Nhập 1000 ký tự vào 1 ô\n2. Quan sát layout\n3. Scroll",
   "- Ô mở rộng theo chiều dọc (word-wrap)\n- Bảng không bị vỡ layout\n- Scroll trong editor hoạt động bình thường",
   "Functional", "P3"),

  (120, "18. Edge Cases", "TC-T120",
   "Copy bảng và paste lại vào editor",
   "Kiểm tra copy/paste toàn bộ bảng.",
   "Bảng đã có dữ liệu.",
   "1. Chọn toàn bộ bảng (Ctrl+A hoặc kéo)\n2. Copy (Ctrl+C)\n3. Click vào paragraph bên dưới\n4. Paste (Ctrl+V)",
   "- Bảng được duplicate\n- Cấu trúc và nội dung giữ nguyên\n- Hai bảng độc lập nhau\n- Nếu bảng có ảnh: ảnh được copy đúng",
   "Functional", "P2"),
]

# ──────────────────────────────────────────────
# BUILD WORKBOOK
# ──────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: Test Cases ────────────────────────
ws = wb.active
ws.title = "Table Test Cases"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A3"

# Banner
ws.merge_cells("A1:M1")
b = ws["A1"]
b.value = "UEDITOR — TABLE FEATURE — TEST CASE SPECIFICATION  (120 TC)"
b.font = Font(bold=True, size=13, color="FFFFFF", name="Segoe UI")
b.fill = fill(C_HEADER_BG)
b.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Column headers
ws.row_dimensions[2].height = 36
for ci, (cname, cw) in enumerate(COLS, 1):
    cell = ws.cell(row=2, column=ci, value=cname)
    cell.font = bfont(10, "FFFFFF")
    cell.fill = fill(C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border()
    ws.column_dimensions[get_column_letter(ci)].width = cw

# Validation
dv_s = DataValidation(type="list", formula1='"Chưa test,Pass,Fail,Blocked,Re-test,Fixed,Skip"')
dv_p = DataValidation(type="list", formula1='"P1,P2,P3"')
dv_t = DataValidation(type="list", formula1='"Functional,UI,Edge case"')
ws.add_data_validation(dv_s)
ws.add_data_validation(dv_p)
ws.add_data_validation(dv_t)

for ri, tc in enumerate(TCS):
    row = ri + 3
    (stt, mod, ma, ten, mota, dk, buoc, kq, loai, prior) = tc
    ws.row_dimensions[row].height = 90
    is_alt = ri % 2 == 1
    bg = C_ALT_ROW if is_alt else C_WHITE
    mc = MODULE_COLORS.get(mod, "1E293B")

    vals = [stt, mod, ma, ten, mota, dk, buoc, kq, "", "Chưa test", prior, loai, ""]

    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.border = border()
        cell.font = nfont()
        cell.alignment = wrap()

        if ci == 1:      # STT
            cell.font = bfont(9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill(bg)
        elif ci == 2:    # Module
            cell.font = Font(bold=True, size=8, color="FFFFFF", name="Segoe UI")
            cell.fill = fill(mc)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        elif ci == 3:    # Mã TC
            cell.font = Font(bold=True, size=9, color="1D4ED8", name="Courier New")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill(bg)
        elif ci == 4:    # Tên
            cell.font = Font(bold=True, size=9, color="0F172A", name="Segoe UI")
            cell.fill = fill(bg)
        elif ci == 10:   # Status
            cell.value = "Chưa test"
            cell.fill = fill(STATUS_COLORS["Chưa test"])
            cell.font = bfont(9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif ci == 11:   # Priority
            cell.fill = fill(PRIORITY_COLORS.get(prior, C_WHITE))
            cell.font = bfont(9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif ci == 12:   # Loại test
            cell.fill = fill(bg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.fill = fill(bg)

    dv_s.add(f"J{row}")
    dv_p.add(f"K{row}")
    dv_t.add(f"L{row}")

# ── Sheet 2: Tổng hợp module ──────────────────
ws2 = wb.create_sheet("Tổng hợp Module")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:G1")
b2 = ws2["A1"]
b2.value = "TỔNG HỢP TEST CASE TABLE THEO MODULE"
b2.font = Font(bold=True, size=12, color="FFFFFF", name="Segoe UI")
b2.fill = fill(C_HEADER_BG)
b2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 24

for ci, (cn, cw) in enumerate([
    ("Module", 30), ("Tổng TC", 9), ("Pass", 8), ("Fail", 8),
    ("Blocked", 9), ("Chưa test", 11), ("Ghi chú", 26)
], 1):
    c = ws2.cell(row=2, column=ci, value=cn)
    c.font = bfont(10, "FFFFFF")
    c.fill = fill("334155")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()
    ws2.column_dimensions[get_column_letter(ci)].width = cw
ws2.row_dimensions[2].height = 22

counts = Counter(tc[1] for tc in TCS)
for ri, (mname, cnt) in enumerate(sorted(counts.items()), 3):
    ws2.row_dimensions[ri].height = 20
    bg = C_ALT_ROW if ri % 2 == 0 else C_WHITE
    mc = MODULE_COLORS.get(mname, "1E293B")
    for ci, v in enumerate([mname, cnt, 0, 0, 0, cnt, ""], 1):
        c = ws2.cell(row=ri, column=ci, value=v)
        c.border = border()
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center", wrap_text=True)
        if ci == 1:
            c.font = Font(bold=True, size=9, color="FFFFFF", name="Segoe UI")
            c.fill = fill(mc)
        elif ci == 2:
            c.font = bfont(9); c.fill = fill(bg)
        elif ci == 6:
            c.font = nfont(9, "64748B"); c.fill = fill("F1F5F9")
        else:
            c.font = nfont(9); c.fill = fill(bg)

total_row = len(counts) + 3
ws2.row_dimensions[total_row].height = 22
for ci, v in enumerate(["TỔNG CỘNG", len(TCS), 0, 0, 0, len(TCS), ""], 1):
    c = ws2.cell(row=total_row, column=ci, value=v)
    c.font = bfont(10, "FFFFFF")
    c.fill = fill(C_HEADER_BG)
    c.border = border()
    c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")

# ── Sheet 3: Hướng dẫn ───────────────────────
ws3 = wb.create_sheet("Hướng dẫn")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 52

ws3.merge_cells("A1:B1")
leg = ws3["A1"]
leg.value = "HƯỚNG DẪN SỬ DỤNG — TABLE TEST CASE"
leg.font = Font(bold=True, size=12, color="FFFFFF", name="Segoe UI")
leg.fill = fill(C_HEADER_BG)
leg.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 24

legend = [
    ("", ""),
    ("── STATUS ──", "Ý nghĩa"),
    ("Chưa test", "Test case chưa được thực thi"),
    ("Pass",      "Kết quả đúng hoàn toàn"),
    ("Fail",      "Kết quả sai, cần tạo bug ticket"),
    ("Blocked",   "Không test được do blocker"),
    ("Re-test",   "Cần test lại sau khi fix"),
    ("Fixed",     "Bug fixed, đang chờ verify"),
    ("Skip",      "Không áp dụng / out of scope"),
    ("", ""),
    ("── ƯU TIÊN ──", "Mô tả"),
    ("P1", "Critical — Phải pass trước khi release"),
    ("P2", "High — Quan trọng, fix trước release"),
    ("P3", "Medium/Low — Fix sau release nếu có thể"),
    ("", ""),
    ("── CÁC MODULE ──", "Tính năng bao phủ"),
    ("01. Tạo bảng",        "Insert grid, slash cmd, menu bar"),
    ("02. Điều hướng",      "Tab, Shift+Tab, arrow keys, click"),
    ("03. Thao tác Hàng",   "Add/Delete/Duplicate/Clear/Drag row"),
    ("04. Thao tác Cột",    "Add/Delete/Duplicate/Clear/Drag col"),
    ("05. Add Rails",       "Quick add row/col bằng kéo nút +"),
    ("06. Merge & Split",   "Gộp ô colspan/rowspan, tách ô"),
    ("07. Resize Cột",      "Kéo thả border để resize column"),
    ("08. Header",          "Toggle header row/column"),
    ("09. Căn lề Bảng",     "Align bảng left/center/right"),
    ("10. Màu & Style Ô",   "backgroundColor, border, cellId"),
    ("11. Formula",         "SUM/AVG/MIN/MAX/COUNT, error types"),
    ("12. Clipboard HTML",  "Paste từ Excel/Sheets (HTML)"),
    ("13. Clipboard TSV",   "Paste tab-separated values"),
    ("14. Keyboard Nav",    "Enter, Backspace, Ctrl+A trong ô"),
    ("15. Undo/Redo",       "History cho table operations"),
    ("16. Read-only",       "editable=false mode"),
    ("17. Save",            "prepareContentForSave pipeline"),
    ("18. Edge Cases",      "1×1, empty, long content, copy"),
]

sc = {
    "Chưa test": "F1F5F9", "Pass": "DCFCE7", "Fail": "FEE2E2",
    "Blocked": "F3E8FF", "Re-test": "DBEAFE", "Fixed": "FEF9C3", "Skip": "F1F5F9",
}
pc = {"P1": "FEE2E2", "P2": "FEF9C3", "P3": "DCFCE7"}

for ri, (k, v) in enumerate(legend, 2):
    ws3.row_dimensions[ri].height = 18
    ca = ws3.cell(row=ri, column=1, value=k)
    cb = ws3.cell(row=ri, column=2, value=v)
    if k.startswith("──"):
        for c in [ca, cb]:
            c.font = bfont(10, "FFFFFF"); c.fill = fill("334155")
            c.alignment = Alignment(vertical="center"); c.border = border()
    elif k in sc:
        ca.fill = fill(sc[k]); ca.font = bfont(9)
        cb.font = nfont(9)
        for c in [ca, cb]: c.border = border()
    elif k in pc:
        ca.fill = fill(pc[k]); ca.font = bfont(9)
        cb.font = nfont(9)
        for c in [ca, cb]: c.border = border()
    elif k.startswith("0") and "." in k:
        ca.font = Font(bold=True, size=9, color="1D4ED8", name="Segoe UI")
        cb.font = nfont(9)
        for c in [ca, cb]:
            c.border = border(); c.fill = fill(C_ALT_ROW)
    elif k == "":
        pass
    else:
        ca.font = nfont(9); cb.font = nfont(9)
        for c in [ca, cb]: c.border = border(); c.fill = fill(C_WHITE)

# ── Save ──────────────────────────────────────
out = "/Users/tran_van_bach/Desktop/project/nextJs/underverse/docs/UEditor_Table_TestCase.xlsx"
wb.save(out)
print(f"✓ Saved: {out}")
print(f"  Total TC: {len(TCS)}")
print(f"  Modules : {len(counts)}")
breakdown = sorted(counts.items())
for m, c in breakdown:
    print(f"  {m}: {c} TC")
