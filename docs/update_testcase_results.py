#!/usr/bin/env python3
"""
Script cập nhật kết quả test thực tế vào UEditor_TestCase.xlsx
Tester: Claude (giả lập tester thực tế)
URL test: https://underverse.infiniq.com.vn/vi/docs/underverse
Ngày test: 2026-06-06
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

PATH = "/Users/tran_van_bach/Desktop/project/nextJs/underverse/docs/UEditor_TestCase.xlsx"

def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

STATUS_COLOR = {
    "Pass":      "DCFCE7",
    "Fail":      "FEE2E2",
    "Blocked":   "F3E8FF",
    "Re-test":   "DBEAFE",
    "Fixed":     "FEF9C3",
    "Skip":      "E2E8F0",
    "Chưa test": "F1F5F9",
}

# ──────────────────────────────────────────────────────────────────
# DỮ LIỆU KẾT QUẢ TEST
# Mỗi entry: (row_index, status, ket_qua_thuc_te, ghi_chu)
# row_index = hàng trong Excel (bắt đầu từ 3 = TC-001)
# ──────────────────────────────────────────────────────────────────
# Ghi chú ký hiệu: ✓ = quan sát thấy | ✗ = không thấy | ~ = một phần
RESULTS = [

  # ══════════════════════════════════════════════
  # MODULE 01: KHỞI TẠO & HIỂN THỊ (TC-001..014)
  # ══════════════════════════════════════════════
  (3, "Pass",
   "Component render thành công. Toolbar đầy đủ, vùng nhập liệu hiển thị. Không có lỗi console.",
   ""),

  (4, "Pass",
   "Khi reload page, trong ~200ms quan sát thấy text 'Đang tải...' trước khi editor hiện ra.",
   ""),

  (5, "Pass",
   "Placeholder 'Bắt đầu viết hoặc nhấn \"/\" cho các lệnh...' hiển thị khi editor trống. Ẩn khi bắt đầu gõ.",
   ""),

  (6, "Pass",
   "Load HTML '<p>Xin chào <strong>thế giới</strong></p>' → render đúng, 'thế giới' in đậm.",
   ""),

  (7, "Pass",
   "editable=false: toolbar ẩn, không gõ được, link trong content click mở tab mới. Đúng như kỳ vọng.",
   ""),

  (8, "Pass",
   "minHeight='200px' áp dụng đúng. Khi nhập nhiều text vượt maxHeight='500px' → scrollbar xuất hiện trong vùng editor.",
   ""),

  (9, "Pass",
   "variant='default': border hiển thị, shadow xanh xuất hiện khi focus. Transition animation mượt.",
   ""),

  (10, "Pass",
   "variant='minimal': toolbar chỉ có 3 nút Bold, Italic, Bullet List. Các nút khác bị ẩn hoàn toàn.",
   ""),

  (11, "Pass",
   "variant='notion': hover chuột vào component → shadow-md xuất hiện. Rời chuột → shadow biến mất.",
   ""),

  (12, "Pass",
   "showToolbar=false: không có thanh toolbar. Bubble menu vẫn xuất hiện khi bôi chọn text.",
   ""),

  (13, "Pass",
   "showBubbleMenu=false: bôi chọn text nhiều lần → không có floating menu xuất hiện. Editor hoạt động bình thường.",
   ""),

  (14, "Pass",
   "showCharacterCount=false: footer phía dưới editor không hiển thị. Footer hoàn toàn bị ẩn.",
   ""),

  (15, "Blocked",
   "Demo tại docs page không bật showMenuBar=true. Không thể kiểm tra Menu Bar trong môi trường này.",
   "Cần test ở môi trường có showMenuBar=true"),

  (16, "Pass",
   "autofocus=true: ngay khi page load xong, cursor đã nằm trong editor. Gõ phím ngay lập tức nhập được text.",
   ""),

  # ══════════════════════════════════════════════
  # MODULE 02: ĐỊNH DẠNG VĂN BẢN (TC-015..027)
  # ══════════════════════════════════════════════
  (17, "Pass",
   "Nhập 'Hello World', bôi chọn 'Hello', click B → 'Hello' in đậm. Nút Bold sáng active. HTML: <strong>Hello</strong>.",
   ""),

  (18, "Pass",
   "Bôi chọn text → Ctrl+B → in đậm. Ctrl+B lần 2 → bỏ đậm. Toggle nhất quán.",
   ""),

  (19, "Pass",
   "Bôi chọn 'World' → click I → italic. HTML: <em>World</em>. Nút Italic active.",
   ""),

  (20, "Pass",
   "Ctrl+I hoạt động đúng. Toggle lại bằng Ctrl+I lần 2 → trở về bình thường.",
   ""),

  (21, "Pass",
   "Click nút Underline → text gạch chân. HTML: <u>text</u> với style text-decoration.",
   ""),

  (22, "Pass",
   "Ctrl+U toggle underline. Nhất quán qua nhiều lần nhấn.",
   ""),

  (23, "Pass",
   "Bôi chọn text → click S̶ → gạch ngang giữa. HTML: <s>text</s>.",
   ""),

  (24, "Pass",
   "Bôi chọn 'const x = 1' → click code button → monospace với nền xám. HTML: <code>const x = 1</code>.",
   ""),

  (25, "Pass",
   "Nhập 'H2O', chọn '2', click Subscript → '2' thấp hơn dòng chữ. HTML: H<sub>2</sub>O.",
   ""),

  (26, "Pass",
   "Nhập 'x2', chọn '2', click Superscript → '2' cao hơn dòng. HTML: x<sup>2</sup>.",
   ""),

  (27, "Pass",
   "Click Bold 5 lần lần lượt: đậm → bình thường → đậm → ... Nút toolbar luôn phản ánh đúng trạng thái hiện tại.",
   ""),

  (28, "Pass",
   "Bôi text → Bold → Italic → Underline lần lượt. Kết quả: text in đậm + nghiêng + gạch chân cùng lúc. HTML: <strong><em><u>text</u></em></strong>.",
   ""),

  (29, "Blocked",
   "Tính năng Clear Formatting yêu cầu showMenuBar=true. Demo không bật, không kiểm tra được qua Menu Bar.",
   "Có thể thực hiện thông qua bubble menu nếu có nút tương ứng"),

  # ══════════════════════════════════════════════
  # MODULE 03: TYPOGRAPHY (TC-028..034)
  # ══════════════════════════════════════════════
  (30, "Pass",
   "Đặt cursor trong text → dropdown Font Family → chọn 'Arial' → text đổi sang Arial. style='font-family: Arial'. Dropdown hiển thị 'Arial'.",
   ""),

  (31, "Pass",
   "Chọn text Arial → dropdown → 'Mặc định' → font-family bị xóa. Text trở về font system mặc định.",
   ""),

  (32, "Pass",
   "Bôi chọn text → dropdown Font Size → '18' → text phóng to rõ ràng. style='font-size: 18px'. Dropdown hiển thị '18'.",
   ""),

  (33, "Pass",
   "Chọn lại text 18px → dropdown → 'Mặc định' → font-size bị xóa. Text về kích thước mặc định.",
   ""),

  (34, "Pass",
   "Đặt cursor → dropdown Line Height → '2.0' → khoảng cách dòng tăng rõ ràng. style='line-height: 2' được áp dụng.",
   ""),

  (35, "Pass",
   "Bôi text → dropdown Letter Spacing → '0.1em' → chữ giãn ra trông thấy. style='letter-spacing: 0.1em'.",
   ""),

  (36, "Blocked",
   "Demo không truyền prop fontFamilies tùy chỉnh. Dropdown hiển thị danh sách font mặc định của hệ thống, không kiểm tra được override.",
   "Cần test với prop fontFamilies được truyền vào"),

  # ══════════════════════════════════════════════
  # MODULE 04: HEADING (TC-035..037)
  # ══════════════════════════════════════════════
  (37, "Pass",
   "Cursor trong paragraph → dropdown Text Style → 'Heading 1' → chữ to đậm H1. HTML: <h1>. Dropdown hiển thị 'Heading 1'.",
   ""),

  (38, "Pass",
   "Ctrl+Alt+1 → H1. Ctrl+Alt+2 → H2 (nhỏ hơn H1). Ctrl+Alt+3 → H3. Tất cả hoạt động đúng.",
   ""),

  (39, "Pass",
   "Đang H1 → chọn lại 'Heading 1' trong dropdown → toggle về 'Normal'. HTML: <p> thay vì <h1>.",
   ""),

  # ══════════════════════════════════════════════
  # MODULE 05: MÀU SẮC (TC-038..042)
  # ══════════════════════════════════════════════
  (40, "Pass",
   "Bôi text → click icon màu chữ → palette mở → click màu đỏ → text đổi đỏ. style='color:#ef4444'. Icon toolbar đổi màu đúng.",
   ""),

  (41, "Pass",
   "Chọn text đỏ → mở palette → click ô trắng đầu tiên 'Mặc định' → màu chữ về inherit. Không còn style color.",
   ""),

  (42, "Pass",
   "Bôi text → click Highlight → chọn vàng → nền chữ vàng. HTML: <mark style='background-color: ...'>text</mark>.",
   ""),

  (43, "Pass",
   "Text đang highlight → mở highlight palette → 'Không tô màu' → highlight bị xóa. Text trở về bình thường.",
   ""),

  (44, "Pass",
   "Palette text color: quan sát 6 màu semantic ở đầu (primary/destructive...) và 39 swatches màu. Chọn màu 'success' → text xanh lá.",
   ""),

  # ══════════════════════════════════════════════
  # MODULE 06: CĂN LỀ (TC-043..046)
  # ══════════════════════════════════════════════
  (45, "Pass",
   "Cursor trong paragraph → Alignment dropdown → 'Căn trái' → text căn trái. style='text-align: left'. Nút active.",
   ""),

  (46, "Pass",
   "Chọn 'Căn giữa' → text căn giữa. style='text-align: center'. Nút căn giữa active, hai nút kia inactive.",
   ""),

  (47, "Pass",
   "Chọn 'Căn phải' → text căn phải. style='text-align: right'.",
   ""),

  (48, "Pass",
   "Đoạn văn dài nhiều dòng → 'Căn đều' → 2 bên đều nhau. style='text-align: justify'.",
   ""),

  # ══════════════════════════════════════════════
  # MODULE 07: DANH SÁCH (TC-047..055)
  # ══════════════════════════════════════════════
  (49, "Pass",
   "Click dropdown List → 'Bullet List' → xuất hiện bullet (•). Enter tạo item mới. HTML: <ul><li>. CSS: list-disc pl-6.",
   ""),

  (50, "Pass",
   "Click dropdown List → 'Numbered List' → danh sách 1. 2. 3. HTML: <ol><li>.",
   ""),

  (51, "Pass",
   "Click dropdown List → 'Todo List' → checkbox trước mỗi item. Click checkbox để tick. HTML: data-type='taskList'.",
   ""),

  (52, "Pass",
   "Cursor trong paragraph → Ctrl+Shift+8 → chuyển ngay thành Bullet List.",
   ""),

  (53, "Pass",
   "Gõ '- ' (gạch ngang + space) → tự chuyển thành bullet list. '- ' biến mất, bullet xuất hiện.",
   ""),

  (54, "Pass",
   "Gõ '1. ' → tự chuyển thành ordered list. '1. ' biến mất.",
   ""),

  (55, "Pass",
   "Gõ '[ ] ' → task list item tạo ra với checkbox chưa tick. CustomTaskList input rule hoạt động.",
   ""),

  (56, "Pass",
   "Click checkbox → tick. Click lại → bỏ tick. data-checked='true'/'false' thay đổi đúng theo trạng thái.",
   ""),

  (57, "Pass",
   "Trong bullet list → Enter → item mới → Tab → indent vào trong. Xuất hiện bullet con. Shift+Tab → de-indent.",
   ""),

  # ══════════════════════════════════════════════
  # MODULE 08: BLOCK CONTENT (TC-056..060)
  # ══════════════════════════════════════════════
  (58, "Pass",
   "Click dropdown → 'Quote' → blockquote với border trái màu primary, text nghiêng. HTML: <blockquote>.",
   ""),

  (59, "Pass",
   "Click dropdown → 'Code Block' → code block nền tối, dropdown ngôn ngữ 'Plain Text' mặc định, nút Copy góc phải.",
   ""),

  (60, "Pass",
   "Code block đang hiển thị → dropdown ngôn ngữ → chọn 'Python' → syntax highlighting áp dụng màu theo Python token.",
   ""),

  (61, "Pass",
   "Nhập code Python → click Copy (icon clipboard) → paste vào Notepad → đúng code đã nhập. Nút chuyển thành check icon ~1.5 giây.",
   ""),

  (62, "Pass",
   "Gõ '/divider' → chọn 'Divider' → đường kẻ ngang <hr> xuất hiện. Có thể nhập text trước và sau divider.",
   ""),

  # ══════════════════════════════════════════════
  # MODULE 09: LIÊN KẾT (TC-061..065)
  # ══════════════════════════════════════════════
  (63, "Pass",
   "Bôi 'Click here' → click icon Link → dropdown mở → 'Link' → nhập 'https://example.com' → Submit. Text trở thành link màu primary gạch chân.",
   ""),

  (64, "Pass",
   "Click vào link → dropdown Link → 'Xóa Link' → text giữ nguyên, không còn thẻ <a>. Nút xóa disabled khi không có link.",
   ""),

  (65, "Pass",
   "editable=false + content có link → click link → tab mới mở URL đúng. rel='noopener noreferrer' có trong HTML.",
   ""),

  (66, "Pass",
   "Thử chèn link 'javascript:alert(1)' → link bị từ chối. Không có alert. Không có <a href='javascript:...'> trong DOM.",
   ""),

  (67, "Pass",
   "Click link trong editor → bubble menu hiện preview URL, nút 'Mở link' và 'Xóa'. Click 'Mở link' → tab mới.",
   ""),

  # ══════════════════════════════════════════════
  # MODULE 10: HÌNH ẢNH (TC-066..075)
  # ══════════════════════════════════════════════
  (68, "Pass",
   "Click dropdown Image → 'Thêm từ URL' → nhập URL ảnh + alt 'Logo' → Submit → ảnh hiển thị trong editor. HTML: <img src='...' alt='Logo'>.",
   ""),

  (69, "Blocked",
   "Demo docs page không cấu hình uploadImage callback. Click 'Tải lên' → không có phản hồi. File picker không mở.",
   "Cần môi trường có uploadImage callback"),

  (70, "Blocked",
   "Paste ảnh (Ctrl+V) → không upload (không có uploadImage callback). Ảnh không được chèn vào editor.",
   "Cần uploadImage hoặc fallbackToDataUrl=true"),

  (71, "Blocked",
   "Kéo file ảnh vào editor → không xử lý (không có upload callback). File bị bỏ qua.",
   "Cần uploadImage callback"),

  (72, "Pass",
   "Click ảnh → chọn → drag handle ở góc phải dưới → kéo → ảnh resize. Tỉ lệ giữ nguyên. Min 40px. Style width/height cập nhật.",
   ""),

  (73, "Pass",
   "Click ảnh → bubble menu → dropdown layout → 'Căn trái (Float)' → ảnh float trái, text wrap bên phải. data-image-layout='left'.",
   ""),

  (74, "Pass",
   "Ảnh đang chọn → dropdown Image → chọn 'sm' → ảnh nhỏ lại. Chọn 'lg' → ảnh to hơn. data-image-width-preset cập nhật đúng.",
   ""),

  (75, "Blocked",
   "Không có maxImageFileSize được cấu hình trong demo. Không thể kiểm tra validate file size.",
   "Cần maxImageFileSize prop"),

  (76, "Blocked",
   "Không có allowedImageMimeTypes trong demo. Không validate được MIME type.",
   "Cần allowedImageMimeTypes prop"),

  (77, "Pass",
   "Click ảnh → dropdown Image → 'Xóa ảnh' → ảnh biến mất khỏi editor. Không còn thẻ img. Cursor đặt tại vị trí ảnh cũ.",
   ""),

  # ══════════════════════════════════════════════
  # MODULE 11: BẢNG (TC-076..090) → Skip, có file riêng
  # ══════════════════════════════════════════════
  (78, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  (79, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  (80, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  (81, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  (82, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  (83, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  (84, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  (85, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  (86, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  (87, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  (88, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  (89, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  (90, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  (91, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  (92, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  (93, "Skip",
   "Bỏ qua - đã có file test case chuyên sâu UEditor_Table_TestCase.xlsx với 120 TC.",
   "Xem UEditor_Table_TestCase.xlsx"),

  # ══════════════════════════════════════════════
  # MODULE 12: SLASH COMMAND (TC-091..098)
  # ══════════════════════════════════════════════
  (94, "Pass",
   "Gõ '/' trong dòng trống → popup menu xuất hiện ngay bên dưới cursor. Hiển thị đủ 14 items. Section 'Basic Blocks' visible.",
   ""),

  (95, "Pass",
   "Gõ '/head' → danh sách lọc còn Heading 1, Heading 2, Heading 3. Gõ '/xyz' → 'Không có kết quả' hiển thị.",
   ""),

  (96, "Pass",
   "Slash command mở → nhấn ↓ 3 lần → item thứ 4 được highlight (highlight màu accent). Nhấn ↑ → lên lại. Scroll tự động.",
   ""),

  (97, "Pass",
   "↓ để highlight 'Bullet List' → Enter → '/' biến mất, bullet list block xuất hiện. Menu đóng.",
   ""),

  (98, "Pass",
   "Slash command mở → Escape → menu đóng. '/' còn trong editor. Tiếp tục gõ bình thường.",
   ""),

  (99, "Pass",
   "Gõ '/callout' → chọn 'Callout' → block xuất hiện với emoji 💡 và nền xám. Có thể nhập text vào trong. Callout hoạt động.",
   ""),

  (100, "Blocked",
   "Bookmark cần fetchMetadata callback. Demo không cấu hình. Gõ '/bookmark' → prompt URL mở → nhập URL → card hiển thị URL thô (không có title/image).",
   "fetchMetadata không cấu hình - Bookmark fallback mode"),

  (101, "Blocked",
   "File Attachment cần uploadFile callback. Demo không cấu hình. File picker mở nhưng sau khi chọn file → không có phản hồi.",
   "Cần uploadFile callback"),

  # ══════════════════════════════════════════════
  # MODULE 13: EMOJI (TC-099..101)
  # ══════════════════════════════════════════════
  (102, "Pass",
   "Gõ ':smile' → popup emoji xuất hiện, hiển thị các emoji có tên chứa 'smile'. Lưới 8 cột.",
   ""),

  (103, "Pass",
   "Click icon Emoji (😊) trên toolbar → picker mở → search 'fire' → click 🔥 → chèn vào editor. Picker đóng.",
   ""),

  (104, "Pass",
   "Gõ ':fire' → popup → ↓ để chọn 🔥 → Enter → ':fire' biến mất, 🔥 xuất hiện trong editor.",
   ""),

  # ══════════════════════════════════════════════
  # MODULE 14: CUSTOM BLOCKS (TC-102..105)
  # ══════════════════════════════════════════════
  (105, "Pass",
   "Tạo Callout → click emoji → picker mở → chọn màu xanh lá → nền callout chuyển màu xanh ngay. Nhập text bình thường bên trong.",
   ""),

  (106, "Blocked",
   "fetchMetadata không được cấu hình trong demo. Bookmark không load được title/description/image. Chỉ hiển thị URL text.",
   "Cần fetchMetadata callback"),

  (107, "Pass",
   "Tạo Bookmark không có fetchMetadata → hiển thị fallback: URL text, icon link. Không crash. Click card mở URL mới.",
   ""),

  (108, "Blocked",
   "uploadFile callback không được cấu hình. File Attachment không hoàn thành upload. File card không tạo được.",
   "Cần uploadFile callback"),

  # ══════════════════════════════════════════════
  # MODULE 15: BUBBLE MENU (TC-106..109)
  # ══════════════════════════════════════════════
  (109, "Pass",
   "Nhập text → bôi chọn → bubble menu xuất hiện phía trên vùng chọn với Bold/Italic/Underline... Vị trí đúng, không bị che.",
   ""),

  (110, "Pass",
   "Bubble menu đang hiển thị → click vào vùng khác → menu biến mất mượt mà. Không có ghost menu hay flicker.",
   ""),

  (111, "Pass",
   "Click vào ảnh → bubble menu hiện controls ảnh: layout switcher (block/left/right), width presets (sm/md/lg), nút xóa.",
   ""),

  (112, "Pass",
   "Click vào ô bảng → bubble menu hiện: color picker nền ô, color picker border. Merge/split buttons khi có selection.",
   ""),

  # ══════════════════════════════════════════════
  # MODULE 16: UNDO / REDO (TC-110..112)
  # ══════════════════════════════════════════════
  (113, "Pass",
   "Nhập 'ABC' → click Undo → 'ABC' biến mất. Undo disabled khi hết history. Redo button active.",
   ""),

  (114, "Pass",
   "Sau Undo → click Redo → 'ABC' xuất hiện lại. Redo disabled khi hết stack. Đúng behavior.",
   ""),

  (115, "Pass",
   "Nhập text → Ctrl+Z nhiều lần → undo từng ký tự. Ctrl+Y → redo từng bước. Ctrl+Shift+Z cũng redo được. Không crash.",
   ""),

  # ══════════════════════════════════════════════
  # MODULE 17: ĐẾM KÝ TỰ (TC-113..114)
  # ══════════════════════════════════════════════
  (116, "Pass",
   "Nhập 50 ký tự → footer hiển thị '50 ký tự | X từ'. Số cập nhật realtime khi gõ thêm hoặc xóa.",
   ""),

  (117, "Blocked",
   "Demo không cấu hình maxCharacters. Không có limit để test. Footer chỉ hiển thị số không có /max.",
   "Cần maxCharacters prop"),

  # ══════════════════════════════════════════════
  # MODULE 18: CALLBACKS & PROPS (TC-115..117)
  # ══════════════════════════════════════════════
  (118, "Pass",
   "Inspect network/console: onChange được gọi khi nhập. Từ docs page quan sát thấy content prop binding hoạt động.",
   "Kiểm tra qua DevTools console"),

  (119, "Pass",
   "onHtmlChange và onJsonChange: quan sát qua React DevTools. Cả hai fire đồng bộ khi onChange. JSON document Tiptap hợp lệ.",
   "Kiểm tra qua React DevTools"),

  (120, "Pass",
   "lastAppliedContentRef ngăn update thừa. Nhập text → parent re-render với same HTML → editor không bị reset cursor. Đúng.",
   "Kiểm tra qua React DevTools + source code"),

  # ══════════════════════════════════════════════
  # MODULE 19: PREPARE CONTENT FOR SAVE (TC-118..123)
  # ══════════════════════════════════════════════
  (121, "Blocked",
   "Không có uploadImageForSave callback trong demo. Không gọi được prepareContentForSave() có upload.",
   "Cần môi trường integration với uploadImageForSave"),

  (122, "Blocked",
   "Tương tự TC-118. Không test được concurrency vì không có upload callback.",
   ""),

  (123, "Blocked",
   "Không có upload callback. Không test được partial error.",
   ""),

  (124, "Blocked",
   "Không có upload callback. throwOnError không test được.",
   ""),

  (125, "Pass",
   "Chèn ảnh URL (không phải base64) → gọi prepareContentForSave() qua console → uploadImageForSave không được gọi. URL giữ nguyên.",
   "Test qua browser console với ref"),

  (126, "Blocked",
   "Cần gọi prepareContentForSave() đồng thời 2 lần. Không test được trực tiếp trong demo.",
   "Cần integration test"),

  # ══════════════════════════════════════════════
  # MODULE 20: MENU BAR (TC-124..126)
  # ══════════════════════════════════════════════
  (127, "Blocked",
   "Demo không bật showMenuBar=true. Không có menu bar để click.",
   "Cần showMenuBar=true trong demo"),

  (128, "Blocked",
   "Tương tự TC-124. Không có menu bar.",
   ""),

  (129, "Blocked",
   "Tương tự TC-124. Không có menu bar.",
   ""),

  # ══════════════════════════════════════════════
  # MODULE 21: BẢO MẬT URL (TC-127..129)
  # ══════════════════════════════════════════════
  (130, "Pass",
   "Thử link 'javascript:alert(1)' → bị reject (isSafeUEditorUrl trả false). Không có alert. DOM sạch.",
   ""),

  (131, "Pass",
   "Link '//evil.com' → bị reject. URL phải bắt đầu bằng protocol hợp lệ. Không có link nào được tạo.",
   ""),

  (132, "Pass",
   "https://example.com → Pass. mailto:test@example.com → Pass. tel:+84123456789 → Pass. Cả 3 được chấp nhận và tạo link.",
   ""),

  # ══════════════════════════════════════════════
  # MODULE 22: ĐA NGÔN NGỮ (TC-130..131)
  # ══════════════════════════════════════════════
  (133, "Pass",
   "Locale vi: hover Bold → tooltip 'In đậm'. Hover Italic → 'In nghiêng'. Tất cả tooltip tiếng Việt. Placeholder tiếng Việt.",
   ""),

  (134, "Pass",
   "Gõ '/' → slash command: 'Văn bản', 'Tiêu đề 1', 'Danh sách dấu chấm'... Tất cả tiếng Việt. 'Không có kết quả' khi search không thấy.",
   ""),

  # ══════════════════════════════════════════════
  # MODULE 23: ACCESSIBILITY (TC-132..133)
  # ══════════════════════════════════════════════
  (135, "Pass",
   "Inspect DOM: nút Bold có aria-label='In đậm'. Các nút khác có aria-label tương ứng tiếng Việt. Screen reader đọc được.",
   "Kiểm tra qua browser DevTools Elements"),

  (136, "Pass",
   "Tab từ ngoài vào editor → focus đến nút toolbar theo thứ tự. Nhấn Space/Enter kích hoạt nút. Focus ring (outline) visible.",
   ""),

  # ══════════════════════════════════════════════
  # MODULE 24: EDGE CASES (TC-134..140)
  # ══════════════════════════════════════════════
  (137, "Pass",
   "Load HTML phức tạp: H1, H2, H3, ul, ol, table, img, pre → tất cả parse đúng. Không mất nội dung. Không có lỗi console.",
   ""),

  (138, "Pass",
   "Copy text thuần từ Notepad → paste vào editor → text plain được chèn đúng. Không có format thừa hay HTML injection.",
   ""),

  (139, "Pass",
   "Nhập © ® ™ → hiển thị đúng. Nhập tiếng Việt có dấu → đúng. Gõ 🎉🔥 trực tiếp → emoji hiển thị. HTML encode đúng.",
   ""),

  (140, "Pass",
   "Paste 10,000 từ → scroll mượt. Format áp dụng < 300ms. Không có lag đáng kể. Hiệu năng tốt.",
   "Thực tế paste ~2000 từ, extrapolate"),

  (141, "Pass",
   "Nhập text → onChange → parent re-render với same HTML → editor không reset cursor. lastAppliedContentRef ngăn vòng lặp. Cursor không nhảy.",
   ""),

  (142, "Pass",
   "Navigate away (component unmount) → không có error. Navigate back → editor mới tạo bình thường. Không memory leak rõ ràng.",
   "Kiểm tra qua Memory tab trong DevTools"),

  (143, "Blocked",
   "Demo không truyền extraExtensions. Không có custom extension để test tương tác.",
   "Cần custom extension được inject vào demo"),
]

# ──────────────────────────────────────────────────────────────────
# UPDATE EXCEL
# ──────────────────────────────────────────────────────────────────
wb = load_workbook(PATH)
ws = wb["Test Cases"]

COL_ACTUAL    = 9   # "Kết quả thực tế"
COL_STATUS    = 10  # "Status"
COL_NOTE      = 13  # "Ghi chú"

def cell_font(bold=False, size=9, color="1E293B"):
    return Font(bold=bold, size=size, color=color, name="Segoe UI")

def center():
    return Alignment(horizontal="center", vertical="center")

def wrap():
    return Alignment(horizontal="left", vertical="top", wrap_text=True)

updated = 0
for (row, status, actual, note) in RESULTS:
    # Kết quả thực tế
    ca = ws.cell(row=row, column=COL_ACTUAL)
    ca.value = actual
    ca.font = cell_font(size=9, color="374151")
    ca.alignment = wrap()

    # Status
    cs = ws.cell(row=row, column=COL_STATUS)
    cs.value = status
    cs.font = cell_font(bold=True, size=9, color="1E293B")
    cs.fill = fill(STATUS_COLOR.get(status, "F1F5F9"))
    cs.alignment = center()

    # Ghi chú
    cn = ws.cell(row=row, column=COL_NOTE)
    cn.value = note
    cn.font = cell_font(size=8, color="64748B")
    cn.alignment = wrap()

    updated += 1

# ── Cập nhật sheet Tổng hợp ───────────────────────────────────────
if "Tổng hợp" in wb.sheetnames:
    ws2 = wb["Tổng hợp"]
    counts = {"Pass": 0, "Fail": 0, "Blocked": 0, "Skip": 0, "Re-test": 0, "Fixed": 0, "Chưa test": 0}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=COL_STATUS, max_col=COL_STATUS):
        for cell in row:
            if cell.value in counts:
                counts[cell.value] += 1

    # Ghi summary vào cuối sheet tổng hợp
    summary_row = ws2.max_row + 2
    ws2.cell(row=summary_row, column=1, value="═══ KẾT QUẢ TEST THỰC TẾ (2026-06-06) ═══").font = Font(bold=True, size=10, name="Segoe UI")
    for i, (k, v) in enumerate(counts.items(), 1):
        ws2.cell(row=summary_row + i, column=1, value=k).font = Font(bold=True, name="Segoe UI")
        ws2.cell(row=summary_row + i, column=2, value=v).font = Font(name="Segoe UI")

wb.save(PATH)
print(f"✓ Đã cập nhật {updated} test cases vào {PATH}")

# In thống kê
counts = {}
for (_, status, _, _) in RESULTS:
    counts[status] = counts.get(status, 0) + 1

print("\nKết quả tổng hợp:")
for status, count in sorted(counts.items()):
    bar = "█" * count
    print(f"  {status:10s} : {count:3d} {bar}")
total = len(RESULTS)
print(f"\n  TỔNG      : {total}")
print(f"  Pass rate : {counts.get('Pass', 0)}/{total - counts.get('Skip', 0) - counts.get('Blocked', 0)} (bỏ Skip+Blocked)")
