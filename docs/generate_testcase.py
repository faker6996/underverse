#!/usr/bin/env python3
"""
Script tạo file test case Excel cho UEditor component
Chạy: python3 docs/generate_testcase.py
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ──────────────────────────────────────────────
# COLOR PALETTE
# ──────────────────────────────────────────────
C_HEADER_BG     = "1E293B"   # slate-800
C_HEADER_FG     = "FFFFFF"
C_MODULE_BG     = "334155"   # slate-700
C_MODULE_FG     = "F1F5F9"
C_ALT_ROW       = "F8FAFC"   # slate-50
C_WHITE         = "FFFFFF"
C_BORDER        = "CBD5E1"   # slate-300

# Status colors
C_TODO          = "F1F5F9"   # slate-100  → Chưa test
C_PASS          = "DCFCE7"   # green-100
C_FAIL          = "FEE2E2"   # red-100
C_FIXED         = "FEF9C3"   # yellow-100
C_RETEST        = "DBEAFE"   # blue-100
C_BLOCKED       = "F3E8FF"   # purple-100
C_SKIP          = "F1F5F9"   # slate-100

# Priority colors
C_P1            = "FEE2E2"
C_P2            = "FEF9C3"
C_P3            = "DCFCE7"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=10, color="000000"):
    return Font(bold=True, size=size, color=color, name="Segoe UI")

def normal_font(size=10, color="1E293B"):
    return Font(size=size, color=color, name="Segoe UI")

def thin_border():
    s = Side(border_style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def wrap_align(h="left", v="top"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)

# ──────────────────────────────────────────────
# TEST DATA
# ──────────────────────────────────────────────
COLUMNS = [
    ("STT",                  4),
    ("Module",               14),
    ("Mã TC",                10),
    ("Tên Test Case",        36),
    ("Mô tả",                42),
    ("Điều kiện\ntiên quyết", 28),
    ("Các bước thực hiện",   58),
    ("Kết quả mong muốn",    52),
    ("Kết quả thực tế",      32),
    ("Status",               12),
    ("Ưu tiên",              9),
    ("Loại test",            13),
    ("Ghi chú",              22),
]

# (stt, module, ma_tc, ten_tc, mo_ta, dieu_kien, cac_buoc, ket_qua_mong_muon, loai_test, uu_tien)
TEST_CASES = [
    # ─── MODULE 1: KHỞI TẠO & HIỂN THỊ ───────────────────────────────────────
    (1,  "01. Khởi tạo & Hiển thị", "TC-001",
     "Render component với props mặc định",
     "Kiểm tra component UEditor render thành công với props mặc định, không báo lỗi.",
     "Môi trường Next.js, React 19. UEditor được import đúng.",
     "1. Mount component <UEditor />\n2. Không truyền bất kỳ prop nào\n3. Quan sát DOM được render",
     "- Component render thành công, không throw error\n- Hiển thị vùng editor có thể nhập liệu\n- Toolbar hiển thị mặc định\n- Không có lỗi console",
     "Functional", "P1"),

    (2,  "01. Khởi tạo & Hiển thị", "TC-002",
     "Hiển thị loading state trước khi editor khởi tạo",
     "Kiểm tra placeholder loading khi Tiptap chưa khởi tạo xong.",
     "Server-side render hoặc lần render đầu tiên.",
     "1. Mount <UEditor />\n2. Quan sát trạng thái ngay khi component mount\n3. Chờ editor khởi tạo xong",
     "- Trong thời gian khởi tạo: hiển thị text 'Đang tải...' (hoặc translation loading key)\n- Sau khi khởi tạo: text biến mất, editor xuất hiện",
     "Functional", "P2"),

    (3,  "01. Khởi tạo & Hiển thị", "TC-003",
     "Hiển thị placeholder khi content rỗng",
     "Kiểm tra placeholder text hiển thị đúng khi editor trống.",
     "UEditor mount với content='' hoặc không truyền content.",
     "1. Mount <UEditor placeholder='Nhập nội dung...' />\n2. Không nhập gì\n3. Quan sát vùng editor",
     "- Placeholder text 'Nhập nội dung...' hiển thị trong vùng editor\n- Placeholder ẩn đi khi bắt đầu nhập\n- Placeholder không hiển thị bên trong ô bảng",
     "UI", "P2"),

    (4,  "01. Khởi tạo & Hiển thị", "TC-004",
     "Render với content HTML ban đầu",
     "Kiểm tra editor load đúng nội dung HTML được truyền qua prop content.",
     "Có sẵn chuỗi HTML hợp lệ: '<p>Xin chào <strong>thế giới</strong></p>'",
     "1. Mount <UEditor content='<p>Xin chào <strong>thế giới</strong></p>' />\n2. Quan sát nội dung hiển thị trong editor",
     "- Nội dung 'Xin chào thế giới' hiển thị\n- 'thế giới' được hiển thị in đậm\n- Cấu trúc DOM khớp với HTML input",
     "Functional", "P1"),

    (5,  "01. Khởi tạo & Hiển thị", "TC-005",
     "Render ở chế độ read-only (editable=false)",
     "Kiểm tra editor không cho phép chỉnh sửa khi editable=false.",
     "Mount <UEditor editable={false} content='<p>Nội dung test</p>' />",
     "1. Mount với editable={false} và content có sẵn\n2. Thử click vào vùng editor\n3. Thử nhập text\n4. Thử click link trong nội dung",
     "- Không thể nhập text\n- Toolbar không hiển thị (bị ẩn)\n- Bubble menu không xuất hiện\n- Click link mở tab mới\n- Không có border focus\n- CSS class 'rounded-none' được áp dụng",
     "Functional", "P1"),

    (6,  "01. Khởi tạo & Hiển thị", "TC-006",
     "Render với minHeight và maxHeight",
     "Kiểm tra kích thước editor được áp dụng đúng.",
     "Mount UEditor với các giá trị kích thước.",
     "1. Mount <UEditor minHeight='300px' maxHeight='600px' />\n2. Quan sát chiều cao vùng editor\n3. Nhập nội dung dài hơn maxHeight\n4. Quan sát scroll behavior",
     "- Vùng editor có minHeight=300px\n- Khi nội dung vượt maxHeight=600px: xuất hiện scrollbar\n- Nội dung không bị tràn ra ngoài container",
     "UI", "P2"),

    (7,  "01. Khởi tạo & Hiển thị", "TC-007",
     "Variant 'default' - giao diện chuẩn",
     "Kiểm tra variant default hiển thị đầy đủ border, shadow, toolbar.",
     "Mount <UEditor variant='default' />",
     "1. Mount với variant='default'\n2. Kiểm tra class CSS được áp dụng\n3. Quan sát khi hover, focus",
     "- Border và background được áp dụng\n- Shadow hiện khi focus\n- Toolbar đầy đủ hiển thị\n- Transition animation khi focus",
     "UI", "P3"),

    (8,  "01. Khởi tạo & Hiển thị", "TC-008",
     "Variant 'minimal' - toolbar tối giản",
     "Kiểm tra variant minimal chỉ hiển thị Bold, Italic, Bullet list.",
     "Mount <UEditor variant='minimal' />",
     "1. Mount với variant='minimal'\n2. Quan sát toolbar",
     "- Toolbar chỉ có 3 nút: Bold, Italic, Bullet List\n- Không có các nút khác (font, color, table...)",
     "UI", "P2"),

    (9,  "01. Khởi tạo & Hiển thị", "TC-009",
     "Variant 'notion' - hover shadow effect",
     "Kiểm tra variant notion có hover shadow.",
     "Mount <UEditor variant='notion' />",
     "1. Mount với variant='notion'\n2. Hover chuột vào component\n3. Quan sát CSS class",
     "- Class 'hover:shadow-md' được áp dụng\n- Shadow xuất hiện khi hover",
     "UI", "P3"),

    (10, "01. Khởi tạo & Hiển thị", "TC-010",
     "Ẩn toolbar (showToolbar=false)",
     "Kiểm tra toolbar bị ẩn hoàn toàn.",
     "Mount <UEditor showToolbar={false} />",
     "1. Mount với showToolbar={false}\n2. Quan sát UI",
     "- Không có thanh toolbar\n- Vùng editor vẫn hoạt động bình thường\n- Bubble menu vẫn hiển thị khi bôi chọn",
     "Functional", "P2"),

    (11, "01. Khởi tạo & Hiển thị", "TC-011",
     "Ẩn bubble menu (showBubbleMenu=false)",
     "Kiểm tra bubble menu không hiển thị khi bôi chọn text.",
     "Mount <UEditor showBubbleMenu={false} />",
     "1. Mount với showBubbleMenu={false}\n2. Nhập text vào editor\n3. Bôi chọn text vừa nhập",
     "- Không có bubble menu xuất hiện khi bôi chọn\n- Editor hoạt động bình thường",
     "Functional", "P2"),

    (12, "01. Khởi tạo & Hiển thị", "TC-012",
     "Ẩn character count (showCharacterCount=false)",
     "Kiểm tra ô đếm ký tự không hiển thị.",
     "Mount <UEditor showCharacterCount={false} />",
     "1. Mount với showCharacterCount={false}\n2. Quan sát footer editor",
     "- Không có dòng đếm ký tự/từ ở cuối editor",
     "UI", "P3"),

    (13, "01. Khởi tạo & Hiển thị", "TC-013",
     "Hiển thị menu bar (showMenuBar=true)",
     "Kiểm tra menu bar kiểu Word hiển thị đúng.",
     "Mount <UEditor showMenuBar={true} />",
     "1. Mount với showMenuBar={true}\n2. Quan sát phần trên editor\n3. Click vào từng menu",
     "- Menu bar hiển thị với các mục: File, Edit, View, Insert, Format, Table\n- Mỗi menu mở dropdown khi click\n- Các action trong menu hoạt động",
     "Functional", "P2"),

    (14, "01. Khởi tạo & Hiển thị", "TC-014",
     "autofocus=true tự động focus editor",
     "Kiểm tra editor tự focus khi mount với autofocus=true.",
     "Mount <UEditor autofocus={true} />",
     "1. Mount với autofocus={true}\n2. Không click vào đâu\n3. Bắt đầu gõ phím",
     "- Editor đã được focus sẵn\n- Gõ phím ngay lập tức có thể nhập text\n- Cursor hiển thị trong editor",
     "Functional", "P3"),

    # ─── MODULE 2: ĐỊNH DẠNG VĂN BẢN ──────────────────────────────────────────
    (15, "02. Định dạng văn bản", "TC-015",
     "In đậm (Bold) - click toolbar button",
     "Kiểm tra toggle Bold qua nút toolbar.",
     "Editor đang focus, đã chọn text.",
     "1. Nhập text 'Hello World'\n2. Bôi chọn 'Hello'\n3. Click nút Bold (B) trên toolbar",
     "- 'Hello' được in đậm\n- Nút Bold sáng lên (active state)\n- HTML output: <strong>Hello</strong> World",
     "Functional", "P1"),

    (16, "02. Định dạng văn bản", "TC-016",
     "In đậm - phím tắt Ctrl+B",
     "Kiểm tra phím tắt Ctrl+B bật/tắt Bold.",
     "Editor đang focus, có text được chọn.",
     "1. Nhập text 'Hello World'\n2. Bôi chọn toàn bộ text\n3. Nhấn Ctrl+B\n4. Nhấn Ctrl+B lần 2",
     "- Lần 1: Toàn bộ text in đậm\n- Lần 2: Tắt in đậm (toggle)\n- Nút Bold trên toolbar phản ánh trạng thái hiện tại",
     "Functional", "P1"),

    (17, "02. Định dạng văn bản", "TC-017",
     "In nghiêng (Italic) - click toolbar",
     "Kiểm tra toggle Italic qua nút toolbar.",
     "Editor đang focus, text được chọn.",
     "1. Nhập text 'Hello World'\n2. Bôi chọn 'World'\n3. Click nút Italic (I) trên toolbar",
     "- 'World' hiển thị in nghiêng\n- Nút Italic ở trạng thái active\n- HTML output: Hello <em>World</em>",
     "Functional", "P1"),

    (18, "02. Định dạng văn bản", "TC-018",
     "In nghiêng - phím tắt Ctrl+I",
     "Kiểm tra phím tắt Ctrl+I.",
     "Editor focus, text được chọn.",
     "1. Bôi chọn text\n2. Nhấn Ctrl+I",
     "- Text được in nghiêng\n- Toggle lại bằng Ctrl+I: trở về bình thường",
     "Functional", "P1"),

    (19, "02. Định dạng văn bản", "TC-019",
     "Gạch chân (Underline) - click toolbar",
     "Kiểm tra toggle Underline.",
     "Editor focus, text được chọn.",
     "1. Bôi chọn text\n2. Click nút Underline (U)",
     "- Text có gạch chân\n- HTML: <u>text</u>",
     "Functional", "P1"),

    (20, "02. Định dạng văn bản", "TC-020",
     "Gạch chân - phím tắt Ctrl+U",
     "Kiểm tra phím tắt Ctrl+U cho Underline.",
     "Editor focus, text được chọn.",
     "1. Bôi chọn text\n2. Nhấn Ctrl+U",
     "- Text gạch chân, toggle được bằng Ctrl+U lần 2",
     "Functional", "P1"),

    (21, "02. Định dạng văn bản", "TC-021",
     "Gạch ngang (Strikethrough) - click toolbar",
     "Kiểm tra toggle Strikethrough.",
     "Editor focus, text được chọn.",
     "1. Bôi chọn text\n2. Click nút Strikethrough (S̶)",
     "- Text có gạch ngang giữa\n- HTML: <s>text</s>",
     "Functional", "P2"),

    (22, "02. Định dạng văn bản", "TC-022",
     "Code inline - click toolbar",
     "Kiểm tra toggle Code inline.",
     "Editor focus, text được chọn.",
     "1. Bôi chọn text 'const x = 1'\n2. Click nút Code (<>)",
     "- Text hiển thị dạng monospace với background\n- HTML: <code>const x = 1</code>",
     "Functional", "P2"),

    (23, "02. Định dạng văn bản", "TC-023",
     "Subscript - click toolbar",
     "Kiểm tra Subscript.",
     "Editor focus, text được chọn.",
     "1. Nhập 'H2O'\n2. Chọn '2'\n3. Click nút Subscript",
     "- '2' hiển thị nhỏ hơn, thấp hơn dòng chữ\n- HTML: H<sub>2</sub>O",
     "Functional", "P3"),

    (24, "02. Định dạng văn bản", "TC-024",
     "Superscript - click toolbar",
     "Kiểm tra Superscript.",
     "Editor focus, text được chọn.",
     "1. Nhập 'x2'\n2. Chọn '2'\n3. Click nút Superscript",
     "- '2' hiển thị nhỏ hơn, cao hơn dòng chữ\n- HTML: x<sup>2</sup>",
     "Functional", "P3"),

    (25, "02. Định dạng văn bản", "TC-025",
     "Toggle format - bật/tắt Bold nhiều lần",
     "Kiểm tra toggle hoạt động nhất quán.",
     "Editor focus.",
     "1. Nhập text\n2. Bôi chọn\n3. Click Bold 5 lần, mỗi lần quan sát",
     "- Lần lẻ: text in đậm\n- Lần chẵn: text thường\n- Nút toolbar luôn phản ánh đúng trạng thái",
     "Functional", "P2"),

    (26, "02. Định dạng văn bản", "TC-026",
     "Kết hợp nhiều format (Bold + Italic + Underline)",
     "Kiểm tra áp dụng nhiều format cùng lúc.",
     "Editor focus, text được chọn.",
     "1. Bôi chọn text\n2. Lần lượt click Bold, Italic, Underline\n3. Quan sát kết quả",
     "- Text in đậm + nghiêng + gạch chân đồng thời\n- HTML: <strong><em><u>text</u></em></strong>\n- Cả 3 nút toolbar ở trạng thái active",
     "Functional", "P2"),

    (27, "02. Định dạng văn bản", "TC-027",
     "Xóa toàn bộ định dạng (Clear Formatting)",
     "Kiểm tra chức năng xóa format qua Menu Bar.",
     "showMenuBar=true, text đang có nhiều format.",
     "1. Áp dụng Bold + Italic + Color cho text\n2. Bôi chọn toàn bộ text\n3. Format > Clear Formatting",
     "- Toàn bộ format bị xóa\n- Text trở về Normal, màu mặc định\n- Không còn style HTML nào",
     "Functional", "P2"),

    # ─── MODULE 3: TYPOGRAPHY ──────────────────────────────────────────────────
    (28, "03. Typography", "TC-028",
     "Đổi font family",
     "Kiểm tra thay đổi font family qua dropdown.",
     "Editor focus, cursor đặt trong đoạn text.",
     "1. Đặt cursor vào text\n2. Click dropdown Font Family trên toolbar\n3. Chọn font 'Arial'",
     "- Text trong đoạn văn đổi sang font Arial\n- Dropdown hiển thị tên font đang dùng\n- style='font-family: Arial' được áp dụng",
     "Functional", "P2"),

    (29, "03. Typography", "TC-029",
     "Reset font family về mặc định",
     "Kiểm tra reset font về mặc định.",
     "Text đang dùng font tùy chỉnh.",
     "1. Chọn text đang dùng font Arial\n2. Mở dropdown Font Family\n3. Click 'Mặc định'",
     "- Font size trả về mặc định hệ thống\n- Không còn style font-family\n- Dropdown hiển thị text mặc định",
     "Functional", "P2"),

    (30, "03. Typography", "TC-030",
     "Đổi font size",
     "Kiểm tra thay đổi cỡ chữ qua dropdown.",
     "Editor focus, cursor trong text.",
     "1. Bôi chọn text\n2. Click dropdown Font Size\n3. Chọn '18px'",
     "- Text được phóng to lên 18px\n- style='font-size: 18px'\n- Dropdown hiện '18px'",
     "Functional", "P2"),

    (31, "03. Typography", "TC-031",
     "Reset font size về mặc định",
     "Kiểm tra reset cỡ chữ.",
     "Text đang dùng font size tùy chỉnh.",
     "1. Chọn text đang 18px\n2. Mở dropdown Font Size\n3. Click 'Mặc định'",
     "- Font size trả về mặc định\n- Không còn style font-size",
     "Functional", "P2"),

    (32, "03. Typography", "TC-032",
     "Đổi line height (giãn dòng)",
     "Kiểm tra thay đổi giãn dòng.",
     "Editor focus, cursor trong đoạn văn.",
     "1. Đặt cursor trong đoạn văn\n2. Click dropdown Line Height\n3. Chọn '2.0'",
     "- Khoảng cách giữa các dòng tăng lên\n- style='line-height: 2.0'\n- Dropdown hiển thị giá trị đang chọn",
     "Functional", "P2"),

    (33, "03. Typography", "TC-033",
     "Đổi letter spacing (giãn chữ)",
     "Kiểm tra thay đổi khoảng cách chữ.",
     "Editor focus, text được chọn.",
     "1. Bôi chọn text\n2. Click dropdown Letter Spacing\n3. Chọn '0.1em'",
     "- Khoảng cách giữa các ký tự tăng\n- style='letter-spacing: 0.1em'",
     "Functional", "P3"),

    (34, "03. Typography", "TC-034",
     "Custom font families qua prop fontFamilies",
     "Kiểm tra custom danh sách font được truyền qua prop.",
     "Mount UEditor với fontFamilies prop.",
     "1. Mount <UEditor fontFamilies={[{label:'Roboto',value:'Roboto'},{label:'Montserrat',value:'Montserrat'}]} />\n2. Mở dropdown Font Family",
     "- Dropdown chỉ hiển thị 2 font: Roboto và Montserrat\n- Không hiển thị font mặc định của hệ thống",
     "Functional", "P3"),

    # ─── MODULE 4: HEADING ─────────────────────────────────────────────────────
    (35, "04. Heading", "TC-035",
     "Chuyển đổi sang Heading 1",
     "Kiểm tra toggle Heading 1 qua dropdown Text Style.",
     "Editor focus, cursor trong đoạn văn.",
     "1. Đặt cursor trong đoạn văn\n2. Click dropdown Text Style (T↓)\n3. Chọn 'Heading 1'",
     "- Đoạn văn chuyển thành H1 (chữ lớn, đậm)\n- HTML: <h1>nội dung</h1>\n- Dropdown hiển thị 'Heading 1'",
     "Functional", "P1"),

    (36, "04. Heading", "TC-036",
     "Heading - phím tắt Ctrl+Alt+1/2/3",
     "Kiểm tra phím tắt tạo Heading.",
     "Editor focus, cursor trong paragraph.",
     "1. Nhấn Ctrl+Alt+1\n2. Nhấn Ctrl+Alt+2\n3. Nhấn Ctrl+Alt+3",
     "- Ctrl+Alt+1: paragraph → H1\n- Ctrl+Alt+2: → H2\n- Ctrl+Alt+3: → H3",
     "Functional", "P2"),

    (37, "04. Heading", "TC-037",
     "Toggle Heading (bật/tắt)",
     "Kiểm tra Heading toggle về Normal text.",
     "Cursor đang trong H1.",
     "1. Chọn Heading 1 cho paragraph\n2. Chọn lại Heading 1 (toggle off)",
     "- Heading chuyển về Normal paragraph\n- HTML: <p> thay vì <h1>",
     "Functional", "P2"),

    # ─── MODULE 5: MÀU SẮC ────────────────────────────────────────────────────
    (38, "05. Màu sắc", "TC-038",
     "Đổi màu chữ (Text Color)",
     "Kiểm tra thay đổi màu chữ qua color palette.",
     "Editor focus, text được chọn.",
     "1. Bôi chọn text\n2. Click icon màu chữ (A̲) trên toolbar\n3. Click màu đỏ trong palette",
     "- Text chuyển sang màu đỏ\n- style='color: #...' được áp dụng\n- Icon màu chữ trên toolbar đổi màu tương ứng",
     "Functional", "P1"),

    (39, "05. Màu sắc", "TC-039",
     "Reset màu chữ về mặc định",
     "Kiểm tra xóa màu chữ.",
     "Text đang có màu tùy chỉnh.",
     "1. Chọn text đang có màu\n2. Mở color palette\n3. Click 'Mặc định' (ô đầu tiên)",
     "- Màu chữ về mặc định (inherit)\n- Không còn style color",
     "Functional", "P2"),

    (40, "05. Màu sắc", "TC-040",
     "Đổi màu highlight (nền chữ)",
     "Kiểm tra tô nền chữ.",
     "Editor focus, text được chọn.",
     "1. Bôi chọn text\n2. Click icon Highlight trên toolbar\n3. Chọn màu vàng",
     "- Text được tô nền vàng\n- HTML: <mark style='background-color: yellow'>text</mark>",
     "Functional", "P1"),

    (41, "05. Màu sắc", "TC-041",
     "Reset màu highlight",
     "Kiểm tra xóa highlight.",
     "Text đang có highlight.",
     "1. Chọn text đang highlight\n2. Mở highlight palette\n3. Click 'Không tô màu'",
     "- Highlight bị xóa\n- Text hiển thị bình thường",
     "Functional", "P2"),

    (42, "05. Màu sắc", "TC-042",
     "Color palette - semantic colors",
     "Kiểm tra các màu semantic (primary, success, warning...) hiển thị đúng.",
     "Mở color palette text color.",
     "1. Mở dropdown Text Color\n2. Quan sát các màu semantic ở đầu palette\n3. Chọn màu 'primary'",
     "- Hiển thị ít nhất 6 màu semantic: primary, secondary, success, warning, destructive, info\n- Chọn được và áp dụng đúng\n- 39 màu swatch hiển thị đầy đủ",
     "UI", "P3"),

    # ─── MODULE 6: CĂN LỀ ─────────────────────────────────────────────────────
    (43, "06. Căn lề", "TC-043",
     "Căn trái (Align Left)",
     "Kiểm tra căn lề trái.",
     "Editor focus, cursor trong paragraph.",
     "1. Đặt cursor trong đoạn văn\n2. Mở dropdown Alignment\n3. Chọn 'Căn trái'",
     "- Text căn trái\n- style='text-align: left'\n- Nút căn trái ở trạng thái active",
     "Functional", "P2"),

    (44, "06. Căn lề", "TC-044",
     "Căn giữa (Align Center)",
     "Kiểm tra căn lề giữa.",
     "Editor focus, cursor trong paragraph.",
     "1. Đặt cursor\n2. Chọn 'Căn giữa'",
     "- Text căn giữa\n- style='text-align: center'",
     "Functional", "P2"),

    (45, "06. Căn lề", "TC-045",
     "Căn phải (Align Right)",
     "Kiểm tra căn phải.",
     "Editor focus, cursor trong paragraph.",
     "1. Chọn 'Căn phải'",
     "- Text căn phải\n- style='text-align: right'",
     "Functional", "P2"),

    (46, "06. Căn lề", "TC-046",
     "Justify (Căn đều 2 bên)",
     "Kiểm tra justify alignment.",
     "Editor focus, đoạn văn dài nhiều dòng.",
     "1. Đặt cursor trong đoạn văn dài\n2. Chọn 'Căn đều'",
     "- Text căn đều 2 bên\n- style='text-align: justify'",
     "Functional", "P3"),

    # ─── MODULE 7: DANH SÁCH ──────────────────────────────────────────────────
    (47, "07. Danh sách", "TC-047",
     "Tạo Bullet List từ toolbar",
     "Kiểm tra tạo danh sách không có số.",
     "Editor focus.",
     "1. Đặt cursor vào dòng trống\n2. Click dropdown List\n3. Chọn 'Bullet List'",
     "- Xuất hiện bullet (•) đầu dòng\n- HTML: <ul><li>...</li></ul>\n- CSS class: list-disc pl-6",
     "Functional", "P1"),

    (48, "07. Danh sách", "TC-048",
     "Tạo Ordered List từ toolbar",
     "Kiểm tra danh sách có số.",
     "Editor focus.",
     "1. Click dropdown List\n2. Chọn 'Numbered List'",
     "- Hiển thị danh sách 1. 2. 3.\n- HTML: <ol><li>...</li></ol>",
     "Functional", "P1"),

    (49, "07. Danh sách", "TC-049",
     "Tạo Task List (Todo) từ toolbar",
     "Kiểm tra danh sách checkbox.",
     "Editor focus.",
     "1. Click dropdown List\n2. Chọn 'Todo List'",
     "- Hiển thị ô checkbox trước mỗi mục\n- HTML: <ul data-type='taskList'>\n- Checkbox có thể tick được",
     "Functional", "P1"),

    (50, "07. Danh sách", "TC-050",
     "Phím tắt Ctrl+Shift+8 (Bullet List)",
     "Kiểm tra phím tắt tạo Bullet List.",
     "Editor focus, cursor trong paragraph.",
     "1. Nhấn Ctrl+Shift+8",
     "- Paragraph chuyển thành Bullet List",
     "Functional", "P2"),

    (51, "07. Danh sách", "TC-051",
     "Input rule '- ' tạo Bullet List",
     "Kiểm tra markdown-style input rule.",
     "Editor focus, dòng trống.",
     "1. Gõ '- ' (gạch ngang + dấu cách)",
     "- Tự động chuyển thành bullet list\n- '- ' biến mất, thay bằng bullet",
     "Functional", "P2"),

    (52, "07. Danh sách", "TC-052",
     "Input rule '1. ' tạo Ordered List",
     "Kiểm tra markdown-style input rule.",
     "Editor focus, dòng trống.",
     "1. Gõ '1. ' (số một + dấu chấm + dấu cách)",
     "- Tự động chuyển thành ordered list\n- '1. ' biến mất",
     "Functional", "P2"),

    (53, "07. Danh sách", "TC-053",
     "Input rule '[ ] ' tạo Task List",
     "Kiểm tra markdown-style task list.",
     "Editor focus, dòng trống.",
     "1. Gõ '[ ] ' hoặc '[] '",
     "- Chuyển thành task list item\n- Checkbox chưa tick",
     "Functional", "P2"),

    (54, "07. Danh sách", "TC-054",
     "Tick/untick checkbox trong Task List",
     "Kiểm tra tương tác checkbox.",
     "Task list đã được tạo, có ít nhất 2 item.",
     "1. Click vào checkbox đầu tiên\n2. Click lại để bỏ tick",
     "- Click 1: checkbox được tick, item có style gạch ngang hoặc mờ\n- Click 2: checkbox bỏ tick\n- HTML: data-checked='true'/'false'",
     "Functional", "P1"),

    (55, "07. Danh sách", "TC-055",
     "Nested list (danh sách lồng nhau)",
     "Kiểm tra tạo list con bên trong list.",
     "Bullet list đã có ít nhất 1 item.",
     "1. Đặt cursor cuối một item\n2. Nhấn Enter để tạo item mới\n3. Nhấn Tab để indent",
     "- Xuất hiện list con (nested)\n- Item được indent vào trong\n- Vẫn có bullet",
     "Functional", "P2"),

    # ─── MODULE 8: BLOCKQUOTE & CODE BLOCK ────────────────────────────────────
    (56, "08. Block Content", "TC-056",
     "Tạo Blockquote từ toolbar",
     "Kiểm tra insert blockquote.",
     "Editor focus.",
     "1. Click dropdown Quote\n2. Chọn 'Quote'",
     "- Xuất hiện khối blockquote với border trái\n- HTML: <blockquote>...\n- CSS: border-l-4 border-primary italic",
     "Functional", "P2"),

    (57, "08. Block Content", "TC-057",
     "Tạo Code Block từ toolbar",
     "Kiểm tra insert code block có syntax highlighting.",
     "Editor focus.",
     "1. Click dropdown Quote\n2. Chọn 'Code Block'",
     "- Xuất hiện code block với nền tối\n- Dropdown chọn ngôn ngữ\n- Nút copy xuất hiện",
     "Functional", "P1"),

    (58, "08. Block Content", "TC-058",
     "Code Block - chọn ngôn ngữ",
     "Kiểm tra dropdown chọn ngôn ngữ trong code block.",
     "Code block đã được tạo.",
     "1. Click vào dropdown ngôn ngữ trong code block\n2. Chọn 'Python'\n3. Nhập code Python",
     "- Ngôn ngữ chuyển sang Python\n- Syntax highlighting áp dụng theo Python\n- Token màu sắc đúng với Python syntax",
     "Functional", "P2"),

    (59, "08. Block Content", "TC-059",
     "Code Block - copy button",
     "Kiểm tra nút copy code.",
     "Code block có nội dung.",
     "1. Nhập code vào code block\n2. Click nút copy (icon clipboard)\n3. Paste vào text editor khác",
     "- Code được copy vào clipboard\n- Nút copy hiển thị feedback (icon check hoặc text 'Đã sao chép')\n- Paste cho ra đúng code đã nhập",
     "Functional", "P2"),

    (60, "08. Block Content", "TC-060",
     "Horizontal Rule (Divider)",
     "Kiểm tra chèn đường kẻ ngang.",
     "Editor focus.",
     "1. Gõ '/' để mở slash command\n2. Chọn 'Divider'",
     "- Xuất hiện đường kẻ ngang phân cách\n- HTML: <hr>\n- Có thể tiếp tục nhập sau divider",
     "Functional", "P3"),

    # ─── MODULE 9: LIÊN KẾT ───────────────────────────────────────────────────
    (61, "09. Liên kết", "TC-061",
     "Thêm link qua toolbar",
     "Kiểm tra chèn hyperlink.",
     "Editor focus, text được chọn.",
     "1. Bôi chọn text 'Click here'\n2. Click nút Link (🔗) trên toolbar\n3. Trong dropdown: click 'Link'\n4. Nhập URL 'https://example.com'\n5. Click Submit",
     "- Text 'Click here' trở thành hyperlink\n- Màu primary, gạch chân\n- HTML: <a href='https://example.com'>Click here</a>",
     "Functional", "P1"),

    (62, "09. Liên kết", "TC-062",
     "Xóa link qua toolbar",
     "Kiểm tra xóa hyperlink.",
     "Cursor đang trong text có link.",
     "1. Click vào link\n2. Mở dropdown Link\n3. Click 'Xóa Link'",
     "- Link bị xóa, text vẫn giữ nguyên\n- Không còn thẻ <a>\n- Nút xóa disabled khi không có link",
     "Functional", "P2"),

    (63, "09. Liên kết", "TC-063",
     "Link mở tab mới khi read-only",
     "Kiểm tra click link trong mode read-only.",
     "editable=false, content có link.",
     "1. Mount với editable={false}\n2. Click vào link\n3. Quan sát browser",
     "- Tab mới được mở với URL của link\n- noopener,noreferrer được thêm\n- Không bôi chọn thì mới mở (khi không select text)",
     "Functional", "P1"),

    (64, "09. Liên kết", "TC-064",
     "URL validation - chặn javascript: protocol",
     "Kiểm tra bảo mật URL.",
     "Toolbar link input đang mở.",
     "1. Thêm link với URL 'javascript:alert(1)'\n2. Submit",
     "- Link bị từ chối\n- Không có <a href='javascript:...'> trong DOM\n- Không có alert xuất hiện",
     "Functional", "P1"),

    (65, "09. Liên kết", "TC-065",
     "Bubble menu - Link preview",
     "Kiểm tra link preview trong bubble menu.",
     "Editor có text dạng link, showBubbleMenu=true.",
     "1. Click vào text có link\n2. Quan sát bubble menu",
     "- Bubble menu hiển thị preview URL\n- Có nút 'Mở link', 'Chỉnh sửa', 'Xóa'\n- Click 'Mở link' mở tab mới",
     "Functional", "P2"),

    # ─── MODULE 10: HÌNH ẢNH ──────────────────────────────────────────────────
    (66, "10. Hình ảnh", "TC-066",
     "Chèn ảnh qua URL",
     "Kiểm tra chèn ảnh từ URL.",
     "Editor focus, imageInsertMode mặc định.",
     "1. Click dropdown Image trên toolbar\n2. Chọn 'Thêm từ URL'\n3. Nhập URL ảnh hợp lệ\n4. Nhập alt text 'Logo'\n5. Submit",
     "- Ảnh hiển thị trong editor\n- Alt text được lưu\n- HTML: <img src='url' alt='Logo'>",
     "Functional", "P1"),

    (67, "10. Hình ảnh", "TC-067",
     "Chèn ảnh qua upload file",
     "Kiểm tra upload ảnh từ máy.",
     "Editor focus, uploadImage callback được cấu hình.",
     "1. Click dropdown Image\n2. Chọn 'Tải lên'\n3. Chọn file ảnh (JPG, PNG, WEBP)\n4. Quan sát kết quả",
     "- File picker mở\n- Ảnh được upload qua uploadImage callback\n- Ảnh hiển thị trong editor sau upload\n- Trạng thái 'Đang tải...' hiện trong quá trình upload",
     "Functional", "P1"),

    (68, "10. Hình ảnh", "TC-068",
     "Paste ảnh từ clipboard",
     "Kiểm tra paste ảnh Ctrl+V.",
     "Editor focus, clipboard có ảnh (copy từ browser hoặc app).",
     "1. Copy ảnh từ nguồn khác\n2. Focus vào editor\n3. Nhấn Ctrl+V",
     "- Ảnh được chèn vào editor\n- Với imageInsertMode='base64': ảnh base64\n- Với imageInsertMode='upload': gọi uploadImage callback",
     "Functional", "P1"),

    (69, "10. Hình ảnh", "TC-069",
     "Kéo thả ảnh vào editor",
     "Kiểm tra drag & drop ảnh.",
     "Editor visible, file ảnh từ File Explorer.",
     "1. Kéo file ảnh từ máy tính\n2. Thả vào vùng editor",
     "- Ảnh được chèn vào vị trí thả\n- Xử lý tương tự paste: base64 hoặc upload tùy config",
     "Functional", "P1"),

    (70, "10. Hình ảnh", "TC-070",
     "Resize ảnh bằng kéo thả",
     "Kiểm tra resize ảnh interactive.",
     "Editor có ảnh được chèn vào.",
     "1. Click vào ảnh để chọn\n2. Kéo handle resize ở góc\n3. Thả chuột",
     "- Ảnh thay đổi kích thước\n- Tỉ lệ được giữ nguyên (aspect ratio)\n- Kích thước tối thiểu 40px\n- Width/height được set trong style",
     "Functional", "P2"),

    (71, "10. Hình ảnh", "TC-071",
     "Đổi layout ảnh - Float Left",
     "Kiểm tra float ảnh sang trái.",
     "Editor có ảnh, ảnh đang được chọn.",
     "1. Click vào ảnh\n2. Click dropdown Image\n3. Chọn 'Căn trái (Float)'",
     "- Ảnh float sang trái\n- Text wrap bên phải ảnh\n- data-image-layout='left'",
     "Functional", "P2"),

    (72, "10. Hình ảnh", "TC-072",
     "Đổi width preset (sm/md/lg)",
     "Kiểm tra các preset kích thước ảnh.",
     "Ảnh đang được chọn.",
     "1. Click dropdown Image\n2. Chọn 'Nhỏ (sm)'\n3. Chọn 'Vừa (md)'\n4. Chọn 'Lớn (lg)'",
     "- sm: ảnh chiều rộng nhỏ\n- md: ảnh vừa\n- lg: ảnh chiều rộng lớn nhất\n- data-image-width-preset='sm/md/lg'",
     "Functional", "P3"),

    (73, "10. Hình ảnh", "TC-073",
     "Validation - file quá lớn",
     "Kiểm tra từ chối file vượt maxImageFileSize.",
     "maxImageFileSize=1MB (1048576 bytes).",
     "1. Click dropdown Image > Tải lên\n2. Chọn file ảnh > 1MB",
     "- File bị từ chối\n- Thông báo lỗi hiển thị\n- Không có ảnh được chèn",
     "Functional", "P2"),

    (74, "10. Hình ảnh", "TC-074",
     "Validation - MIME type không hỗ trợ",
     "Kiểm tra từ chối file không phải ảnh.",
     "allowedImageMimeTypes=['image/jpeg','image/png'].",
     "1. Thử paste hoặc upload file .gif, .webp\n2. Upload file .pdf",
     "- File không nằm trong allowedMimeTypes bị từ chối\n- PDF bị từ chối vì không phải ảnh\n- Thông báo lỗi rõ ràng",
     "Functional", "P2"),

    (75, "10. Hình ảnh", "TC-075",
     "Xóa ảnh",
     "Kiểm tra xóa ảnh khỏi editor.",
     "Ảnh đang được chọn.",
     "1. Click vào ảnh để chọn\n2. Click dropdown Image\n3. Chọn 'Xóa ảnh'",
     "- Ảnh bị xóa khỏi editor\n- Không còn thẻ img trong DOM\n- Cursor đặt tại vị trí ảnh vừa bị xóa",
     "Functional", "P2"),

    # ─── MODULE 11: BẢNG ──────────────────────────────────────────────────────
    (76, "11. Bảng (Table)", "TC-076",
     "Tạo bảng qua TableInsertGrid",
     "Kiểm tra tạo bảng bằng grid picker.",
     "Editor focus.",
     "1. Click dropdown Table trên toolbar\n2. Trong grid picker, di chuột đến ô 3x4\n3. Click để tạo",
     "- Bảng 3 hàng x 4 cột được tạo\n- Hàng đầu là header (th)\n- Các ô có thể nhập nội dung",
     "Functional", "P1"),

    (77, "11. Bảng (Table)", "TC-077",
     "Tạo bảng qua Slash Command",
     "Kiểm tra tạo bảng 3x3 qua slash command.",
     "Editor focus, dòng trống.",
     "1. Gõ '/table'\n2. Chọn 'Table' từ menu",
     "- Bảng 3x3 với header row được tạo\n- withHeaderRow=true",
     "Functional", "P1"),

    (78, "11. Bảng (Table)", "TC-078",
     "Thêm cột trước/sau",
     "Kiểm tra insert column.",
     "Bảng đã có, cursor trong ô.",
     "1. Click vào ô trong cột\n2. Mở dropdown Table > 'Thêm cột trước'\n3. 'Thêm cột sau'",
     "- Cột mới được thêm vào vị trí tương ứng\n- Cột mới rỗng\n- Số cột tăng thêm 1",
     "Functional", "P1"),

    (79, "11. Bảng (Table)", "TC-079",
     "Thêm hàng trước/sau",
     "Kiểm tra insert row.",
     "Bảng đã có, cursor trong ô.",
     "1. Mở Table menu\n2. Chọn 'Thêm hàng trước'\n3. Chọn 'Thêm hàng sau'",
     "- Hàng mới được thêm\n- Hàng mới rỗng\n- Số hàng tăng thêm 1",
     "Functional", "P1"),

    (80, "11. Bảng (Table)", "TC-080",
     "Xóa cột",
     "Kiểm tra xóa column.",
     "Bảng có ít nhất 2 cột, cursor trong ô.",
     "1. Mở Table menu\n2. Click 'Xóa cột'",
     "- Cột hiện tại bị xóa\n- Dữ liệu trong cột mất\n- Cột khác không bị ảnh hưởng",
     "Functional", "P1"),

    (81, "11. Bảng (Table)", "TC-081",
     "Xóa hàng",
     "Kiểm tra xóa row.",
     "Bảng có ít nhất 2 hàng (không kể header).",
     "1. Mở Table menu\n2. Click 'Xóa hàng'",
     "- Hàng hiện tại bị xóa\n- Nội dung hàng đó mất\n- Hàng header không thể xóa bằng cách này nếu là hàng duy nhất",
     "Functional", "P1"),

    (82, "11. Bảng (Table)", "TC-082",
     "Xóa toàn bộ bảng",
     "Kiểm tra delete table.",
     "Cursor trong bảng.",
     "1. Mở Table menu\n2. Click 'Xóa bảng'",
     "- Toàn bộ bảng bị xóa\n- Không còn thẻ <table> trong DOM\n- Cursor đặt tại paragraph trước/sau bảng",
     "Functional", "P1"),

    (83, "11. Bảng (Table)", "TC-083",
     "Toggle header row",
     "Kiểm tra bật/tắt header row.",
     "Bảng đã tạo với header row.",
     "1. Mở Table menu\n2. Click 'Toggle Header Row'\n3. Click lại để toggle",
     "- Lần 1: header row chuyển thành <td> bình thường\n- Lần 2: khôi phục lại <th>",
     "Functional", "P2"),

    (84, "11. Bảng (Table)", "TC-084",
     "Merge cells (gộp ô)",
     "Kiểm tra merge nhiều ô thành 1.",
     "Bảng đã có data, editor focus.",
     "1. Click ô đầu tiên\n2. Kéo chọn nhiều ô liên tiếp\n3. Click nút Merge (⊞) trên toolbar context\n4. Hoặc Mở Table menu > 'Gộp ô'",
     "- Các ô được gộp thành 1\n- colspan/rowspan được áp dụng đúng\n- Nội dung các ô được gộp lại",
     "Functional", "P1"),

    (85, "11. Bảng (Table)", "TC-085",
     "Split cell (tách ô đã gộp)",
     "Kiểm tra tách ô colspan/rowspan.",
     "Có ô đã merge (colspan > 1 hoặc rowspan > 1).",
     "1. Click vào ô đã merge\n2. Click nút Split (⊟) hoặc 'Tách ô'",
     "- Ô được tách thành nhiều ô riêng lẻ\n- Kích thước bảng khôi phục đúng\n- Merge button chuyển thành Split khi active",
     "Functional", "P1"),

    (86, "11. Bảng (Table)", "TC-086",
     "Căn lề bảng (Table Alignment)",
     "Kiểm tra căn lề toàn bộ bảng.",
     "Bảng đã tạo, cursor trong bảng.",
     "1. Mở Table menu\n2. Chọn 'Căn giữa'\n3. Chọn 'Căn phải'",
     "- Bảng di chuyển theo căn lề được chọn\n- margin: auto cho căn giữa\n- data-table-align attribute được cập nhật",
     "Functional", "P2"),

    (87, "11. Bảng (Table)", "TC-087",
     "Màu nền cell (Cell Background Color)",
     "Kiểm tra đổi màu nền ô bảng.",
     "Cursor đang trong ô bảng, bubble menu hiển thị.",
     "1. Click vào ô bảng\n2. Trong bubble menu chọn icon màu nền ô\n3. Chọn màu",
     "- Ô bảng có màu nền được chọn\n- data-background-color được lưu\n- style='background-color:...' được render",
     "Functional", "P2"),

    (88, "11. Bảng (Table)", "TC-088",
     "Resize cột bảng bằng kéo thả",
     "Kiểm tra thay đổi độ rộng cột.",
     "Bảng đã tạo, editable=true.",
     "1. Hover chuột vào border giữa 2 cột\n2. Cursor chuyển thành ↔\n3. Kéo để thay đổi độ rộng",
     "- Cột được resize\n- Column width được lưu\n- Không ảnh hưởng đến cột khác",
     "Functional", "P2"),

    (89, "11. Bảng (Table)", "TC-089",
     "Paste bảng từ Excel/Google Sheets",
     "Kiểm tra paste TSV/HTML từ spreadsheet.",
     "Clipboard có data từ Excel (TSV hoặc HTML table).",
     "1. Copy vùng data từ Excel/Sheets\n2. Paste vào editor (Ctrl+V)",
     "- Data được chuyển thành bảng HTML trong editor\n- Cấu trúc hàng/cột được giữ nguyên\n- Style cơ bản (background color nếu có) được giữ",
     "Functional", "P2"),

    (90, "11. Bảng (Table)", "TC-090",
     "Toolbar context bảng - quick action buttons",
     "Kiểm tra các nút quick action xuất hiện khi trong bảng.",
     "Cursor đang ở trong bảng.",
     "1. Click vào ô bảng\n2. Quan sát toolbar",
     "- Toolbar tự động hiện thêm nhóm nút: ←↑↓→ (add row/col) và nút merge/split\n- Nút bị disabled đúng khi không thể thực hiện",
     "UI", "P2"),

    # ─── MODULE 12: SLASH COMMAND ─────────────────────────────────────────────
    (91, "12. Slash Command", "TC-091",
     "Gõ '/' mở slash command menu",
     "Kiểm tra trigger slash command.",
     "Editor focus, dòng trống hoặc đầu dòng.",
     "1. Gõ '/' trong editor",
     "- Popup menu xuất hiện phía dưới cursor\n- Hiển thị danh sách các block type\n- Section 'Basic Blocks' hiển thị",
     "Functional", "P1"),

    (92, "12. Slash Command", "TC-092",
     "Tìm kiếm trong slash command",
     "Kiểm tra filter theo tên.",
     "Slash command menu đang mở.",
     "1. Gõ '/head'\n2. Quan sát danh sách\n3. Gõ '/xyz' (không tìm thấy)",
     "- '/head': chỉ hiện Heading 1/2/3\n- '/xyz': hiển thị 'No results'\n- Tìm kiếm case-insensitive",
     "Functional", "P1"),

    (93, "12. Slash Command", "TC-093",
     "Điều hướng slash command bằng phím mũi tên",
     "Kiểm tra keyboard navigation.",
     "Slash command menu đang mở.",
     "1. Nhấn phím mũi tên xuống ↓\n2. Nhấn ↑\n3. Quan sát item được highlight",
     "- Item tiếp theo/trước được highlight\n- Cuộn tự động khi item ngoài viewport\n- Vòng tròn: từ cuối quay lại đầu",
     "Functional", "P2"),

    (94, "12. Slash Command", "TC-094",
     "Chọn block bằng Enter",
     "Kiểm tra thực thi slash command bằng Enter.",
     "Slash command menu mở, item được highlight.",
     "1. Dùng ↓ để highlight 'Bullet List'\n2. Nhấn Enter",
     "- '/' và text tìm kiếm bị xóa\n- Bullet list block được insert\n- Menu đóng lại",
     "Functional", "P1"),

    (95, "12. Slash Command", "TC-095",
     "Đóng menu bằng Escape",
     "Kiểm tra Escape đóng slash command.",
     "Slash command menu đang mở.",
     "1. Nhấn Escape",
     "- Menu đóng lại\n- '/' và text tìm kiếm vẫn còn trong editor\n- Có thể tiếp tục gõ bình thường",
     "Functional", "P2"),

    (96, "12. Slash Command", "TC-096",
     "Insert Callout qua Slash Command",
     "Kiểm tra insert Callout block.",
     "Editor focus.",
     "1. Gõ '/callout'\n2. Chọn 'Callout'",
     "- Callout block được insert\n- Có emoji và màu nền mặc định\n- Có thể nhập text vào trong",
     "Functional", "P2"),

    (97, "12. Slash Command", "TC-097",
     "Insert Bookmark qua Slash Command",
     "Kiểm tra insert Bookmark.",
     "Editor focus.",
     "1. Gõ '/bookmark'\n2. Chọn 'Bookmark Card'\n3. Nhập URL vào prompt",
     "- Bookmark card được tạo với URL đã nhập\n- Sau khi fetch: hiển thị title, description, image\n- Click vào card mở URL",
     "Functional", "P2"),

    (98, "12. Slash Command", "TC-098",
     "Insert File Attachment qua Slash Command",
     "Kiểm tra insert FileCard.",
     "uploadFile callback được cấu hình.",
     "1. Gõ '/file'\n2. Chọn 'File Attachment'\n3. Chọn file từ file picker",
     "- File card được hiển thị với tên file, kích thước\n- Icon theo loại file\n- Upload qua uploadFile callback",
     "Functional", "P2"),

    # ─── MODULE 13: EMOJI ─────────────────────────────────────────────────────
    (99, "13. Emoji", "TC-099",
     "Gõ ':' mở emoji suggestion",
     "Kiểm tra trigger emoji suggestion.",
     "Editor focus.",
     "1. Gõ ':smile'\n2. Quan sát popup",
     "- Popup emoji xuất hiện\n- Các emoji có tên chứa 'smile' được hiển thị\n- Lưới 8x8, tối đa 64 emoji",
     "Functional", "P2"),

    (100, "13. Emoji", "TC-100",
     "Chèn emoji từ toolbar Emoji Picker",
     "Kiểm tra emoji picker đầy đủ trong toolbar.",
     "Editor focus.",
     "1. Click nút Emoji (😊) trên toolbar\n2. Tìm kiếm emoji\n3. Click emoji muốn chèn",
     "- Emoji được chèn tại vị trí cursor\n- Picker đóng sau khi chọn",
     "Functional", "P2"),

    (101, "13. Emoji", "TC-101",
     "Chọn emoji bằng phím Enter",
     "Kiểm tra keyboard selection emoji.",
     "Emoji suggestion popup đang mở.",
     "1. Gõ ':fire'\n2. Dùng mũi tên điều hướng\n3. Nhấn Enter",
     "- Emoji được insert vào editor\n- ':fire' text bị xóa thay bằng 🔥",
     "Functional", "P3"),

    # ─── MODULE 14: CUSTOM BLOCKS ─────────────────────────────────────────────
    (102, "14. Custom Blocks", "TC-102",
     "Callout - tạo và đổi màu",
     "Kiểm tra Callout block đầy đủ tính năng.",
     "Editor focus.",
     "1. Tạo Callout qua slash command\n2. Click vào emoji của callout\n3. Đổi sang màu xanh lá\n4. Nhập nội dung",
     "- Callout có 5 màu: Gray, Blue, Green, Yellow, Red\n- Màu nền thay đổi ngay lập tức\n- Emoji có thể được chỉnh sửa\n- Nội dung có thể nhập bình thường",
     "Functional", "P2"),

    (103, "14. Custom Blocks", "TC-103",
     "Bookmark - hiển thị OpenGraph metadata",
     "Kiểm tra Bookmark fetch và hiển thị metadata.",
     "fetchMetadata callback cấu hình, trả về title/description/image.",
     "1. Tạo Bookmark với URL có OpenGraph\n2. Chờ metadata load",
     "- Title hiển thị\n- Description hiển thị\n- Thumbnail image hiển thị\n- Publisher/domain hiển thị\n- Click vào card mở URL trong tab mới",
     "Functional", "P2"),

    (104, "14. Custom Blocks", "TC-104",
     "Bookmark - fallback khi không fetch được",
     "Kiểm tra fallback UI của Bookmark.",
     "fetchMetadata trả về lỗi hoặc không được cấu hình.",
     "1. Tạo Bookmark với URL bất kỳ\n2. fetchMetadata fail hoặc không được truyền",
     "- Hiển thị URL thô\n- Không có ảnh thumbnail\n- Không crash\n- Vẫn click được",
     "Functional", "P2"),

    (105, "14. Custom Blocks", "TC-105",
     "File Card - hiển thị đầy đủ thông tin",
     "Kiểm tra File Card UI.",
     "File card đã được tạo với file PDF.",
     "1. Chọn file PDF qua slash command\n2. Quan sát card sau khi upload",
     "- Icon PDF đúng với loại file\n- Tên file hiển thị\n- Kích thước file được format đúng (KB/MB)\n- Nút download hiển thị\n- URL upload thành công được lưu",
     "Functional", "P2"),

    # ─── MODULE 15: BUBBLE MENU ───────────────────────────────────────────────
    (106, "15. Bubble Menu", "TC-106",
     "Bubble menu hiện khi bôi chọn text",
     "Kiểm tra bubble menu trigger.",
     "showBubbleMenu=true (default), editor có text.",
     "1. Nhập text\n2. Bôi chọn một phần text\n3. Quan sát",
     "- Bubble menu xuất hiện phía trên vùng chọn\n- Có các nút format: Bold, Italic, Underline...\n- Vị trí không bị che khuất",
     "UI", "P1"),

    (107, "15. Bubble Menu", "TC-107",
     "Bubble menu ẩn khi bỏ chọn",
     "Kiểm tra bubble menu tự ẩn.",
     "Bubble menu đang hiển thị.",
     "1. Click vào vùng khác\n2. Quan sát bubble menu",
     "- Bubble menu biến mất\n- Không có animation bug",
     "UI", "P2"),

    (108, "15. Bubble Menu", "TC-108",
     "Bubble menu - Image controls",
     "Kiểm tra controls khi chọn ảnh.",
     "Editor có ảnh.",
     "1. Click vào ảnh\n2. Quan sát bubble menu",
     "- Bubble menu hiện controls riêng cho ảnh\n- Có layout switcher (block/left/right)\n- Có width presets\n- Có nút xóa ảnh",
     "UI", "P2"),

    (109, "15. Bubble Menu", "TC-109",
     "Bubble menu - Table cell controls",
     "Kiểm tra cell-specific controls.",
     "Cursor trong ô bảng.",
     "1. Click vào ô bảng\n2. Quan sát bubble menu",
     "- Hiển thị controls màu nền ô\n- Hiển thị controls màu border\n- Merge/split cell buttons",
     "UI", "P2"),

    # ─── MODULE 16: UNDO/REDO ─────────────────────────────────────────────────
    (110, "16. Undo / Redo", "TC-110",
     "Undo thao tác vừa thực hiện",
     "Kiểm tra Undo hoàn tác thao tác.",
     "Editor có nội dung.",
     "1. Nhập text 'ABC'\n2. Click nút Undo trên toolbar",
     "- Text 'ABC' biến mất\n- Nút Undo disabled nếu không còn history\n- Nút Redo trở nên active",
     "Functional", "P1"),

    (111, "16. Undo / Redo", "TC-111",
     "Redo sau khi Undo",
     "Kiểm tra Redo khôi phục thao tác.",
     "Đã Undo ít nhất 1 lần.",
     "1. Sau khi Undo\n2. Click nút Redo",
     "- Thao tác được khôi phục\n- Nút Redo disabled khi hết history\n- Ctrl+Y và Ctrl+Shift+Z đều hoạt động",
     "Functional", "P1"),

    (112, "16. Undo / Redo", "TC-112",
     "Phím tắt Ctrl+Z / Ctrl+Y",
     "Kiểm tra keyboard shortcut Undo/Redo.",
     "Editor focus.",
     "1. Nhập text\n2. Nhấn Ctrl+Z nhiều lần\n3. Nhấn Ctrl+Y nhiều lần",
     "- Ctrl+Z: Undo từng bước\n- Ctrl+Y: Redo từng bước\n- Không crash khi nhấn nhiều lần",
     "Functional", "P1"),

    # ─── MODULE 17: CHARACTER COUNT ───────────────────────────────────────────
    (113, "17. Đếm ký tự", "TC-113",
     "Hiển thị số ký tự hiện tại",
     "Kiểm tra character count.",
     "showCharacterCount=true (default).",
     "1. Nhập 50 ký tự vào editor\n2. Quan sát footer",
     "- Số ký tự hiển thị: 50\n- Cập nhật realtime khi gõ\n- Hiển thị số từ",
     "Functional", "P2"),

    (114, "17. Đếm ký tự", "TC-114",
     "Limit ký tự với maxCharacters",
     "Kiểm tra giới hạn ký tự.",
     "maxCharacters=100.",
     "1. Mount <UEditor maxCharacters={100} />\n2. Nhập 90 ký tự\n3. Nhập thêm đến 96 ký tự\n4. Tiếp tục nhập",
     "- 90 ký tự: hiển thị 90/100, màu bình thường\n- 96 ký tự: màu warning (đỏ/cam), vì > 90%\n- Không thể nhập quá 100 ký tự",
     "Functional", "P1"),

    # ─── MODULE 18: CALLBACKS ─────────────────────────────────────────────────
    (115, "18. Callbacks & Props", "TC-115",
     "onChange callback khi nội dung thay đổi",
     "Kiểm tra onChange được gọi.",
     "onChange handler được truyền vào.",
     "1. Mount với onChange={handler}\n2. Nhập text vào editor",
     "- onChange được gọi mỗi khi nội dung thay đổi\n- Tham số trả về là HTML string\n- Không gọi onChange khi content không đổi",
     "Functional", "P1"),

    (116, "18. Callbacks & Props", "TC-116",
     "onHtmlChange và onJsonChange callbacks",
     "Kiểm tra các callback phụ.",
     "onHtmlChange và onJsonChange được cấu hình.",
     "1. Nhập text\n2. Quan sát giá trị callback",
     "- onHtmlChange: nhận HTML string\n- onJsonChange: nhận JSON object (Tiptap document)\n- Cả hai đồng bộ với onChange",
     "Functional", "P2"),

    (117, "18. Callbacks & Props", "TC-117",
     "Cập nhật content từ bên ngoài",
     "Kiểm tra controlled content update.",
     "Component mounted, content đang là 'Hello'.",
     "1. Thay đổi prop content thành 'Goodbye' từ parent\n2. Quan sát editor",
     "- Editor hiển thị nội dung mới 'Goodbye'\n- Không trigger onChange\n- Không có vòng lặp vô hạn\n- lastAppliedContentRef đảm bảo không update thừa",
     "Functional", "P1"),

    # ─── MODULE 19: PREPARE CONTENT FOR SAVE ──────────────────────────────────
    (118, "19. prepareContentForSave", "TC-118",
     "Upload ảnh base64 khi save",
     "Kiểm tra pipeline upload ảnh lúc save.",
     "uploadImageForSave callback được cấu hình.",
     "1. Paste ảnh (base64) vào editor\n2. Gọi ref.current.prepareContentForSave()\n3. Quan sát kết quả",
     "- uploadImageForSave được gọi với File object\n- HTML output: src thay bằng URL từ server\n- result.uploaded: danh sách file đã upload\n- result.errors: rỗng nếu thành công",
     "Functional", "P1"),

    (119, "19. prepareContentForSave", "TC-119",
     "Upload song song với concurrency",
     "Kiểm tra upload concurrency limit.",
     "Editor có 5 ảnh base64, uploadImageConcurrency=3.",
     "1. Paste 5 ảnh vào editor\n2. Gọi prepareContentForSave()\n3. Monitor network requests",
     "- Chỉ 3 request upload xảy ra đồng thời\n- Sau khi 1 complete, request thứ 4 bắt đầu\n- Tất cả 5 ảnh được upload cuối cùng",
     "Functional", "P2"),

    (120, "19. prepareContentForSave", "TC-120",
     "Xử lý lỗi upload riêng lẻ",
     "Kiểm tra partial error handling.",
     "uploadImageForSave sẽ fail cho ảnh thứ 2.",
     "1. Paste 3 ảnh vào editor\n2. Mock uploadImageForSave fail cho ảnh 2\n3. Gọi prepareContentForSave()",
     "- result.errors có 1 entry với index của ảnh lỗi\n- 2 ảnh còn lại vẫn được upload thành công\n- HTML: ảnh lỗi giữ nguyên base64\n- Không throw nếu throwOnError=false",
     "Functional", "P1"),

    (121, "19. prepareContentForSave", "TC-121",
     "throwOnError=true khi có lỗi",
     "Kiểm tra throw exception mode.",
     "uploadImageForSave fail cho ít nhất 1 ảnh.",
     "1. Có lỗi upload\n2. Gọi ref.prepareContentForSave({throwOnError:true})",
     "- Hàm throw UEditorPrepareContentForSaveError\n- Error object chứa result đầy đủ\n- Có thể catch và xử lý\n- Với throwOnError=false (default): không throw",
     "Functional", "P1"),

    (122, "19. prepareContentForSave", "TC-122",
     "Không upload ảnh đã có URL (không base64)",
     "Kiểm tra skip ảnh đã có URL.",
     "Editor có ảnh từ URL (không phải base64).",
     "1. Chèn ảnh qua URL https://example.com/img.jpg\n2. Gọi prepareContentForSave()",
     "- uploadImageForSave không được gọi\n- URL ảnh giữ nguyên trong HTML output\n- result.inlineImageUrls chứa URL ảnh",
     "Functional", "P2"),

    (123, "19. prepareContentForSave", "TC-123",
     "Dedup in-flight concurrent calls",
     "Kiểm tra gọi prepareContentForSave đồng thời.",
     "Editor có ảnh base64.",
     "1. Gọi prepareContentForSave() lần 1\n2. Ngay lập tức gọi lần 2 (chưa đợi lần 1 xong)",
     "- Cả 2 calls share cùng 1 Promise\n- uploadImageForSave chỉ được gọi 1 lần (không duplicate)\n- Cả 2 resolve với cùng kết quả",
     "Functional", "P2"),

    # ─── MODULE 20: MENU BAR ──────────────────────────────────────────────────
    (124, "20. Menu Bar", "TC-124",
     "File > Save - gọi onSave callback",
     "Kiểm tra menu bar File > Save.",
     "showMenuBar=true, onSave callback được cấu hình.",
     "1. Click menu 'File'\n2. Click 'Lưu'",
     "- onSave callback được gọi\n- Phím tắt Ctrl+S cũng gọi onSave",
     "Functional", "P2"),

    (125, "20. Menu Bar", "TC-125",
     "View > Source Code",
     "Kiểm tra xem HTML source.",
     "showMenuBar=true, có nội dung trong editor.",
     "1. Click menu 'View'\n2. Click 'Source Code'\n3. Quan sát dialog/panel",
     "- Dialog mở với HTML source của nội dung\n- Có thể chỉnh sửa trực tiếp HTML\n- Sau khi lưu: editor cập nhật với HTML mới",
     "Functional", "P2"),

    (126, "20. Menu Bar", "TC-126",
     "View > Preview",
     "Kiểm tra chế độ preview.",
     "showMenuBar=true, editor có nội dung.",
     "1. Click menu 'View'\n2. Click 'Preview'\n3. Quan sát output",
     "- Modal hoặc panel hiển thị nội dung dạng read-only\n- Formatting được render đúng\n- Có thể đóng preview",
     "Functional", "P2"),

    # ─── MODULE 21: URL SAFETY ────────────────────────────────────────────────
    (127, "21. Bảo mật URL", "TC-127",
     "Chặn URL javascript: protocol (XSS)",
     "Kiểm tra isSafeUEditorUrl từ chối javascript:.",
     "Cố tình chèn URL độc hại.",
     "1. Thử thêm link với href='javascript:alert(1)'\n2. Thử chèn ảnh với src='javascript:alert(1)'",
     "- URL bị reject, không được áp dụng\n- Không có alert hoặc code execution\n- Không có <a href='javascript:...'> trong DOM",
     "Functional", "P1"),

    (128, "21. Bảo mật URL", "TC-128",
     "Chặn protocol-relative URL",
     "Kiểm tra từ chối //evil.com.",
     "Cố tình chèn protocol-relative URL.",
     "1. Thêm link với URL '//evil.com'\n2. Quan sát",
     "- URL bị sanitize hoặc reject\n- Không có link đến //evil.com",
     "Functional", "P2"),

    (129, "21. Bảo mật URL", "TC-129",
     "Cho phép https:// và mailto: và tel:",
     "Kiểm tra whitelist protocols.",
     "Editor focus.",
     "1. Thêm link 'https://example.com'\n2. Thêm link 'mailto:test@example.com'\n3. Thêm link 'tel:+84123456789'",
     "- Tất cả 3 URL được chấp nhận\n- Link được tạo thành công trong editor",
     "Functional", "P1"),

    # ─── MODULE 22: I18N ──────────────────────────────────────────────────────
    (130, "22. Đa ngôn ngữ (i18n)", "TC-130",
     "Toolbar hiển thị đúng tiếng Việt",
     "Kiểm tra i18n locale vi.",
     "App đang dùng locale 'vi'.",
     "1. Hover vào nút Bold\n2. Hover vào các nút khác trên toolbar\n3. Quan sát tooltip",
     "- Tooltip 'In đậm' (không phải 'Bold')\n- Tất cả tooltip tiếng Việt\n- Placeholder text tiếng Việt",
     "UI", "P2"),

    (131, "22. Đa ngôn ngữ (i18n)", "TC-131",
     "Slash command hiển thị tiếng Việt",
     "Kiểm tra slash command labels.",
     "Locale 'vi'.",
     "1. Gõ '/'\n2. Quan sát các item trong menu",
     "- Tên và mô tả các block hiển thị bằng tiếng Việt\n- 'Heading 1' → 'Tiêu đề 1'\n- 'No results' → 'Không có kết quả'",
     "UI", "P2"),

    # ─── MODULE 23: ACCESSIBILITY ─────────────────────────────────────────────
    (132, "23. Accessibility (A11Y)", "TC-132",
     "Toolbar buttons có aria-label",
     "Kiểm tra accessibility labels.",
     "Editor mounted.",
     "1. Inspect DOM của toolbar\n2. Kiểm tra attribute aria-label\n3. Dùng screen reader",
     "- Mỗi toolbar button có aria-label mô tả chức năng\n- aria-label đúng ngôn ngữ hiện tại\n- Screen reader đọc được",
     "UI", "P3"),

    (133, "23. Accessibility (A11Y)", "TC-133",
     "Tab navigation trong editor",
     "Kiểm tra keyboard-only navigation.",
     "Editor mounted.",
     "1. Dùng Tab để navigate qua toolbar\n2. Nhấn Enter/Space để activate button",
     "- Tab focus đi qua các nút toolbar theo thứ tự\n- Enter/Space kích hoạt button\n- Focus ring visible",
     "UI", "P3"),

    # ─── MODULE 24: EDGE CASES ────────────────────────────────────────────────
    (134, "24. Edge Cases", "TC-134",
     "Editor với content HTML phức tạp",
     "Kiểm tra parser với HTML nested phức tạp.",
     "Có chuỗi HTML phức tạp với nested elements.",
     "1. Mount với content chứa: headings, lists, tables, images, code blocks\n2. Quan sát rendering",
     "- Tất cả elements được parse đúng\n- Không mất nội dung\n- Không có error console",
     "Functional", "P1"),

    (135, "24. Edge Cases", "TC-135",
     "Paste text thuần từ clipboard",
     "Kiểm tra paste plain text.",
     "Editor focus.",
     "1. Copy text thuần từ Notepad\n2. Paste vào editor",
     "- Text được paste vào\n- Không có HTML injection\n- Không có format không mong muốn",
     "Functional", "P2"),

    (136, "24. Edge Cases", "TC-136",
     "Nhập ký tự đặc biệt và Unicode",
     "Kiểm tra support ký tự đặc biệt.",
     "Editor focus.",
     "1. Nhập: © ® ™ → ← ↑ ↓ ∑ ∫\n2. Nhập tiếng Việt có dấu\n3. Nhập emoji trực tiếp 🎉🔥",
     "- Tất cả ký tự hiển thị đúng\n- Không bị mất hoặc replaced\n- HTML output encode đúng",
     "Functional", "P2"),

    (137, "24. Edge Cases", "TC-137",
     "Hiệu năng với content rất dài",
     "Kiểm tra performance khi có nhiều nội dung.",
     "Editor được mount.",
     "1. Paste văn bản dài 10,000 từ vào editor\n2. Scroll\n3. Thực hiện format",
     "- Không có lag đáng kể khi nhập\n- Scroll mượt mà\n- Format áp dụng trong vòng < 500ms",
     "Functional", "P3"),

    (138, "24. Edge Cases", "TC-138",
     "Editor không re-render khi content prop không đổi",
     "Kiểm tra optimization - tránh vòng lặp.",
     "Editor mounted, onChange cập nhật state parent.",
     "1. Nhập text → onChange → parent state update → content prop không đổi (same HTML)\n2. Quan sát editor",
     "- Editor không reset cursor về đầu\n- Không có flicker hay re-render\n- lastAppliedContentRef ngăn update thừa",
     "Functional", "P1"),

    (139, "24. Edge Cases", "TC-139",
     "Destroy và remount editor",
     "Kiểm tra cleanup khi unmount.",
     "Editor đang hiển thị.",
     "1. Unmount component (navigate away)\n2. Remount lại\n3. Quan sát không có memory leak/error",
     "- Unmount không gây error\n- Remount tạo editor mới bình thường\n- Không có dangling event listeners",
     "Functional", "P2"),

    (140, "24. Edge Cases", "TC-140",
     "extraExtensions - thêm extension tùy chỉnh",
     "Kiểm tra inject custom Tiptap extension.",
     "Có custom Tiptap extension sẵn sàng.",
     "1. Mount <UEditor extraExtensions={[MyCustomExtension]} />\n2. Test tính năng của extension",
     "- Custom extension được load\n- Không conflict với extension có sẵn\n- Tính năng của extension hoạt động",
     "Functional", "P3"),
]

STATUS_OPTIONS = '"Chưa test,Pass,Fail,Blocked,Skip,Re-test,Fixed"'
PRIORITY_OPTIONS = '"P1,P2,P3"'
TYPE_OPTIONS = '"Functional,UI,Edge case"'

# ──────────────────────────────────────────────
# BUILD WORKBOOK
# ──────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: Test Cases ────────────────────────
ws = wb.active
ws.title = "Test Cases"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A3"

# ── Header row 1: title banner ────────────────
ws.merge_cells("A1:M1")
banner = ws["A1"]
banner.value = "UEDITOR COMPONENT - TEST CASE SPECIFICATION"
banner.font = Font(bold=True, size=14, color=C_HEADER_FG, name="Segoe UI")
banner.fill = fill(C_HEADER_BG)
banner.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# ── Header row 2: columns ─────────────────────
ws.row_dimensions[2].height = 36
for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=2, column=col_idx, value=col_name)
    cell.font = bold_font(10, C_HEADER_FG)
    cell.fill = fill(C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
    ws.column_dimensions[get_column_letter(col_idx)].width = col_width

# ── Data rows ─────────────────────────────────
module_fill_map = {}
module_fill_colors = [
    "1E40AF",  # blue-800
    "065F46",  # green-800
    "7C2D12",  # orange-800
    "4C1D95",  # purple-800
    "881337",  # rose-800
    "134E4A",  # teal-800
    "1C1917",  # stone-900
    "1E3A5F",  # dark blue
    "3F3F1E",  # dark yellow
    "1F2D3D",  # dark navy
    "2D1B69",  # deep purple
    "0F3460",  # royal blue
    "1A1A2E",  # midnight
    "16213E",  # dark ocean
    "0F3460",  # deep blue
    "533483",  # medium purple
    "155263",  # teal dark
    "E84545",  # red accent - lighter for variety
    "2B2D42",  # dark slate
    "3D5A80",  # muted blue
    "6B4226",  # brown
    "003049",  # dark teal
    "403D39",  # warm dark
    "023E8A",  # deep navy
]

def get_module_color(module_name):
    if module_name not in module_fill_map:
        idx = len(module_fill_map) % len(module_fill_colors)
        module_fill_map[module_name] = module_fill_colors[idx]
    return module_fill_map[module_name]

STATUS_COLOR = {
    "Chưa test": C_TODO,
    "Pass":       C_PASS,
    "Fail":       C_FAIL,
    "Fixed":      C_FIXED,
    "Re-test":    C_RETEST,
    "Blocked":    C_BLOCKED,
    "Skip":       C_SKIP,
}
PRIORITY_COLOR = {"P1": C_P1, "P2": C_P2, "P3": C_P3}

# Data validation
dv_status = DataValidation(type="list", formula1=STATUS_OPTIONS, allow_blank=True)
dv_priority = DataValidation(type="list", formula1=PRIORITY_OPTIONS, allow_blank=True)
dv_type = DataValidation(type="list", formula1=TYPE_OPTIONS, allow_blank=True)
ws.add_data_validation(dv_status)
ws.add_data_validation(dv_priority)
ws.add_data_validation(dv_type)

for row_offset, tc in enumerate(TEST_CASES):
    row = row_offset + 3
    (stt, module, ma_tc, ten_tc, mo_ta, dieu_kien, cac_buoc, ket_qua, loai_test, uu_tien) = tc

    is_alt = row_offset % 2 == 1
    row_bg = C_ALT_ROW if is_alt else C_WHITE
    ws.row_dimensions[row].height = 80

    mod_color = get_module_color(module)

    values = [
        stt,          # A: STT
        module,       # B: Module
        ma_tc,        # C: Mã TC
        ten_tc,       # D: Tên
        mo_ta,        # E: Mô tả
        dieu_kien,    # F: Điều kiện
        cac_buoc,     # G: Các bước
        ket_qua,      # H: Kết quả mong muốn
        "",           # I: Kết quả thực tế
        "Chưa test",  # J: Status
        uu_tien,      # K: Ưu tiên
        loai_test,    # L: Loại test
        "",           # M: Ghi chú
    ]

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border = thin_border()

        # Default style
        cell.font = normal_font(9)
        cell.alignment = wrap_align("left", "top")

        # Column-specific overrides
        if col_idx == 1:  # STT
            cell.font = bold_font(9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill(row_bg)

        elif col_idx == 2:  # Module
            cell.font = Font(bold=True, size=8, color="FFFFFF", name="Segoe UI")
            cell.fill = fill(mod_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        elif col_idx == 3:  # Mã TC
            cell.font = Font(bold=True, size=9, color="1D4ED8", name="Courier New")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill(row_bg)

        elif col_idx == 4:  # Tên
            cell.font = Font(bold=True, size=9, color="1E293B", name="Segoe UI")
            cell.fill = fill(row_bg)

        elif col_idx in (5, 6, 7, 8, 9, 13):  # Text content
            cell.fill = fill(row_bg)

        elif col_idx == 10:  # Status
            status_val = "Chưa test"
            cell.value = status_val
            cell.fill = fill(STATUS_COLOR.get(status_val, C_TODO))
            cell.font = bold_font(9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            dv_status.sqref = f"J{row}:J{row}"

        elif col_idx == 11:  # Priority
            cell.fill = fill(PRIORITY_COLOR.get(uu_tien, C_WHITE))
            cell.font = bold_font(9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            dv_priority.sqref = f"K{row}:K{row}"

        elif col_idx == 12:  # Loại test
            cell.fill = fill(row_bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            dv_type.sqref = f"L{row}:L{row}"

    # Apply validations to range
    dv_status.add(f"J{row}")
    dv_priority.add(f"K{row}")
    dv_type.add(f"L{row}")

# ── Sheet 2: Summary ──────────────────────────
ws2 = wb.create_sheet("Tổng hợp")
ws2.sheet_view.showGridLines = False

# Banner
ws2.merge_cells("A1:G1")
b2 = ws2["A1"]
b2.value = "TỔNG HỢP TEST CASE THEO MODULE"
b2.font = Font(bold=True, size=13, color=C_HEADER_FG, name="Segoe UI")
b2.fill = fill(C_HEADER_BG)
b2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

# Header
summary_cols = [
    ("Module", 32), ("Tổng TC", 10), ("Pass", 8),
    ("Fail", 8), ("Blocked", 9), ("Chưa test", 11), ("Ghi chú", 26),
]
for col_idx, (col_name, col_width) in enumerate(summary_cols, start=1):
    cell = ws2.cell(row=2, column=col_idx, value=col_name)
    cell.font = bold_font(10, C_HEADER_FG)
    cell.fill = fill(C_MODULE_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()
    ws2.column_dimensions[get_column_letter(col_idx)].width = col_width
ws2.row_dimensions[2].height = 22

# Count TCs per module
from collections import Counter
module_counts = Counter(tc[1] for tc in TEST_CASES)

for row_idx, (module_name, count) in enumerate(sorted(module_counts.items()), start=3):
    ws2.row_dimensions[row_idx].height = 20
    bg = C_ALT_ROW if row_idx % 2 == 0 else C_WHITE
    mod_c = get_module_color(module_name)
    row_data = [module_name, count, 0, 0, 0, count, ""]
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left",
                                   vertical="center", wrap_text=True)
        if col_idx == 1:
            cell.font = Font(bold=True, size=9, color="FFFFFF", name="Segoe UI")
            cell.fill = fill(mod_c)
        elif col_idx == 2:
            cell.font = bold_font(9)
            cell.fill = fill(bg)
        elif col_idx == 6:
            cell.font = Font(size=9, color="64748B", name="Segoe UI")
            cell.fill = fill(C_TODO)
        else:
            cell.font = normal_font(9)
            cell.fill = fill(bg)

# Total row
total_row = len(module_counts) + 3
ws2.row_dimensions[total_row].height = 22
total_data = ["TỔNG CỘNG", len(TEST_CASES), 0, 0, 0, len(TEST_CASES), ""]
for col_idx, val in enumerate(total_data, start=1):
    cell = ws2.cell(row=total_row, column=col_idx, value=val)
    cell.font = bold_font(10, C_HEADER_FG)
    cell.fill = fill(C_HEADER_BG)
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left",
                               vertical="center")

# ── Sheet 3: Legend ───────────────────────────
ws3 = wb.create_sheet("Hướng dẫn")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 18
ws3.column_dimensions["B"].width = 48

ws3.merge_cells("A1:B1")
leg = ws3["A1"]
leg.value = "HƯỚNG DẪN SỬ DỤNG FILE TEST CASE"
leg.font = Font(bold=True, size=12, color=C_HEADER_FG, name="Segoe UI")
leg.fill = fill(C_HEADER_BG)
leg.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 24

legend_data = [
    ("", ""),
    ("── STATUS ──", "Ý nghĩa"),
    ("Chưa test", "Test case chưa được thực thi"),
    ("Pass", "Test case thực thi thành công, đúng kết quả mong muốn"),
    ("Fail", "Test case thất bại, kết quả không đúng"),
    ("Blocked", "Không thể test do blocker (bug khác, môi trường...)"),
    ("Re-test", "Cần test lại sau khi fix bug"),
    ("Fixed", "Bug đã được fix, cần verify lại"),
    ("Skip", "Bỏ qua (out-of-scope, không áp dụng)"),
    ("", ""),
    ("── ƯU TIÊN ──", "Mô tả"),
    ("P1", "Critical - Chặn release nếu fail"),
    ("P2", "High - Quan trọng, cần fix trước release"),
    ("P3", "Medium/Low - Nice to have, fix sau release"),
    ("", ""),
    ("── LOẠI TEST ──", "Mô tả"),
    ("Functional", "Kiểm tra chức năng hoạt động đúng"),
    ("UI", "Kiểm tra giao diện, layout, visual"),
    ("Edge case", "Kiểm tra trường hợp đặc biệt, biên"),
    ("", ""),
    ("── QUY TRÌNH ──", ""),
    ("Bước 1", "Người test chọn TC cần thực thi"),
    ("Bước 2", "Thực hiện các bước trong cột 'Các bước thực hiện'"),
    ("Bước 3", "Ghi kết quả thực tế vào cột 'Kết quả thực tế'"),
    ("Bước 4", "So sánh với 'Kết quả mong muốn'"),
    ("Bước 5", "Cập nhật Status phù hợp"),
    ("Bước 6", "Nếu Fail: tạo bug ticket, note vào cột Ghi chú"),
]

status_row_colors = {
    "Chưa test": C_TODO, "Pass": C_PASS, "Fail": C_FAIL,
    "Blocked": C_BLOCKED, "Re-test": C_RETEST, "Fixed": C_FIXED, "Skip": C_SKIP,
}
priority_row_colors = {"P1": C_P1, "P2": C_P2, "P3": C_P3}

for ridx, (key, val) in enumerate(legend_data, start=2):
    ws3.row_dimensions[ridx].height = 18
    ca = ws3.cell(row=ridx, column=1, value=key)
    cb = ws3.cell(row=ridx, column=2, value=val)

    if key.startswith("──"):
        for c in [ca, cb]:
            c.font = bold_font(10, C_HEADER_FG)
            c.fill = fill(C_MODULE_BG)
            c.alignment = Alignment(vertical="center")
            c.border = thin_border()
    elif key in status_row_colors:
        ca.fill = fill(status_row_colors[key])
        ca.font = bold_font(9)
        cb.font = normal_font(9)
        for c in [ca, cb]: c.border = thin_border()
    elif key in priority_row_colors:
        ca.fill = fill(priority_row_colors[key])
        ca.font = bold_font(9)
        cb.font = normal_font(9)
        for c in [ca, cb]: c.border = thin_border()
    elif key.startswith("Bước"):
        ca.font = bold_font(9, "1D4ED8")
        cb.font = normal_font(9)
        for c in [ca, cb]:
            c.border = thin_border()
            c.fill = fill(C_ALT_ROW)
    elif key == "":
        pass
    else:
        ca.font = normal_font(9)
        cb.font = normal_font(9)
        for c in [ca, cb]:
            c.border = thin_border()
            c.fill = fill(C_WHITE)

# ── Save ──────────────────────────────────────
output_path = "/Users/tran_van_bach/Desktop/project/nextJs/underverse/docs/UEditor_TestCase.xlsx"
wb.save(output_path)
print(f"✓ File saved: {output_path}")
print(f"  Total test cases: {len(TEST_CASES)}")
print(f"  Modules: {len(module_counts)}")
print(f"  Sheets: Test Cases, Tổng hợp, Hướng dẫn")
